"""
Circuit IDE v8.0 - Enterprise AI Coding Assistant

Modern, professional IDE with AI-powered coding assistance.
Features:
- VS Code-style welcome screen
- Professional SVG icons (no emojis)
- Compact, enterprise-grade UI
- System-wide search with permissions
- Full Git integration
- AI context awareness with token tracking
"""

import sys
import os
import json
import asyncio
import subprocess
import re
import time
import logging
from pathlib import Path

# Configure logging for GUI
logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger("circuit_ide")
from typing import Optional, List, Dict, Any, Callable
from datetime import datetime
from enum import Enum
import threading

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QSplitter, QTreeView, QPlainTextEdit, QTextEdit, QLineEdit,
    QLabel, QStatusBar, QMenuBar, QMenu, QFileDialog,
    QMessageBox, QFrame, QPushButton, QComboBox, QFileSystemModel,
    QScrollArea, QStackedWidget, QListWidget, QListWidgetItem,
    QCheckBox, QSpinBox, QTabWidget, QToolButton, QDialog,
    QDialogButtonBox, QSizePolicy, QTextBrowser, QStyle, QGroupBox,
    QProgressBar, QFormLayout, QRadioButton, QButtonGroup,
    QTreeWidget, QTreeWidgetItem, QHeaderView, QGridLayout,
    QStyledItemDelegate, QTabBar
)
from PySide6.QtCore import (
    Qt, QDir, Signal, Slot, QThread, QObject, QSize, QTimer,
    QMargins, QPropertyAnimation, QEasingCurve, Property, QPoint, QPointF,
    QRect, QRectF, QProcess, QProcessEnvironment
)
from PySide6.QtGui import (
    QFont, QTextCharFormat, QColor, QSyntaxHighlighter,
    QAction, QKeySequence, QTextCursor, QIcon, QPainter,
    QPixmap, QPen, QBrush, QPalette, QLinearGradient, QFontDatabase,
    QPainterPath, QPolygonF, QShortcut
)

# Add parent to path for imports
_parent = Path(__file__).parent.parent
if str(_parent) not in sys.path:
    sys.path.insert(0, str(_parent))


# ============================================================================
# Theme System
# ============================================================================

THEMES = {
    "dark": {
        "name": "Dark",
        "BG_DARK": "#1e1e1e",
        "BG_MAIN": "#252526",
        "BG_SECONDARY": "#2d2d30",
        "BG_TERTIARY": "#3c3c3c",
        "BG_HOVER": "#454545",
        "BG_SELECTED": "#094771",
        "BG_INPUT": "#3c3c3c",
        "BORDER": "#3c3c3c",
        "BORDER_LIGHT": "#505050",
        "TEXT_PRIMARY": "#cccccc",
        "TEXT_SECONDARY": "#9d9d9d",
        "TEXT_MUTED": "#6d6d6d",
        "TEXT_LINK": "#3794ff",
        "ACCENT_BLUE": "#0078d4",
        "ACCENT_GREEN": "#4ec9b0",
        "ACCENT_PURPLE": "#c586c0",
        "ACCENT_ORANGE": "#ce9178",
        "ACCENT_RED": "#f14c4c",
        "ACCENT_CYAN": "#4fc1ff",
        "ACCENT_YELLOW": "#dcdcaa",
        "SUCCESS": "#4ec9b0",
        "WARNING": "#cca700",
        "ERROR": "#f14c4c",
        "INFO": "#3794ff",
        # Provider-specific colors
        "CLAUDE_COLOR": "#E8A87C",     # Light orange/coral for Claude
        "CIRCUIT_COLOR": "#88CFFF",    # Light blue for Circuit/Cisco
    },
    "light": {
        "name": "Light",
        "BG_DARK": "#f3f3f3",
        "BG_MAIN": "#ffffff",
        "BG_SECONDARY": "#f3f3f3",
        "BG_TERTIARY": "#e8e8e8",
        "BG_HOVER": "#e0e0e0",
        "BG_SELECTED": "#0078d4",
        "BG_INPUT": "#ffffff",
        "BORDER": "#e0e0e0",
        "BORDER_LIGHT": "#d4d4d4",
        "TEXT_PRIMARY": "#333333",
        "TEXT_SECONDARY": "#616161",
        "TEXT_MUTED": "#9e9e9e",
        "TEXT_LINK": "#0066bf",
        "ACCENT_BLUE": "#0066bf",
        "ACCENT_GREEN": "#16825d",
        "ACCENT_PURPLE": "#af00db",
        "ACCENT_ORANGE": "#c17e00",
        "ACCENT_RED": "#cd3131",
        "ACCENT_CYAN": "#0098c4",
        "ACCENT_YELLOW": "#795e26",
        "SUCCESS": "#16825d",
        "WARNING": "#bf8803",
        "ERROR": "#cd3131",
        "INFO": "#0066bf",
        # Provider-specific colors
        "CLAUDE_COLOR": "#D4845C",     # Darker orange for light theme
        "CIRCUIT_COLOR": "#4A90C2",    # Darker blue for light theme
    },
    "midnight": {
        "name": "Midnight Blue",
        "BG_DARK": "#0a0a12",
        "BG_MAIN": "#0e0e18",
        "BG_SECONDARY": "#14142a",
        "BG_TERTIARY": "#1e1e3c",
        "BG_HOVER": "#28284a",
        "BG_SELECTED": "#0050a0",
        "BG_INPUT": "#1e1e3c",
        "BORDER": "#2a2a4a",
        "BORDER_LIGHT": "#3a3a5a",
        "TEXT_PRIMARY": "#e0e0f0",
        "TEXT_SECONDARY": "#a0a0c0",
        "TEXT_MUTED": "#606080",
        "TEXT_LINK": "#6090ff",
        "ACCENT_BLUE": "#4080ff",
        "ACCENT_GREEN": "#40d090",
        "ACCENT_PURPLE": "#b060ff",
        "ACCENT_ORANGE": "#ff9060",
        "ACCENT_RED": "#ff5050",
        "ACCENT_CYAN": "#40d0ff",
        "ACCENT_YELLOW": "#ffd040",
        "SUCCESS": "#40d090",
        "WARNING": "#ffa040",
        "ERROR": "#ff5050",
        "INFO": "#4080ff",
        # Provider-specific colors
        "CLAUDE_COLOR": "#FFB080",     # Light orange for midnight
        "CIRCUIT_COLOR": "#99DDFF",    # Light blue for midnight
    },
    "forest": {
        "name": "Forest",
        "BG_DARK": "#0d120e",
        "BG_MAIN": "#111a14",
        "BG_SECONDARY": "#182218",
        "BG_TERTIARY": "#223020",
        "BG_HOVER": "#2a3a28",
        "BG_SELECTED": "#1a5030",
        "BG_INPUT": "#223020",
        "BORDER": "#2a3a28",
        "BORDER_LIGHT": "#3a4a38",
        "TEXT_PRIMARY": "#d0e0d0",
        "TEXT_SECONDARY": "#90a090",
        "TEXT_MUTED": "#607060",
        "TEXT_LINK": "#60c090",
        "ACCENT_BLUE": "#50a0d0",
        "ACCENT_GREEN": "#50c080",
        "ACCENT_PURPLE": "#a070c0",
        "ACCENT_ORANGE": "#d09050",
        "ACCENT_RED": "#d05050",
        "ACCENT_CYAN": "#50c0c0",
        "ACCENT_YELLOW": "#c0c050",
        "SUCCESS": "#50c080",
        "WARNING": "#c0a040",
        "ERROR": "#d05050",
        "INFO": "#50a0d0",
        # Provider-specific colors
        "CLAUDE_COLOR": "#E8A070",     # Light orange for forest
        "CIRCUIT_COLOR": "#88CFFF",    # Light blue for forest
    },
}


class ThemeManager:
    """Manages application themes with live switching."""

    _current_theme = "dark"
    _callbacks: List[Callable] = []

    @classmethod
    def get_theme(cls) -> Dict[str, str]:
        return THEMES.get(cls._current_theme, THEMES["dark"])

    @classmethod
    def get(cls, key: str) -> str:
        return cls.get_theme().get(key, "#000000")

    @classmethod
    def set_theme(cls, name: str):
        if name in THEMES:
            cls._current_theme = name
            for cb in cls._callbacks:
                try:
                    cb(cls.get_theme())
                except Exception as e:
                    logger.warning(f"Theme callback error: {e}")

    @classmethod
    def on_change(cls, callback: Callable):
        cls._callbacks.append(callback)

    @classmethod
    def list_themes(cls) -> List[str]:
        return list(THEMES.keys())


class Theme:
    """Theme accessor with static methods."""

    @staticmethod
    def get(key: str) -> str:
        return ThemeManager.get(key)


# Update Theme class attributes dynamically
def _update_theme_class():
    theme = ThemeManager.get_theme()
    for key, value in theme.items():
        if key != "name":
            setattr(Theme, key, value)


_update_theme_class()
ThemeManager.on_change(lambda t: _update_theme_class())


# ============================================================================
# Configuration
# ============================================================================

CONFIG_DIR = Path.home() / ".config" / "circuit-agent"
PROFILES_DIR = CONFIG_DIR / "profiles"
SEARCH_PERMISSIONS_FILE = CONFIG_DIR / "search_permissions.json"
RECENT_PROJECTS_FILE = CONFIG_DIR / "recent_projects.json"


class RecentProjects:
    """Manages recent project list."""

    MAX_RECENT = 10

    @classmethod
    def get(cls) -> List[str]:
        if RECENT_PROJECTS_FILE.exists():
            try:
                data = json.loads(RECENT_PROJECTS_FILE.read_text())
                return [p for p in data.get("projects", []) if Path(p).exists()][:cls.MAX_RECENT]
            except Exception as e:
                logger.warning(f"Failed to load recent projects: {e}")
        return []

    @classmethod
    def add(cls, path: str):
        projects = cls.get()
        path = str(Path(path).resolve())
        if path in projects:
            projects.remove(path)
        projects.insert(0, path)
        projects = projects[:cls.MAX_RECENT]
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        RECENT_PROJECTS_FILE.write_text(json.dumps({"projects": projects}, indent=2))


class SearchPermissions:
    """Manages search directory permissions."""

    def __init__(self):
        self.allowed_dirs: List[Path] = []
        self._load_permissions()

    def _load_permissions(self):
        if SEARCH_PERMISSIONS_FILE.exists():
            try:
                data = json.loads(SEARCH_PERMISSIONS_FILE.read_text())
                self.allowed_dirs = [Path(p) for p in data.get("allowed_directories", [])]
            except Exception as e:
                logger.warning(f"Failed to load search permissions: {e}")
                self.allowed_dirs = []

    def _save_permissions(self):
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        data = {
            "allowed_directories": [str(p) for p in self.allowed_dirs],
            "last_updated": datetime.now().isoformat()
        }
        SEARCH_PERMISSIONS_FILE.write_text(json.dumps(data, indent=2))

    def has_access(self, directory: Path) -> bool:
        directory = directory.resolve()
        for allowed in self.allowed_dirs:
            try:
                directory.relative_to(allowed.resolve())
                return True
            except ValueError:
                continue
        return False

    def grant_access(self, directory: Path):
        directory = directory.resolve()
        if directory not in self.allowed_dirs:
            self.allowed_dirs.append(directory)
            self._save_permissions()

    def revoke_access(self, directory: Path):
        directory = directory.resolve()
        self.allowed_dirs = [d for d in self.allowed_dirs if d.resolve() != directory]
        self._save_permissions()


class WorkspaceProfile:
    """Workspace configuration profiles."""

    @classmethod
    def save(cls, name: str, settings: Dict[str, Any]):
        PROFILES_DIR.mkdir(parents=True, exist_ok=True)
        filepath = PROFILES_DIR / f"{name}.json"
        filepath.write_text(json.dumps(settings, indent=2))

    @classmethod
    def load(cls, name: str) -> Optional[Dict[str, Any]]:
        filepath = PROFILES_DIR / f"{name}.json"
        if filepath.exists():
            return json.loads(filepath.read_text())
        return None

    @classmethod
    def list_profiles(cls) -> List[str]:
        if not PROFILES_DIR.exists():
            return ["Default"]
        profiles = [p.stem for p in PROFILES_DIR.glob("*.json")]
        return profiles if profiles else ["Default"]

    @classmethod
    def delete(cls, name: str):
        filepath = PROFILES_DIR / f"{name}.json"
        if filepath.exists():
            filepath.unlink()


# ============================================================================
# Icon System - Vector Icons (No Emojis)
# ============================================================================

class Icons:
    """VS Code-style icons using official SVG paths with caching for performance."""

    # Icon cache to avoid recreating the same icons
    _icon_cache: Dict[tuple, QIcon] = {}

    # Official VS Code icon SVG paths (from microsoft/vscode-icons)
    SVG_CLOSE = '<path fill-rule="evenodd" clip-rule="evenodd" d="M8.00004 8.70711L11.6465 12.3536L12.3536 11.6465L8.70714 8.00001L12.3536 4.35356L11.6465 3.64645L8.00004 7.2929L4.35359 3.64645L3.64648 4.35356L7.29293 8.00001L3.64648 11.6465L4.35359 12.3536L8.00004 8.70711Z" fill="{color}"/>'
    SVG_REFRESH = '<path fill-rule="evenodd" clip-rule="evenodd" d="M5.56253 2.51577C3.46348 3.4501 2 5.55414 2 7.99999C2 11.3137 4.68629 14 8 14C11.3137 14 14 11.3137 14 7.99999C14 5.32519 12.2497 3.05919 9.83199 2.28482L9.52968 3.23832C11.5429 3.88454 13 5.7721 13 7.99999C13 10.7614 10.7614 13 8 13C5.23858 13 3 10.7614 3 7.99999C3 6.31104 3.83742 4.81767 5.11969 3.91245L5.56253 2.51577Z" fill="{color}"/><path fill-rule="evenodd" clip-rule="evenodd" d="M5 3H2V2H5.5L6 2.5V6H5V3Z" fill="{color}"/>'
    SVG_FILES = '<path d="M2 2H6L7 3V5H10V3L11 2H14V13H2V2ZM3 3V12H13V3H11V6H7V3H3Z" fill="{color}"/>'
    SVG_SEARCH = '<path fill-rule="evenodd" clip-rule="evenodd" d="M10 4C10 6.20914 8.20914 8 6 8C3.79086 8 2 6.20914 2 4C2 1.79086 3.79086 0 6 0C8.20914 0 10 1.79086 10 4ZM9.17 8.58C8.23 9.15 7.14 9.5 6 9.5C2.96 9.5 0.5 7.04 0.5 4C0.5 0.96 2.96 -1.5 6 -1.5C9.04 -1.5 11.5 0.96 11.5 4C11.5 5.14 11.15 6.23 10.58 7.17L14.71 11.29L14 12L9.87 7.87L9.17 8.58Z" fill="{color}" transform="translate(0,2)"/>'
    SVG_SETTINGS = '<path d="M9.1 4.4L8.6 2H7.4L6.9 4.4L6.2 4.7L4.2 3.4L3.3 4.2L4.6 6.2L4.4 6.9L2 7.4V8.6L4.4 9.1L4.7 9.9L3.4 11.9L4.2 12.7L6.2 11.4L7 11.7L7.4 14H8.6L9.1 11.6L9.9 11.3L11.9 12.6L12.7 11.8L11.4 9.8L11.7 9L14 8.6V7.4L11.6 6.9L11.3 6.1L12.6 4.1L11.8 3.3L9.8 4.6L9.1 4.4ZM9.4 1L9.9 3.4L12 2.1L14 4.1L12.6 6.2L15 6.6V9.4L12.6 9.9L14 12L12 14L9.9 12.6L9.4 15H6.6L6.1 12.6L4 13.9L2 11.9L3.4 9.8L1 9.4V6.6L3.4 6.1L2.1 4L4.1 2L6.2 3.4L6.6 1H9.4ZM10 8C10 9.1 9.1 10 8 10C6.9 10 6 9.1 6 8C6 6.9 6.9 6 8 6C9.1 6 10 6.9 10 8ZM8 9C8.6 9 9 8.6 9 8C9 7.4 8.6 7 8 7C7.4 7 7 7.4 7 8C7 8.6 7.4 9 8 9Z" fill="{color}"/>'
    SVG_GIT = '<path d="M4 4C4 2.9 4.9 2 6 2C7.1 2 8 2.9 8 4C8 4.7 7.6 5.4 7 5.7V10.3C7.6 10.6 8 11.3 8 12C8 13.1 7.1 14 6 14C4.9 14 4 13.1 4 12C4 11.3 4.4 10.6 5 10.3V5.7C4.4 5.4 4 4.7 4 4ZM6 3C5.4 3 5 3.4 5 4C5 4.6 5.4 5 6 5C6.6 5 7 4.6 7 4C7 3.4 6.6 3 6 3ZM6 11C5.4 11 5 11.4 5 12C5 12.6 5.4 13 6 13C6.6 13 7 12.6 7 12C7 11.4 6.6 11 6 11ZM12 4C12 2.9 11.1 2 10 2C8.9 2 8 2.9 8 4H9C9 3.4 9.4 3 10 3C10.6 3 11 3.4 11 4C11 4.6 10.6 5 10 5H9V6H10C10.6 6 11 6.4 11 7L9 9H7V10H9.6L12 7.6V7C12 6.5 11.8 6.1 11.5 5.7C11.8 5.4 12 4.7 12 4Z" fill="{color}"/>'
    SVG_ROBOT = '<path d="M14.5 2H10V0H9V2H5.5L5 2.5V6.5L5.5 7H6V12H5.5L5 12.5V14.5L5.5 15H14.5L15 14.5V12.5L14.5 12H14V7H14.5L15 6.5V2.5L14.5 2ZM6 3H14V6H6V3ZM14 14H6V13H14V14ZM7 7H13V12H7V7ZM9 9V10H8V9H9ZM12 9V10H11V9H12Z" fill="{color}"/>'
    SVG_NEW_FILE = '<path fill-rule="evenodd" clip-rule="evenodd" d="M4 7H3V4H0V3H3V0H4V3H7V4H4V7ZM10.5 1.09998L13.9 4.59998L14 5V13.5L13.5 14H3.5L3 13.5V8H4V13H13V6H9V2H5V1H10.2L10.5 1.09998ZM10 2V5H12.9L10 2Z" fill="{color}"/>'
    SVG_NEW_FOLDER = '<path fill-rule="evenodd" clip-rule="evenodd" d="M7 3H4V0H3V3H0V4H3V7H4V4H7V3ZM5.5 7H5V6H5.3L6.1 5.1L6.5 5H14V4H8V3H14.5L15 3.5V13.5L14.5 14H1.5L1 13.5V6.5V6V5H2V6V6.5V13H14V7V6H6.7L5.9 6.9L5.5 7Z" fill="{color}"/>'
    SVG_COLLAPSE = '<path fill-rule="evenodd" clip-rule="evenodd" d="M7.97612 10.0719L12.3334 5.7146L12.9521 6.33332L8.28548 11L7.66676 11L3.0001 6.33332L3.61882 5.7146L7.97612 10.0719Z" fill="{color}" transform="rotate(180,8,8)"/>'
    SVG_ADD = '<path d="M14 7V8H8V14H7V8H1V7H7V1H8V7H14Z" fill="{color}"/>'
    SVG_CHEVRON_DOWN = '<path fill-rule="evenodd" clip-rule="evenodd" d="M7.97612 10.0719L12.3334 5.7146L12.9521 6.33332L8.28548 11L7.66676 11L3.0001 6.33332L3.61882 5.7146L7.97612 10.0719Z" fill="{color}"/>'
    SVG_CHEVRON_RIGHT = '<path fill-rule="evenodd" clip-rule="evenodd" d="M10.0719 8.02388L5.7146 3.66658L6.33332 3.04785L11 7.71452V8.33324L6.33332 12.9999L5.7146 12.3812L10.0719 8.02388Z" fill="{color}"/>'
    SVG_FOLDER = '<path d="M14.5 3H7.71L6.86 2.15L6.5 2H1.5L1 2.5V12.5L1.5 13H14.5L15 12.5V3.5L14.5 3ZM14 12H2V6H14V12ZM14 5H2V3H6.29L7.14 3.85L7.5 4H14V5Z" fill="{color}"/>'
    SVG_FILE = '<path d="M13.71 4.29L10 0.59L9.71 0.29L9.5 0H3.5L3 0.5V15.5L3.5 16H12.5L13 15.5V4.71L13.71 4.29ZM10 2.41L11.59 4H10V2.41ZM12 15H4V1H9V4.5L9.5 5H12V15Z" fill="{color}"/>'
    SVG_CHECK = '<path fill-rule="evenodd" clip-rule="evenodd" d="M14.431 3.323L5.569 12.177L1.56904 8.177L2.431 7.323L5.569 10.461L13.569 2.461L14.431 3.323Z" fill="{color}"/>'
    SVG_ERROR = '<path fill-rule="evenodd" clip-rule="evenodd" d="M8.6 1C9.2 1 9.7 1.3 10 1.8L14.7 10.3C15 10.8 15 11.4 14.7 11.9C14.4 12.4 13.9 12.7 13.3 12.7H3.7C3.1 12.7 2.6 12.4 2.3 11.9C2 11.4 2 10.8 2.3 10.3L7 1.8C7.3 1.3 7.8 1 8.5 1H8.6ZM8.5 2.5C8.5 2.5 8.5 2.5 8.5 2.5L3.8 11C3.7 11.2 3.8 11.4 4 11.5H13C13.2 11.4 13.3 11.2 13.2 11L8.5 2.5ZM9 10H8V9H9V10ZM9 8H8V5H9V8Z" fill="{color}"/>'
    SVG_WARNING = '<path fill-rule="evenodd" clip-rule="evenodd" d="M7.56 1H8.44L14.98 13.5799L14.54 14.28H1.46L1.02 13.5799L7.56 1ZM8 2.28L2.28 13.28H13.72L8 2.28ZM8.5 12H7.5V11H8.5V12ZM7.5 10V6.5H8.5V10H7.5Z" fill="{color}"/>'
    SVG_INFO = '<path fill-rule="evenodd" clip-rule="evenodd" d="M8.568 1.031C6.38 1.031 4.279 1.901 2.732 3.448C1.185 4.995 0.316 7.095 0.316 9.284C0.316 11.472 1.186 13.573 2.733 15.12C4.28 16.667 6.38 17.536 8.568 17.536C10.757 17.536 12.857 16.667 14.404 15.12C15.951 13.573 16.821 11.472 16.821 9.284C16.821 7.096 15.951 4.995 14.404 3.448C12.857 1.901 10.757 1.031 8.568 1.031ZM8.568 16.536C5.677 16.536 3.068 14.936 1.822 12.408C0.577 9.879 0.86 6.866 2.567 4.62C4.274 2.374 7.132 1.331 9.874 1.908C12.616 2.485 14.817 4.579 15.527 7.284C16.238 9.989 15.336 12.872 13.18 14.676C11.681 15.917 9.759 16.572 7.784 16.536L8.568 16.536ZM9.068 5.284H8.068V4.284H9.068V5.284ZM9.068 6.784V13.784H8.068V6.784H9.068Z" fill="{color}" transform="translate(0,-1) scale(0.9)"/>'
    SVG_SEND = '<path d="M1 1.91L1.78 1.5L15 8L1.78 14.5L1 14.09V9.09L8 8L1 6.91V1.91Z" fill="{color}"/>'
    SVG_EYE = '<path fill-rule="evenodd" clip-rule="evenodd" d="M16 8C16 8 13 3 8 3C3 3 0 8 0 8C0 8 3 13 8 13C13 13 16 8 16 8ZM1.173 8C1.654 7.38 2.3 6.64 3.085 5.925C4.64 4.5 6.343 3.5 8 3.5C9.657 3.5 11.36 4.5 12.915 5.925C13.7 6.64 14.346 7.38 14.827 8C14.346 8.62 13.7 9.36 12.915 10.075C11.36 11.5 9.657 12.5 8 12.5C6.343 12.5 4.64 11.5 3.085 10.075C2.3 9.36 1.654 8.62 1.173 8ZM8 10.5C9.38 10.5 10.5 9.38 10.5 8C10.5 6.62 9.38 5.5 8 5.5C6.62 5.5 5.5 6.62 5.5 8C5.5 9.38 6.62 10.5 8 10.5Z" fill="{color}"/>'
    SVG_CLEAR = '<path d="M8 2C8.55228 2 9 2.44772 9 3V4H13C13.5523 4 14 4.44772 14 5C14 5.55228 13.5523 6 13 6H12.9199L12.1504 13.166C12.0603 14.195 11.2006 15 10.167 15H5.83301C4.79943 15 3.93965 14.195 3.84962 13.166L3.0801 6H3C2.44772 6 2 5.55228 2 5C2 4.44772 2.44772 4 3 4H7V3C7 2.44772 7.44772 2 8 2ZM5.0918 6L5.8418 13.055C5.8618 13.276 6.0518 13.5 6.3418 13.5L9.6418 13.5C9.9318 13.5 10.1218 13.276 10.1418 13.055L10.9082 6H5.0918Z" fill="{color}"/>'
    SVG_TERMINAL = '<path d="M1 2.795L1.5 2H14.5L15 2.795V13.205L14.5 14H1.5L1 13.205V2.795ZM2 13H14V3H2V13ZM5.146 10.146L4.439 9.439L6.878 7L4.439 4.561L5.146 3.854L8.292 7L5.146 10.146ZM11 10H8V9H11V10Z" fill="{color}"/>'

    @classmethod
    def create_icon(cls, icon_name_or_func, size: int = 16, color: str = None) -> QIcon:
        """Create a QIcon from an SVG path or drawing function with caching."""
        from PySide6.QtSvg import QSvgRenderer
        from PySide6.QtCore import QByteArray

        # Resolve color for cache key
        resolved_color = color or Theme.TEXT_SECONDARY

        # Create cache key - use id() for callables since they're not hashable by content
        if callable(icon_name_or_func):
            cache_key = (id(icon_name_or_func), size, resolved_color)
        else:
            cache_key = (icon_name_or_func, size, resolved_color)

        # Return cached icon if available
        if cache_key in cls._icon_cache:
            return cls._icon_cache[cache_key]

        # If it's a callable (old-style), use the old method
        if callable(icon_name_or_func):
            pixmap = QPixmap(size, size)
            pixmap.fill(Qt.GlobalColor.transparent)
            painter = QPainter(pixmap)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            c = QColor(resolved_color)
            painter.setPen(QPen(c, 1.5))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            icon_name_or_func(painter, size, c)
            painter.end()
            icon = QIcon(pixmap)
        else:
            # New style: SVG string
            svg_path = icon_name_or_func
            svg_content = f'''<svg width="{size}" height="{size}" viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg">{svg_path.format(color=resolved_color)}</svg>'''

            pixmap = QPixmap(size, size)
            pixmap.fill(Qt.GlobalColor.transparent)

            renderer = QSvgRenderer(QByteArray(svg_content.encode()))
            painter = QPainter(pixmap)
            renderer.render(painter)
            painter.end()
            icon = QIcon(pixmap)

        # Cache and return
        cls._icon_cache[cache_key] = icon
        return icon

    @classmethod
    def clear_cache(cls):
        """Clear the icon cache. Call when theme changes."""
        cls._icon_cache.clear()

    @classmethod
    def close(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_CLOSE, size, color)

    @classmethod
    def refresh(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_REFRESH, size, color)

    @classmethod
    def files(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_FILES, size, color)

    @classmethod
    def search(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_SEARCH, size, color)

    @classmethod
    def settings(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_SETTINGS, size, color)

    @classmethod
    def git_branch(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_GIT, size, color)

    @classmethod
    def git(cls, size: int = 16, color: str = None) -> QIcon:
        """Alias for git_branch."""
        return cls.create_icon(cls.SVG_GIT, size, color)

    @classmethod
    def robot(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_ROBOT, size, color)

    @classmethod
    def new_file(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_NEW_FILE, size, color)

    @classmethod
    def new_folder(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_NEW_FOLDER, size, color)

    @classmethod
    def collapse_all(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_COLLAPSE, size, color)

    @classmethod
    def add(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_ADD, size, color)

    @classmethod
    def folder(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_FOLDER, size, color)

    @classmethod
    def file(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_FILE, size, color)

    @classmethod
    def check(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_CHECK, size, color)

    @classmethod
    def send(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_SEND, size, color)

    @classmethod
    def eye(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_EYE, size, color)

    @classmethod
    def clear(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_CLEAR, size, color)

    @classmethod
    def terminal(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_TERMINAL, size, color)

    @classmethod
    def chevron_down(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_CHEVRON_DOWN, size, color)

    @classmethod
    def chevron_right(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_CHEVRON_RIGHT, size, color)

    @classmethod
    def warning(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_WARNING, size, color)

    @classmethod
    def error(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_ERROR, size, color)

    @classmethod
    def info(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_INFO, size, color)

    @classmethod
    def plus(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_ADD, size, color)

    @classmethod
    def arrow_down(cls, size: int = 16, color: str = None) -> QIcon:
        return cls.create_icon(cls.SVG_CHEVRON_DOWN, size, color)

    @classmethod
    def arrow_up(cls, size: int = 16, color: str = None) -> QIcon:
        # Rotate the chevron up
        svg_up = '<path fill-rule="evenodd" clip-rule="evenodd" d="M7.97612 5.9281L12.3334 10.2854L12.9521 9.66668L8.28548 5L7.66676 5L3.0001 9.66668L3.61882 10.2854L7.97612 5.9281Z" fill="{color}"/>'
        return cls.create_icon(svg_up, size, color)


# Clear icon cache when theme changes
ThemeManager.on_change(lambda t: Icons.clear_cache())


# ============================================================================
# File Icon Provider
# ============================================================================

class FileTypeIconProvider:
    """Provides file type-specific icons for the file explorer using VS Code-style SVGs."""

    # SVG paths for specific file types (from VS Code icons)
    SVG_PYTHON = '<path d="M8 1C6.27 1 5.56 1.84 5.56 3.28V4.56H8.22V5H3.56C2.11 5 1 6.25 1 7.78C1 9.31 2.11 10.56 3.56 10.56H4.56V9C4.56 7.47 5.75 6.22 7.28 6.22H10.17C11.36 6.22 12.33 5.25 12.33 4.06V3.28C12.33 1.84 11.17 1 9.44 1H8ZM6.78 2.39C7.19 2.39 7.5 2.7 7.5 3.11C7.5 3.52 7.19 3.83 6.78 3.83C6.36 3.83 6.06 3.52 6.06 3.11C6.06 2.7 6.36 2.39 6.78 2.39Z" fill="#3572A5"/><path d="M11.44 5V6.56C11.44 8.08 10.25 9.33 8.72 9.33H5.83C4.64 9.33 3.67 10.31 3.67 11.5V12.28C3.67 13.72 4.83 14.56 6.56 14.56H8C9.73 14.56 10.44 13.72 10.44 12.28V11H7.78V10.56H12.44C13.89 10.56 15 9.31 15 7.78C15 6.25 13.89 5 12.44 5H11.44ZM9.22 11.72C9.64 11.72 9.94 12.03 9.94 12.44C9.94 12.86 9.64 13.17 9.22 13.17C8.81 13.17 8.5 12.86 8.5 12.44C8.5 12.03 8.81 11.72 9.22 11.72Z" fill="#FFD845"/>'
    SVG_JAVASCRIPT = '<path d="M1 1H15V15H1V1Z" fill="#F7DF1E"/><path d="M10.7 12.4C11 13.1 11.5 13.5 12.3 13.5C13 13.5 13.4 13.2 13.4 12.7C13.4 12.2 13 11.9 12.2 11.6C10.9 11.1 10.1 10.5 10.1 9.2C10.1 8 11 7 12.5 7C13.6 7 14.4 7.4 14.9 8.3L13.7 9.1C13.4 8.6 13 8.3 12.5 8.3C12 8.3 11.6 8.5 11.6 9C11.6 9.5 12 9.7 12.8 10C14.2 10.5 15 11.1 15 12.5C15 13.8 14 14.8 12.3 14.8C10.7 14.8 9.6 14.1 9.2 13L10.7 12.4Z" fill="black"/><path d="M5.7 12.5C5.9 13 6.3 13.5 7 13.5C7.6 13.5 8 13.2 8 12.4V7.2H9.5V12.4C9.5 14 8.6 14.8 7.1 14.8C5.7 14.8 4.9 14 4.5 13.1L5.7 12.5Z" fill="black"/>'
    SVG_TYPESCRIPT = '<path d="M1 1H15V15H1V1Z" fill="#3178C6"/><path d="M5.5 10.1H7V14H8.4V10.1H9.9V9H5.5V10.1Z" fill="white"/><path d="M10.4 13.9C10.7 14.1 11.2 14.3 11.8 14.3C13.1 14.3 13.9 13.6 13.9 12.5C13.9 11.6 13.4 11.1 12.4 10.7C11.6 10.4 11.4 10.2 11.4 9.9C11.4 9.5 11.7 9.3 12.2 9.3C12.6 9.3 13 9.5 13.3 9.7L13.8 8.9C13.4 8.5 12.8 8.3 12.2 8.3C11.1 8.3 10.3 9 10.3 10C10.3 10.8 10.8 11.4 11.7 11.7C12.5 12 12.7 12.2 12.7 12.6C12.7 13 12.4 13.3 11.8 13.3C11.3 13.3 10.7 13.1 10.3 12.7L9.8 13.5C10 13.8 10.2 13.9 10.4 13.9Z" fill="white"/>'
    SVG_JSON = '<path d="M5 3C3.9 3 3 3.9 3 5V6C3 6.6 2.6 7 2 7V9C2.6 9 3 9.4 3 10V11C3 12.1 3.9 13 5 13H6V11H5V10C5 9.2 4.5 8.5 3.8 8C4.5 7.5 5 6.8 5 6V5H6V3H5Z" fill="#CBB068"/><path d="M11 3H10V5H11V6C11 6.8 11.5 7.5 12.2 8C11.5 8.5 11 9.2 11 10V11H10V13H11C12.1 13 13 12.1 13 11V10C13 9.4 13.4 9 14 9V7C13.4 7 13 6.6 13 6V5C13 3.9 12.1 3 11 3Z" fill="#CBB068"/>'
    SVG_MARKDOWN = '<path d="M2 4H4L6 8L8 4H10V12H8V7.5L6 11L4 7.5V12H2V4Z" fill="#519ABA"/><path d="M11 8L14 11V8.5H15V12H14L11 9V11.5H10V8H11Z" fill="#519ABA"/>'
    SVG_HTML = '<path d="M2 1L3 14L8 15L13 14L14 1H2Z" fill="#E34C26"/><path d="M8 13.5V2.5H12.5L12 11.5L8 12.5V13.5Z" fill="#F06529"/><path d="M5 4H11L10.8 5.5H5.5L5.7 7H10.5L10 11L8 11.5V11L6 10.5L5.9 9.5H4.5L4.7 11.5L8 12.5L11.3 11.5L11.8 4.5L12 4H4L5 4Z" fill="white"/>'
    SVG_CSS = '<path d="M2 1L3 14L8 15L13 14L14 1H2Z" fill="#1572B6"/><path d="M8 13.5V2.5H12.5L12 11.5L8 12.5V13.5Z" fill="#33A9DC"/><path d="M5.3 7H8V8.5H4L3.8 6H8V4.5H3.5L3.3 3H12.5L12.3 4.5H9.5V6H12L11.5 11L8 12L4.5 11L4.3 9H5.8L5.9 10L8 10.5L10.1 10L10.3 8.5H5L5.3 7Z" fill="white"/>'
    SVG_CONFIG = '<path d="M9.1 4.4L8.6 2H7.4L6.9 4.4L6.2 4.7L4.2 3.4L3.3 4.2L4.6 6.2L4.4 6.9L2 7.4V8.6L4.4 9.1L4.7 9.9L3.4 11.9L4.2 12.7L6.2 11.4L7 11.7L7.4 14H8.6L9.1 11.6L9.9 11.3L11.9 12.6L12.7 11.8L11.4 9.8L11.7 9L14 8.6V7.4L11.6 6.9L11.3 6.1L12.6 4.1L11.8 3.3L9.8 4.6L9.1 4.4Z" fill="{color}"/><circle cx="8" cy="8" r="2" fill="{bg}"/>'
    SVG_GIT = '<path d="M14.7 7.3L8.7 1.3C8.3 0.9 7.7 0.9 7.3 1.3L5.8 2.8L7.6 4.6C8.1 4.4 8.7 4.5 9.1 4.9C9.5 5.3 9.6 5.9 9.4 6.4L11.2 8.2C11.7 8 12.3 8.1 12.7 8.5C13.3 9.1 13.3 10 12.7 10.6C12.1 11.2 11.2 11.2 10.6 10.6C10.2 10.2 10 9.5 10.3 9L8.6 7.3V11.4C8.8 11.5 9 11.7 9.1 11.9C9.7 12.5 9.7 13.4 9.1 14C8.5 14.6 7.6 14.6 7 14C6.4 13.4 6.4 12.5 7 11.9C7.2 11.7 7.4 11.5 7.6 11.4V7.2C7.4 7.1 7.2 6.9 7 6.7C6.6 6.3 6.4 5.6 6.7 5.1L4.9 3.3L1.3 7C0.9 7.4 0.9 8 1.3 8.4L7.3 14.4C7.7 14.8 8.3 14.8 8.7 14.4L14.7 8.4C15.1 8 15.1 7.7 14.7 7.3Z" fill="#F05032"/>'
    SVG_IMAGE = '<path d="M14 3H2C1.5 3 1 3.5 1 4V12C1 12.5 1.5 13 2 13H14C14.5 13 15 12.5 15 12V4C15 3.5 14.5 3 14 3ZM14 12H2V4H14V12Z" fill="{color}"/><circle cx="5" cy="6.5" r="1.5" fill="{color}"/><path d="M14 12H2L5 8L7 10L10 6L14 12Z" fill="{color}"/>'

    # Map file extensions to SVG paths and colors
    EXTENSION_MAP = {
        # Python
        '.py': ('python', None),
        '.pyw': ('python', None),
        '.pyi': ('python', None),
        # JavaScript
        '.js': ('javascript', None),
        '.jsx': ('javascript', None),
        '.mjs': ('javascript', None),
        # TypeScript
        '.ts': ('typescript', None),
        '.tsx': ('typescript', None),
        # JSON
        '.json': ('json', None),
        '.jsonc': ('json', None),
        # Markdown
        '.md': ('markdown', None),
        '.markdown': ('markdown', None),
        '.mdx': ('markdown', None),
        # HTML
        '.html': ('html', None),
        '.htm': ('html', None),
        '.xhtml': ('html', None),
        # CSS
        '.css': ('css', None),
        '.scss': ('css', None),
        '.sass': ('css', None),
        '.less': ('css', None),
        # Config
        '.yaml': ('config', '#CB171E'),
        '.yml': ('config', '#CB171E'),
        '.toml': ('config', '#9C4121'),
        '.ini': ('config', '#6D8086'),
        '.cfg': ('config', '#6D8086'),
        '.conf': ('config', '#6D8086'),
        '.env': ('config', '#ECD53F'),
        # Git
        '.gitignore': ('git', None),
        '.gitattributes': ('git', None),
        '.gitmodules': ('git', None),
        # Images
        '.png': ('image', '#A074C4'),
        '.jpg': ('image', '#A074C4'),
        '.jpeg': ('image', '#A074C4'),
        '.gif': ('image', '#A074C4'),
        '.svg': ('image', '#FFB13B'),
        '.ico': ('image', '#A074C4'),
        '.webp': ('image', '#A074C4'),
    }

    # Special filenames
    FILENAME_MAP = {
        'Dockerfile': ('config', '#2496ED'),
        'docker-compose.yml': ('config', '#2496ED'),
        'docker-compose.yaml': ('config', '#2496ED'),
        'Makefile': ('config', '#6D8086'),
        'README.md': ('markdown', None),
        'LICENSE': ('file', '#D4AF37'),
        'requirements.txt': ('python', None),
        'setup.py': ('python', None),
        'pyproject.toml': ('python', None),
        'package.json': ('json', None),
        'tsconfig.json': ('typescript', None),
        '.eslintrc': ('config', '#4B32C3'),
        '.prettierrc': ('config', '#F7B93E'),
    }

    _icon_cache: Dict[str, QIcon] = {}

    @classmethod
    def get_icon(cls, file_path: Path) -> Optional[QIcon]:
        """Get the appropriate icon for a file."""
        if file_path.is_dir():
            return cls._get_folder_icon()

        filename = file_path.name
        ext = file_path.suffix.lower()

        # Check special filenames first
        cache_key = filename if filename in cls.FILENAME_MAP else ext

        if cache_key in cls._icon_cache:
            return cls._icon_cache[cache_key]

        icon_type = None
        color = None

        if filename in cls.FILENAME_MAP:
            icon_type, color = cls.FILENAME_MAP[filename]
        elif ext in cls.EXTENSION_MAP:
            icon_type, color = cls.EXTENSION_MAP[ext]

        if icon_type:
            icon = cls._create_icon(icon_type, color)
            cls._icon_cache[cache_key] = icon
            return icon

        return None

    @classmethod
    def _get_folder_icon(cls) -> QIcon:
        """Get folder icon."""
        if 'folder' not in cls._icon_cache:
            cls._icon_cache['folder'] = Icons.folder(16, "#DCB67A")
        return cls._icon_cache['folder']

    @classmethod
    def _create_icon(cls, icon_type: str, color: str = None) -> QIcon:
        """Create an icon based on type using SVG paths."""
        from PySide6.QtSvg import QSvgRenderer
        from PySide6.QtCore import QByteArray

        svg_map = {
            'python': cls.SVG_PYTHON,
            'javascript': cls.SVG_JAVASCRIPT,
            'typescript': cls.SVG_TYPESCRIPT,
            'json': cls.SVG_JSON,
            'markdown': cls.SVG_MARKDOWN,
            'html': cls.SVG_HTML,
            'css': cls.SVG_CSS,
            'git': cls.SVG_GIT,
            'config': cls.SVG_CONFIG,
            'image': cls.SVG_IMAGE,
        }

        svg_path = svg_map.get(icon_type, Icons.SVG_FILE)
        fill_color = color or Theme.TEXT_SECONDARY

        # Handle special replacements for config and image icons
        svg_formatted = svg_path.format(color=fill_color, bg=Theme.BG_DARK)

        svg_content = f'''<svg width="16" height="16" viewBox="0 0 16 16" fill="none" xmlns="http://www.w3.org/2000/svg">{svg_formatted}</svg>'''

        pixmap = QPixmap(16, 16)
        pixmap.fill(Qt.GlobalColor.transparent)

        renderer = QSvgRenderer(QByteArray(svg_content.encode()))
        painter = QPainter(pixmap)
        renderer.render(painter)
        painter.end()

        return QIcon(pixmap)


class FileExplorerDelegate(QStyledItemDelegate):
    """Custom delegate to render file type icons in the explorer."""

    def __init__(self, model: QFileSystemModel, parent=None):
        super().__init__(parent)
        self.model = model

    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        file_path = Path(self.model.filePath(index))
        custom_icon = FileTypeIconProvider.get_icon(file_path)
        if custom_icon:
            option.icon = custom_icon


# ============================================================================
# Compact UI Components
# ============================================================================

class CompactLineEdit(QLineEdit):
    """Compact styled line edit."""

    def __init__(self, placeholder: str = "", parent=None):
        super().__init__(parent)
        self.setPlaceholderText(placeholder)
        self.setStyleSheet(f"""
            QLineEdit {{
                background: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 3px;
                padding: 5px 8px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 12px;
                selection-background-color: {Theme.ACCENT_BLUE};
            }}
            QLineEdit:focus {{
                border-color: {Theme.ACCENT_BLUE};
            }}
            QLineEdit::placeholder {{
                color: {Theme.TEXT_MUTED};
            }}
        """)


class CompactButton(QPushButton):
    """Compact primary button with enhanced visual feedback."""

    def __init__(self, text: str = "", parent=None):
        super().__init__(text, parent)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {Theme.ACCENT_BLUE}, stop:1 #0066cc);
                border: none;
                border-radius: 4px;
                padding: 6px 14px;
                color: white;
                font-size: 12px;
                font-weight: 500;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 {Theme.TEXT_LINK}, stop:1 {Theme.ACCENT_BLUE});
            }}
            QPushButton:pressed {{
                background: {Theme.BG_SELECTED};
                padding-top: 7px;
                padding-bottom: 5px;
            }}
            QPushButton:disabled {{
                background: {Theme.BG_TERTIARY};
                color: {Theme.TEXT_MUTED};
            }}
        """)


class SecondaryButton(QPushButton):
    """Compact secondary button with enhanced visual feedback."""

    def __init__(self, text: str = "", parent=None):
        super().__init__(text, parent)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setStyleSheet(f"""
            QPushButton {{
                background: {Theme.BG_TERTIARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 5px 12px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 11px;
            }}
            QPushButton:hover {{
                background: {Theme.BG_HOVER};
                border-color: {Theme.ACCENT_BLUE};
                color: {Theme.TEXT_PRIMARY};
            }}
            QPushButton:pressed {{
                background: {Theme.BG_SECONDARY};
                border-color: {Theme.ACCENT_BLUE};
            }}
            QPushButton:disabled {{
                background: {Theme.BG_SECONDARY};
                color: {Theme.TEXT_MUTED};
                border-color: {Theme.BORDER};
            }}
        """)


class IconButton(QPushButton):
    """Small icon-only button with hover effects."""

    def __init__(self, icon_method=None, tooltip: str = "", parent=None):
        super().__init__(parent)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFixedSize(24, 24)
        self.setToolTip(tooltip)
        self._icon_method = icon_method
        if icon_method:
            self.setIcon(icon_method(14, Theme.TEXT_SECONDARY))
            self.setIconSize(QSize(14, 14))
        self.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                border: none;
                border-radius: 4px;
            }}
            QPushButton:hover {{
                background: {Theme.BG_HOVER};
            }}
            QPushButton:pressed {{
                background: {Theme.BG_TERTIARY};
            }}
        """)

    def enterEvent(self, event):
        """Brighten icon on hover."""
        super().enterEvent(event)
        if self._icon_method:
            self.setIcon(self._icon_method(14, Theme.TEXT_PRIMARY))

    def leaveEvent(self, event):
        """Restore icon on leave."""
        super().leaveEvent(event)
        if self._icon_method:
            self.setIcon(self._icon_method(14, Theme.TEXT_SECONDARY))


class CompactCheckBox(QCheckBox):
    """Compact styled checkbox."""

    def __init__(self, text: str = "", parent=None):
        super().__init__(text, parent)
        self.setStyleSheet(f"""
            QCheckBox {{
                color: {Theme.TEXT_PRIMARY};
                font-size: 11px;
                spacing: 6px;
            }}
            QCheckBox::indicator {{
                width: 14px;
                height: 14px;
                border-radius: 3px;
                border: 1px solid {Theme.BORDER_LIGHT};
                background: {Theme.BG_INPUT};
            }}
            QCheckBox::indicator:checked {{
                background: {Theme.ACCENT_BLUE};
                border-color: {Theme.ACCENT_BLUE};
            }}
            QCheckBox::indicator:hover {{
                border-color: {Theme.ACCENT_BLUE};
            }}
        """)


class CompactComboBox(QComboBox):
    """Compact styled combo box."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QComboBox {{
                background: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 3px;
                padding: 4px 8px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 11px;
                min-width: 80px;
            }}
            QComboBox:hover {{
                border-color: {Theme.BORDER_LIGHT};
            }}
            QComboBox::drop-down {{
                border: none;
                width: 16px;
            }}
            QComboBox::down-arrow {{
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 4px solid {Theme.TEXT_SECONDARY};
                margin-right: 6px;
            }}
            QComboBox QAbstractItemView {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                selection-background-color: {Theme.BG_SELECTED};
                color: {Theme.TEXT_PRIMARY};
                font-size: 11px;
            }}
        """)


class CollapsibleSection(QWidget):
    """Collapsible section with compact header."""

    def __init__(self, title: str, parent=None):
        super().__init__(parent)
        self.is_collapsed = False
        self._title = title

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.header = QPushButton(f"  {title}")
        self.header.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                border: none;
                padding: 6px 4px;
                color: {Theme.TEXT_SECONDARY};
                font-size: 11px;
                font-weight: 600;
                text-transform: uppercase;
                text-align: left;
            }}
            QPushButton:hover {{
                color: {Theme.TEXT_PRIMARY};
            }}
        """)
        self.header.clicked.connect(self._toggle)
        layout.addWidget(self.header)

        self.content = QWidget()
        self.content_layout = QVBoxLayout(self.content)
        self.content_layout.setContentsMargins(8, 4, 8, 8)
        self.content_layout.setSpacing(6)
        layout.addWidget(self.content)

    def _toggle(self):
        self.is_collapsed = not self.is_collapsed
        self.content.setVisible(not self.is_collapsed)
        arrow = ">" if self.is_collapsed else "v"
        self.header.setText(f"{arrow}  {self._title}")

    def add_widget(self, widget):
        self.content_layout.addWidget(widget)

    def add_layout(self, layout):
        self.content_layout.addLayout(layout)


# ============================================================================
# Toast Notifications
# ============================================================================

class ToastNotification(QFrame):
    """Non-blocking toast notification that auto-dismisses."""

    class Type(Enum):
        SUCCESS = "success"
        ERROR = "error"
        WARNING = "warning"
        INFO = "info"

    _active_toasts: List["ToastNotification"] = []

    def __init__(self, message: str, toast_type: Type = Type.INFO, duration: int = 3000, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Tool | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setAttribute(Qt.WidgetAttribute.WA_ShowWithoutActivating)

        # Colors based on type
        colors = {
            self.Type.SUCCESS: (Theme.SUCCESS, "#1a3d1a"),
            self.Type.ERROR: (Theme.ERROR, "#3d1a1a"),
            self.Type.WARNING: (Theme.WARNING, "#3d3d1a"),
            self.Type.INFO: (Theme.ACCENT_BLUE, "#1a2a3d"),
        }
        accent_color, bg_color = colors.get(toast_type, colors[self.Type.INFO])

        self.setStyleSheet(f"""
            ToastNotification {{
                background: {bg_color};
                border: 1px solid {accent_color};
                border-left: 3px solid {accent_color};
                border-radius: 4px;
            }}
        """)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(12, 8, 12, 8)
        layout.setSpacing(8)

        # Icon based on type
        icons = {
            self.Type.SUCCESS: Icons.check,
            self.Type.ERROR: Icons.error,
            self.Type.WARNING: Icons.warning,
            self.Type.INFO: Icons.info,
        }
        icon_fn = icons.get(toast_type, Icons.info)
        icon_label = QLabel()
        icon_label.setPixmap(icon_fn(14, accent_color).pixmap(14, 14))
        layout.addWidget(icon_label)

        label = QLabel(message)
        label.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 12px;")
        label.setWordWrap(True)
        layout.addWidget(label, 1)

        # Close button
        close_btn = QPushButton()
        close_btn.setFixedSize(16, 16)
        close_btn.setIcon(Icons.close(12, Theme.TEXT_MUTED))
        close_btn.setStyleSheet("QPushButton { background: transparent; border: none; }")
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn)

        self.setFixedWidth(320)
        self.adjustSize()

        # Auto dismiss timer
        self._timer = QTimer(self)
        self._timer.setSingleShot(True)
        self._timer.timeout.connect(self.close)
        self._timer.start(duration)

        # Animation for fade in
        self._opacity = 0.0
        self._fade_timer = QTimer(self)
        self._fade_timer.timeout.connect(self._fade_in)
        self._fade_timer.start(16)

    def _fade_in(self):
        self._opacity = min(1.0, self._opacity + 0.1)
        self.setWindowOpacity(self._opacity)
        if self._opacity >= 1.0:
            self._fade_timer.stop()

    def show(self):
        # Position in bottom-right of parent window
        if self.parent():
            parent_rect = self.parent().rect()
            parent_pos = self.parent().mapToGlobal(parent_rect.bottomRight())
            # Stack toasts
            offset = len(ToastNotification._active_toasts) * (self.height() + 8)
            self.move(parent_pos.x() - self.width() - 20, parent_pos.y() - self.height() - 20 - offset)
        ToastNotification._active_toasts.append(self)
        super().show()

    def closeEvent(self, event):
        if self in ToastNotification._active_toasts:
            ToastNotification._active_toasts.remove(self)
        super().closeEvent(event)

    @classmethod
    def success(cls, message: str, parent=None, duration: int = 3000):
        toast = cls(message, cls.Type.SUCCESS, duration, parent)
        toast.show()
        return toast

    @classmethod
    def error(cls, message: str, parent=None, duration: int = 5000):
        toast = cls(message, cls.Type.ERROR, duration, parent)
        toast.show()
        return toast

    @classmethod
    def warning(cls, message: str, parent=None, duration: int = 4000):
        toast = cls(message, cls.Type.WARNING, duration, parent)
        toast.show()
        return toast

    @classmethod
    def info(cls, message: str, parent=None, duration: int = 3000):
        toast = cls(message, cls.Type.INFO, duration, parent)
        toast.show()
        return toast


# ============================================================================
# Pulsing Status Indicator
# ============================================================================

class StatusDot(QWidget):
    """Simple status dot indicator."""

    def __init__(self, color: str = None, size: int = 6, parent=None):
        super().__init__(parent)
        self._color = QColor(color or Theme.TEXT_MUTED)
        self._size = size
        self.setFixedSize(size + 4, size + 4)

    def set_color(self, color: str):
        self._color = QColor(color)
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setBrush(self._color)
        painter.setPen(Qt.PenStyle.NoPen)
        offset = (self.width() - self._size) // 2
        painter.drawEllipse(offset, offset, self._size, self._size)


# ============================================================================
# Syntax Highlighter
# ============================================================================

# Try to import Pygments for enhanced syntax highlighting
try:
    from pygments import highlight
    from pygments.lexers import get_lexer_for_filename, get_lexer_by_name, TextLexer
    from pygments.token import Token
    PYGMENTS_AVAILABLE = True
except ImportError:
    PYGMENTS_AVAILABLE = False


class PygmentsHighlighter(QSyntaxHighlighter):
    """Enhanced syntax highlighter using Pygments for multi-language support."""

    # Token type to color mapping (VS Code dark theme inspired)
    TOKEN_COLORS = {
        Token.Keyword: Theme.ACCENT_PURPLE,
        Token.Keyword.Constant: Theme.ACCENT_PURPLE,
        Token.Keyword.Declaration: Theme.ACCENT_PURPLE,
        Token.Keyword.Namespace: Theme.ACCENT_PURPLE,
        Token.Keyword.Pseudo: Theme.ACCENT_PURPLE,
        Token.Keyword.Reserved: Theme.ACCENT_PURPLE,
        Token.Keyword.Type: Theme.ACCENT_CYAN,
        Token.Name.Builtin: Theme.ACCENT_CYAN,
        Token.Name.Builtin.Pseudo: Theme.ACCENT_CYAN,
        Token.Name.Class: Theme.ACCENT_CYAN,
        Token.Name.Decorator: Theme.ACCENT_YELLOW,
        Token.Name.Exception: Theme.ACCENT_RED,
        Token.Name.Function: Theme.ACCENT_YELLOW,
        Token.Name.Function.Magic: Theme.ACCENT_YELLOW,
        Token.Name.Variable: Theme.TEXT_PRIMARY,
        Token.Name.Variable.Class: Theme.ACCENT_CYAN,
        Token.Name.Variable.Global: Theme.TEXT_PRIMARY,
        Token.Name.Variable.Instance: Theme.TEXT_PRIMARY,
        Token.Name.Constant: Theme.ACCENT_BLUE,
        Token.Name.Tag: Theme.ACCENT_RED,
        Token.Name.Attribute: Theme.ACCENT_YELLOW,
        Token.String: Theme.ACCENT_ORANGE,
        Token.String.Doc: Theme.ACCENT_GREEN,
        Token.String.Escape: Theme.ACCENT_CYAN,
        Token.String.Interpol: Theme.ACCENT_CYAN,
        Token.String.Regex: Theme.ACCENT_RED,
        Token.Number: Theme.ACCENT_GREEN,
        Token.Number.Float: Theme.ACCENT_GREEN,
        Token.Number.Hex: Theme.ACCENT_GREEN,
        Token.Number.Integer: Theme.ACCENT_GREEN,
        Token.Number.Oct: Theme.ACCENT_GREEN,
        Token.Operator: Theme.TEXT_SECONDARY,
        Token.Operator.Word: Theme.ACCENT_PURPLE,
        Token.Comment: Theme.TEXT_MUTED,
        Token.Comment.Multiline: Theme.TEXT_MUTED,
        Token.Comment.Single: Theme.TEXT_MUTED,
        Token.Comment.Special: Theme.TEXT_MUTED,
        Token.Generic.Deleted: Theme.ACCENT_RED,
        Token.Generic.Inserted: Theme.ACCENT_GREEN,
        Token.Generic.Heading: Theme.ACCENT_BLUE,
        Token.Generic.Subheading: Theme.ACCENT_CYAN,
        Token.Generic.Strong: Theme.TEXT_PRIMARY,
        Token.Generic.Emph: Theme.TEXT_SECONDARY,
    }

    def __init__(self, document, filename: str = None):
        super().__init__(document)
        self._lexer = None
        self._formats = {}
        self._setup_formats()
        if filename:
            self.set_language_from_file(filename)

    def _setup_formats(self):
        """Create QTextCharFormat for each token type."""
        for token_type, color in self.TOKEN_COLORS.items():
            fmt = QTextCharFormat()
            fmt.setForeground(QColor(color))
            # Make keywords bold
            if token_type in (Token.Keyword, Token.Keyword.Constant,
                            Token.Keyword.Declaration, Token.Keyword.Namespace,
                            Token.Keyword.Reserved):
                fmt.setFontWeight(QFont.Weight.Bold)
            # Make comments italic
            if str(token_type).startswith('Token.Comment'):
                fmt.setFontItalic(True)
            self._formats[token_type] = fmt

    def set_language_from_file(self, filename: str):
        """Set lexer based on filename."""
        if not PYGMENTS_AVAILABLE:
            return
        try:
            self._lexer = get_lexer_for_filename(filename)
        except Exception:
            self._lexer = TextLexer()
        self.rehighlight()

    def set_language(self, language: str):
        """Set lexer by language name."""
        if not PYGMENTS_AVAILABLE:
            return
        try:
            self._lexer = get_lexer_by_name(language)
        except Exception:
            self._lexer = TextLexer()
        self.rehighlight()

    def _get_format(self, token_type):
        """Get format for token type, checking parent types."""
        if token_type in self._formats:
            return self._formats[token_type]
        # Check parent token types
        for parent_type in self.TOKEN_COLORS:
            if str(token_type).startswith(str(parent_type)):
                return self._formats.get(parent_type)
        return None

    def highlightBlock(self, text):
        """Highlight a block of text using Pygments."""
        if not PYGMENTS_AVAILABLE or not self._lexer:
            return

        try:
            tokens = list(self._lexer.get_tokens(text))
            index = 0
            for token_type, value in tokens:
                length = len(value)
                fmt = self._get_format(token_type)
                if fmt:
                    self.setFormat(index, length, fmt)
                index += length
        except Exception:
            pass


class PythonHighlighter(QSyntaxHighlighter):
    """Fallback syntax highlighter for Python code when Pygments unavailable."""

    def __init__(self, document):
        super().__init__(document)
        self._rules = []
        self._setup_rules()

    def _setup_rules(self):
        keyword_fmt = QTextCharFormat()
        keyword_fmt.setForeground(QColor(Theme.ACCENT_PURPLE))
        keyword_fmt.setFontWeight(QFont.Weight.Bold)

        keywords = [
            "and", "as", "assert", "async", "await", "break", "class",
            "continue", "def", "del", "elif", "else", "except", "finally",
            "for", "from", "global", "if", "import", "in", "is", "lambda",
            "not", "or", "pass", "raise", "return", "try", "while", "with", "yield"
        ]
        for word in keywords:
            self._rules.append((re.compile(rf"\b{word}\b"), keyword_fmt))

        string_fmt = QTextCharFormat()
        string_fmt.setForeground(QColor(Theme.ACCENT_ORANGE))
        self._rules.append((re.compile(r'"[^"\\]*(\\.[^"\\]*)*"'), string_fmt))
        self._rules.append((re.compile(r"'[^'\\]*(\\.[^'\\]*)*'"), string_fmt))

        comment_fmt = QTextCharFormat()
        comment_fmt.setForeground(QColor(Theme.TEXT_MUTED))
        comment_fmt.setFontItalic(True)
        self._rules.append((re.compile(r"#.*"), comment_fmt))

        function_fmt = QTextCharFormat()
        function_fmt.setForeground(QColor(Theme.ACCENT_YELLOW))
        self._rules.append((re.compile(r"\b([a-zA-Z_][a-zA-Z0-9_]*)\s*\("), function_fmt))

        number_fmt = QTextCharFormat()
        number_fmt.setForeground(QColor(Theme.ACCENT_GREEN))
        self._rules.append((re.compile(r"\b\d+\.?\d*\b"), number_fmt))

        # Class names
        class_fmt = QTextCharFormat()
        class_fmt.setForeground(QColor(Theme.ACCENT_CYAN))
        self._rules.append((re.compile(r"\bclass\s+(\w+)"), class_fmt))

        # Decorators
        decorator_fmt = QTextCharFormat()
        decorator_fmt.setForeground(QColor(Theme.ACCENT_YELLOW))
        self._rules.append((re.compile(r"@\w+"), decorator_fmt))

    def highlightBlock(self, text):
        for pattern, fmt in self._rules:
            for match in pattern.finditer(text):
                self.setFormat(match.start(), match.end() - match.start(), fmt)


# ============================================================================
# Welcome Screen (VS Code Style)
# ============================================================================

class WelcomeScreen(QWidget):
    """Professional VS Code-style welcome screen with icons."""

    open_folder_requested = Signal()
    open_file_requested = Signal()
    new_file_requested = Signal()
    project_selected = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"background: {Theme.BG_MAIN};")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(60, 40, 60, 40)
        layout.setSpacing(0)

        # Center content
        center = QWidget()
        center_layout = QVBoxLayout(center)
        center_layout.setContentsMargins(0, 0, 0, 0)
        center_layout.setSpacing(40)

        # Logo/Title with circuit icon
        title_container = QWidget()
        title_layout = QHBoxLayout(title_container)
        title_layout.setContentsMargins(0, 0, 0, 0)
        title_layout.setSpacing(16)

        # Circuit logo icon with robot inside gradient box
        logo_container = QWidget()
        logo_container.setFixedSize(48, 48)
        logo_container.setStyleSheet(f"""
            QWidget {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {Theme.ACCENT_BLUE}, stop:1 {Theme.ACCENT_CYAN});
                border-radius: 12px;
            }}
        """)
        logo_layout = QVBoxLayout(logo_container)
        logo_layout.setContentsMargins(0, 0, 0, 0)
        logo_layout.setSpacing(0)
        logo_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo_icon = QLabel()
        logo_icon.setFixedSize(24, 24)
        logo_icon.setPixmap(Icons.robot(24, "#1E1E1E").pixmap(24, 24))
        logo_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo_layout.addWidget(logo_icon)
        title_layout.addWidget(logo_container)

        # Title text
        title_text = QVBoxLayout()
        title_text.setSpacing(4)

        logo = QLabel("Circuit IDE")
        logo.setStyleSheet(f"""
            color: {Theme.TEXT_PRIMARY};
            font-size: 28px;
            font-weight: 500;
        """)
        title_text.addWidget(logo)

        version = QLabel("v8.0 - AI-Powered Coding Assistant")
        version.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 12px;")
        title_text.addWidget(version)

        title_layout.addLayout(title_text)
        title_layout.addStretch()
        center_layout.addWidget(title_container)

        # Two column layout
        columns = QHBoxLayout()
        columns.setSpacing(80)

        # Left column - Start
        left_col = QVBoxLayout()
        left_col.setSpacing(16)

        start_label = QLabel("Start")
        start_label.setStyleSheet(f"""
            color: {Theme.TEXT_SECONDARY};
            font-size: 11px;
            font-weight: 600;
            letter-spacing: 1px;
        """)
        left_col.addWidget(start_label)

        # Action buttons with icons
        actions = [
            (Icons.new_file, "New File", self.new_file_requested, "Create a new file", "Ctrl+N"),
            (Icons.file, "Open File...", self.open_file_requested, "Open an existing file", ""),
            (Icons.folder, "Open Folder...", self.open_folder_requested, "Open a folder as workspace", "Ctrl+O"),
        ]

        for icon_method, text, signal, tooltip, shortcut in actions:
            btn_container = QWidget()
            btn_layout = QHBoxLayout(btn_container)
            btn_layout.setContentsMargins(0, 4, 0, 4)
            btn_layout.setSpacing(10)

            icon_label = QLabel()
            icon_label.setPixmap(icon_method(16, Theme.TEXT_LINK).pixmap(16, 16))
            icon_label.setFixedSize(16, 16)
            btn_layout.addWidget(icon_label)

            btn = QPushButton(text)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setToolTip(tooltip)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: transparent;
                    border: none;
                    color: {Theme.TEXT_LINK};
                    font-size: 13px;
                    text-align: left;
                    padding: 0;
                }}
                QPushButton:hover {{
                    color: {Theme.ACCENT_CYAN};
                }}
            """)
            btn.clicked.connect(signal.emit)
            btn_layout.addWidget(btn)

            if shortcut:
                shortcut_label = QLabel(shortcut)
                shortcut_label.setStyleSheet(f"""
                    color: {Theme.TEXT_MUTED};
                    font-size: 10px;
                    background: {Theme.BG_TERTIARY};
                    padding: 2px 6px;
                    border-radius: 3px;
                """)
                btn_layout.addWidget(shortcut_label)

            btn_layout.addStretch()
            left_col.addWidget(btn_container)

        left_col.addStretch()
        columns.addLayout(left_col)

        # Right column - Recent
        right_col = QVBoxLayout()
        right_col.setSpacing(16)

        recent_header = QHBoxLayout()
        recent_label = QLabel("Recent")
        recent_label.setStyleSheet(f"""
            color: {Theme.TEXT_SECONDARY};
            font-size: 11px;
            font-weight: 600;
            letter-spacing: 1px;
        """)
        recent_header.addWidget(recent_label)
        recent_header.addStretch()
        right_col.addLayout(recent_header)

        # Recent projects list with scroll
        self.recent_list = QWidget()
        self.recent_layout = QVBoxLayout(self.recent_list)
        self.recent_layout.setContentsMargins(0, 0, 0, 0)
        self.recent_layout.setSpacing(8)
        right_col.addWidget(self.recent_list)

        self._load_recent_projects()

        right_col.addStretch()
        columns.addLayout(right_col)

        center_layout.addLayout(columns)
        center_layout.addStretch()

        # Footer with helpful shortcuts
        footer_container = QWidget()
        footer_layout = QHBoxLayout(footer_container)
        footer_layout.setContentsMargins(0, 0, 0, 0)
        footer_layout.setSpacing(24)

        shortcuts = [
            ("Ctrl+O", "Open folder"),
            ("Ctrl+`", "Toggle terminal"),
            ("Ctrl+Shift+F", "Search files"),
        ]

        for key, desc in shortcuts:
            shortcut_widget = QHBoxLayout()
            shortcut_widget.setSpacing(6)

            key_label = QLabel(key)
            key_label.setStyleSheet(f"""
                color: {Theme.TEXT_PRIMARY};
                font-size: 10px;
                background: {Theme.BG_TERTIARY};
                padding: 3px 6px;
                border-radius: 3px;
                font-weight: 500;
            """)
            shortcut_widget.addWidget(key_label)

            desc_label = QLabel(desc)
            desc_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 10px;")
            shortcut_widget.addWidget(desc_label)

            footer_layout.addLayout(shortcut_widget)

        footer_layout.addStretch()
        center_layout.addWidget(footer_container)

        layout.addWidget(center, alignment=Qt.AlignmentFlag.AlignCenter)

    def _load_recent_projects(self):
        # Clear existing
        while self.recent_layout.count():
            item = self.recent_layout.takeAt(0)
            if w := item.widget():
                w.deleteLater()

        recent = RecentProjects.get()
        if not recent:
            no_recent = QLabel("No recent folders")
            no_recent.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 12px;")
            self.recent_layout.addWidget(no_recent)
            return

        for path in recent[:6]:
            p = Path(path)

            # Container for each recent project
            item = QWidget()
            item_layout = QHBoxLayout(item)
            item_layout.setContentsMargins(0, 4, 0, 4)
            item_layout.setSpacing(10)

            # Folder icon
            folder_icon = QLabel()
            folder_icon.setPixmap(Icons.folder(16, Theme.TEXT_LINK).pixmap(16, 16))
            folder_icon.setFixedSize(16, 16)
            item_layout.addWidget(folder_icon)

            # Project info
            info_layout = QVBoxLayout()
            info_layout.setSpacing(1)

            name_btn = QPushButton(p.name)
            name_btn.setToolTip(str(p))
            name_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            name_btn.setStyleSheet(f"""
                QPushButton {{
                    background: transparent;
                    border: none;
                    color: {Theme.TEXT_LINK};
                    font-size: 12px;
                    text-align: left;
                    padding: 0;
                }}
                QPushButton:hover {{
                    color: {Theme.ACCENT_CYAN};
                }}
            """)
            name_btn.clicked.connect(lambda checked, pt=path: self.project_selected.emit(pt))
            info_layout.addWidget(name_btn)

            path_label = QLabel(str(p.parent)[:50] + ("..." if len(str(p.parent)) > 50 else ""))
            path_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 10px;")
            info_layout.addWidget(path_label)

            item_layout.addLayout(info_layout)
            item_layout.addStretch()

            self.recent_layout.addWidget(item)

    def refresh_recent(self):
        self._load_recent_projects()


# ============================================================================
# File Explorer - VS Code Style
# ============================================================================

class FileExplorer(QWidget):
    """VS Code-style file explorer with full functionality."""

    file_selected = Signal(Path)
    file_created = Signal(Path)
    file_deleted = Signal(Path)
    file_renamed = Signal(Path, Path)  # old_path, new_path

    def __init__(self, root_path: str = None, parent=None):
        super().__init__(parent)
        self.root_path = Path(root_path) if root_path else None
        self._editing_index = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Header with title and action buttons
        header = QWidget()
        header.setFixedHeight(28)
        header.setStyleSheet(f"background: {Theme.BG_SECONDARY};")
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(8, 0, 4, 0)
        header_layout.setSpacing(0)

        title = QLabel("EXPLORER")
        title.setStyleSheet(f"""
            color: {Theme.TEXT_SECONDARY};
            font-size: 10px;
            font-weight: 600;
            letter-spacing: 0.5px;
        """)
        header_layout.addWidget(title)
        header_layout.addStretch()

        # Action buttons
        btn_style = f"""
            QPushButton {{
                background: transparent;
                border: none;
                padding: 4px;
                border-radius: 3px;
            }}
            QPushButton:hover {{
                background: {Theme.BG_HOVER};
            }}
        """

        self.new_file_btn = QPushButton()
        self.new_file_btn.setIcon(Icons.new_file(14, Theme.TEXT_MUTED))
        self.new_file_btn.setFixedSize(22, 22)
        self.new_file_btn.setToolTip("New File (Ctrl+N)")
        self.new_file_btn.setStyleSheet(btn_style)
        self.new_file_btn.clicked.connect(self._new_file)
        header_layout.addWidget(self.new_file_btn)

        self.new_folder_btn = QPushButton()
        self.new_folder_btn.setIcon(Icons.new_folder(14, Theme.TEXT_MUTED))
        self.new_folder_btn.setFixedSize(22, 22)
        self.new_folder_btn.setToolTip("New Folder")
        self.new_folder_btn.setStyleSheet(btn_style)
        self.new_folder_btn.clicked.connect(self._new_folder)
        header_layout.addWidget(self.new_folder_btn)

        self.refresh_btn = QPushButton()
        self.refresh_btn.setIcon(Icons.refresh(14, Theme.TEXT_MUTED))
        self.refresh_btn.setFixedSize(22, 22)
        self.refresh_btn.setToolTip("Refresh Explorer")
        self.refresh_btn.setStyleSheet(btn_style)
        self.refresh_btn.clicked.connect(self._refresh)
        header_layout.addWidget(self.refresh_btn)

        self.collapse_btn = QPushButton()
        self.collapse_btn.setIcon(Icons.collapse_all(14, Theme.TEXT_MUTED))
        self.collapse_btn.setFixedSize(22, 22)
        self.collapse_btn.setToolTip("Collapse All")
        self.collapse_btn.setStyleSheet(btn_style)
        self.collapse_btn.clicked.connect(self._collapse_all)
        header_layout.addWidget(self.collapse_btn)

        layout.addWidget(header)

        # File system model
        self.model = QFileSystemModel()
        self.model.setReadOnly(False)  # Allow rename/delete
        if self.root_path:
            self.model.setRootPath(str(self.root_path))

        # Tree view
        self.tree = QTreeView()
        self.tree.setModel(self.model)
        if self.root_path:
            self.tree.setRootIndex(self.model.index(str(self.root_path)))
        self.tree.setHeaderHidden(True)
        self.tree.hideColumn(1)  # Size
        self.tree.hideColumn(2)  # Type
        self.tree.hideColumn(3)  # Date Modified
        self.tree.setAnimated(True)
        self.tree.setIndentation(16)
        self.tree.setIconSize(QSize(16, 16))
        self.tree.setEditTriggers(QTreeView.EditTrigger.NoEditTriggers)  # Manual trigger only
        self.tree.setDragEnabled(True)
        self.tree.setAcceptDrops(True)
        self.tree.setDropIndicatorShown(True)
        self.tree.setDragDropMode(QTreeView.DragDropMode.InternalMove)
        self.tree.setSelectionMode(QTreeView.SelectionMode.ExtendedSelection)

        # Custom delegate for file type icons
        self.delegate = FileExplorerDelegate(self.model, self.tree)
        self.tree.setItemDelegate(self.delegate)

        self.tree.setStyleSheet(f"""
            QTreeView {{
                background: {Theme.BG_SECONDARY};
                border: none;
                color: {Theme.TEXT_PRIMARY};
                font-size: 12px;
                outline: none;
                show-decoration-selected: 1;
            }}
            QTreeView::item {{
                padding: 2px 6px;
                border-radius: 0;
                height: 22px;
                border-left: 1px solid transparent;
            }}
            QTreeView::item:hover {{
                background: {Theme.BG_HOVER};
                border-left: 1px solid {Theme.BORDER};
            }}
            QTreeView::item:selected {{
                background: {Theme.BG_SELECTED};
            }}
            QTreeView::branch {{
                background: transparent;
                border-left: 1px solid transparent;
            }}
            QTreeView::branch:hover {{
                border-left: 1px solid {Theme.BORDER};
            }}
            QTreeView::branch:has-siblings:!adjoins-item {{
                border-left: 1px solid {Theme.BORDER};
            }}
            QTreeView::branch:has-children:!has-siblings:closed,
            QTreeView::branch:closed:has-children:has-siblings {{
                image: url(data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTYiIGhlaWdodD0iMTYiIHZpZXdCb3g9IjAgMCAxNiAxNiIgZmlsbD0ibm9uZSIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIj48cGF0aCBkPSJNNiA0TDEwIDhMNiAxMiIgc3Ryb2tlPSIjODA4MDgwIiBzdHJva2Utd2lkdGg9IjEuNSIvPjwvc3ZnPg==);
            }}
            QTreeView::branch:open:has-children:!has-siblings,
            QTreeView::branch:open:has-children:has-siblings {{
                image: url(data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTYiIGhlaWdodD0iMTYiIHZpZXdCb3g9IjAgMCAxNiAxNiIgZmlsbD0ibm9uZSIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIj48cGF0aCBkPSJNNCAxMEw4IDZMMTIgMTAiIHN0cm9rZT0iIzgwODA4MCIgc3Ryb2tlLXdpZHRoPSIxLjUiIHRyYW5zZm9ybT0icm90YXRlKDE4MCwgOCwgOCkiLz48L3N2Zz4=);
            }}
        """)

        # Connect signals - single click for folders, double click for files
        self.tree.clicked.connect(self._on_click)
        self.tree.doubleClicked.connect(self._on_double_click)
        self.tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tree.customContextMenuRequested.connect(self._show_context_menu)

        # Enable mouse tracking for indent guides
        self.tree.setMouseTracking(True)
        self.tree.viewport().setMouseTracking(True)

        layout.addWidget(self.tree)

        # Install event filter for keyboard shortcuts and hover effects
        self.tree.installEventFilter(self)
        self.tree.viewport().installEventFilter(self)

    def eventFilter(self, obj, event):
        """Handle keyboard shortcuts."""
        if obj == self.tree and event.type() == event.Type.KeyPress:
            key = event.key()
            modifiers = event.modifiers()

            if key == Qt.Key.Key_F2:  # Rename
                self._rename_selected()
                return True
            elif key == Qt.Key.Key_Delete or key == Qt.Key.Key_Backspace:  # Delete
                self._delete_selected()
                return True
            elif key == Qt.Key.Key_Return or key == Qt.Key.Key_Enter:  # Open
                index = self.tree.currentIndex()
                if index.isValid():
                    path = Path(self.model.filePath(index))
                    if path.is_file():
                        self.file_selected.emit(path)
                    else:
                        self.tree.setExpanded(index, not self.tree.isExpanded(index))
                return True
            elif key == Qt.Key.Key_C and modifiers & Qt.KeyboardModifier.ControlModifier:  # Copy path
                self._copy_path()
                return True
            elif key == Qt.Key.Key_N and modifiers & Qt.KeyboardModifier.ControlModifier:  # New file
                self._new_file()
                return True

        return super().eventFilter(obj, event)

    def set_root(self, path: str):
        """Set the root directory."""
        self.root_path = Path(path)
        self.model.setRootPath(str(self.root_path))
        self.tree.setRootIndex(self.model.index(str(self.root_path)))

    def _on_click(self, index):
        """Handle single-click - expand/collapse folders."""
        path = Path(self.model.filePath(index))
        if path.is_dir():
            # Toggle folder expansion on single click
            self.tree.setExpanded(index, not self.tree.isExpanded(index))

    def _on_double_click(self, index):
        """Handle double-click to open file."""
        path = Path(self.model.filePath(index))
        if path.is_file():
            self.file_selected.emit(path)

    def _show_context_menu(self, pos):
        """Show VS Code-style context menu."""
        index = self.tree.indexAt(pos)
        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 4px;
            }}
            QMenu::item {{
                padding: 6px 24px 6px 12px;
                border-radius: 3px;
                color: {Theme.TEXT_PRIMARY};
            }}
            QMenu::item:selected {{
                background: {Theme.BG_SELECTED};
            }}
            QMenu::separator {{
                height: 1px;
                background: {Theme.BORDER};
                margin: 4px 8px;
            }}
        """)

        # New File/Folder
        new_file_action = menu.addAction("New File")
        new_file_action.setShortcut("Ctrl+N")
        new_file_action.triggered.connect(self._new_file)

        new_folder_action = menu.addAction("New Folder")
        new_folder_action.triggered.connect(self._new_folder)

        menu.addSeparator()

        if index.isValid():
            path = Path(self.model.filePath(index))

            # Open actions
            if path.is_file():
                open_action = menu.addAction("Open")
                open_action.triggered.connect(lambda: self.file_selected.emit(path))
                menu.addSeparator()

            # Edit actions
            rename_action = menu.addAction("Rename")
            rename_action.setShortcut("F2")
            rename_action.triggered.connect(self._rename_selected)

            delete_action = menu.addAction("Delete")
            delete_action.setShortcut("Delete")
            delete_action.triggered.connect(self._delete_selected)

            menu.addSeparator()

            # Path actions
            copy_path_action = menu.addAction("Copy Path")
            copy_path_action.setShortcut("Ctrl+C")
            copy_path_action.triggered.connect(self._copy_path)

            copy_rel_path_action = menu.addAction("Copy Relative Path")
            copy_rel_path_action.triggered.connect(self._copy_relative_path)

            menu.addSeparator()

            # System actions
            reveal_action = menu.addAction("Reveal in Finder")
            reveal_action.triggered.connect(self._reveal_in_finder)

            terminal_action = menu.addAction("Open in Terminal")
            terminal_action.triggered.connect(self._open_in_terminal)

        menu.exec(self.tree.mapToGlobal(pos))

    def _new_file(self):
        """Create a new file."""
        index = self.tree.currentIndex()
        if index.isValid():
            path = Path(self.model.filePath(index))
            parent_dir = path if path.is_dir() else path.parent
        else:
            parent_dir = self.root_path

        if not parent_dir:
            return

        # Create file atomically to avoid TOCTOU race condition
        import os as _os
        base_name = "untitled"
        counter = 0
        new_path = None

        while counter < 1000:  # Prevent infinite loop
            try:
                candidate = parent_dir / (base_name if counter == 0 else f"{base_name}_{counter}")
                # Use O_CREAT | O_EXCL for atomic creation (fails if exists)
                fd = _os.open(str(candidate), _os.O_CREAT | _os.O_EXCL | _os.O_WRONLY, 0o644)
                _os.close(fd)
                new_path = candidate
                break
            except FileExistsError:
                counter += 1
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Could not create file: {e}")
                return

        if new_path:
            self.file_created.emit(new_path)
            # Select and start editing
            new_index = self.model.index(str(new_path))
            self.tree.setCurrentIndex(new_index)
            self.tree.scrollTo(new_index)
            QTimer.singleShot(100, lambda: self.tree.edit(new_index))

    def _new_folder(self):
        """Create a new folder."""
        index = self.tree.currentIndex()
        if index.isValid():
            path = Path(self.model.filePath(index))
            parent_dir = path if path.is_dir() else path.parent
        else:
            parent_dir = self.root_path

        if not parent_dir:
            return

        # Create folder atomically to avoid TOCTOU race condition
        import os as _os
        base_name = "new_folder"
        counter = 0
        new_path = None

        while counter < 1000:  # Prevent infinite loop
            try:
                candidate = parent_dir / (base_name if counter == 0 else f"{base_name}_{counter}")
                # mkdir with exist_ok=False raises FileExistsError atomically
                candidate.mkdir(mode=0o755, exist_ok=False)
                new_path = candidate
                break
            except FileExistsError:
                counter += 1
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Could not create folder: {e}")
                return

        if new_path:
            # Select and start editing
            new_index = self.model.index(str(new_path))
            self.tree.setCurrentIndex(new_index)
            self.tree.scrollTo(new_index)
            QTimer.singleShot(100, lambda: self.tree.edit(new_index))

    def _rename_selected(self):
        """Rename selected file/folder."""
        index = self.tree.currentIndex()
        if index.isValid():
            self.tree.edit(index)

    def _delete_selected(self):
        """Delete selected files/folders."""
        indexes = self.tree.selectedIndexes()
        if not indexes:
            return

        # Get unique paths (each file has multiple column indexes)
        paths = list(set(Path(self.model.filePath(idx)) for idx in indexes))

        if len(paths) == 1:
            msg = f"Are you sure you want to delete '{paths[0].name}'?"
        else:
            msg = f"Are you sure you want to delete {len(paths)} items?"

        reply = QMessageBox.question(
            self, "Confirm Delete", msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            for path in paths:
                try:
                    if path.is_dir():
                        import shutil
                        shutil.rmtree(path)
                    else:
                        path.unlink()
                    self.file_deleted.emit(path)
                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Could not delete {path.name}: {e}")

    def _copy_path(self):
        """Copy absolute path to clipboard."""
        index = self.tree.currentIndex()
        if index.isValid():
            path = self.model.filePath(index)
            QApplication.clipboard().setText(path)

    def _copy_relative_path(self):
        """Copy relative path to clipboard."""
        index = self.tree.currentIndex()
        if index.isValid() and self.root_path:
            path = Path(self.model.filePath(index))
            try:
                rel_path = path.relative_to(self.root_path)
                QApplication.clipboard().setText(str(rel_path))
            except ValueError:
                QApplication.clipboard().setText(str(path))

    def _reveal_in_finder(self):
        """Open file/folder in system file manager."""
        index = self.tree.currentIndex()
        if index.isValid():
            path = self.model.filePath(index)
            if sys.platform == "darwin":
                subprocess.run(["open", "-R", path])
            elif sys.platform == "win32":
                subprocess.run(["explorer", "/select,", path])
            else:
                subprocess.run(["xdg-open", str(Path(path).parent)])

    def _open_in_terminal(self):
        """Open terminal at selected location."""
        index = self.tree.currentIndex()
        if index.isValid():
            path = Path(self.model.filePath(index))
            folder = path if path.is_dir() else path.parent
            if sys.platform == "darwin":
                subprocess.run(["open", "-a", "Terminal", str(folder)])
            elif sys.platform == "win32":
                subprocess.run(["cmd", "/c", "start", "cmd", "/K", f"cd /d {folder}"])
            else:
                subprocess.run(["x-terminal-emulator", "--working-directory", str(folder)])

    def _refresh(self):
        """Refresh the file tree."""
        if self.root_path:
            self.model.setRootPath("")
            self.model.setRootPath(str(self.root_path))

    def _collapse_all(self):
        """Collapse all expanded folders."""
        self.tree.collapseAll()


# ============================================================================
# Activity Bar
# ============================================================================

class ActivityBar(QWidget):
    """VS Code-style activity bar with enhanced visual feedback."""

    view_changed = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedWidth(48)
        self.setStyleSheet(f"""
            background: {Theme.BG_DARK};
            border-right: 1px solid {Theme.BORDER};
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 8, 0, 8)
        layout.setSpacing(2)

        self.buttons = {}
        self.icon_methods = {}  # Store icon method references
        icons = [
            ("files", Icons.files, "Explorer (Ctrl+Shift+E)"),
            ("search", Icons.search, "Search (Ctrl+Shift+F)"),
            ("agent", Icons.robot, "AI Agent (Ctrl+Shift+A)"),
            ("git", Icons.git_branch, "Source Control (Ctrl+Shift+G)"),
            ("settings", Icons.settings, "Settings (Ctrl+,)"),
        ]

        for name, icon_method, tooltip in icons:
            btn = QPushButton()
            btn.setFixedSize(48, 44)
            btn.setCheckable(True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setToolTip(tooltip)
            btn.setIcon(icon_method(18, Theme.TEXT_MUTED))
            btn.setIconSize(QSize(18, 18))
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: transparent;
                    border: none;
                    border-left: 3px solid transparent;
                    margin: 2px 0;
                }}
                QPushButton:hover {{
                    background: {Theme.BG_TERTIARY};
                }}
                QPushButton:checked {{
                    border-left: 3px solid {Theme.ACCENT_BLUE};
                    background: rgba(0, 122, 204, 0.1);
                }}
            """)
            btn.clicked.connect(lambda checked, n=name: self._on_click(n))
            self.buttons[name] = btn
            self.icon_methods[name] = icon_method
            layout.addWidget(btn)

        layout.addStretch()
        self.buttons["files"].setChecked(True)
        self._update_icons("files")

    def _on_click(self, name: str):
        for n, btn in self.buttons.items():
            btn.setChecked(n == name)
        self._update_icons(name)
        self.view_changed.emit(name)

    def _update_icons(self, active: str):
        """Update icon colors based on active state."""
        for name, btn in self.buttons.items():
            color = Theme.TEXT_PRIMARY if name == active else Theme.TEXT_MUTED
            btn.setIcon(self.icon_methods[name](18, color))


# ============================================================================
# Command Palette (Ctrl+Shift+P)
# ============================================================================

class CommandPalette(QDialog):
    """VS Code-style command palette with fuzzy search."""

    command_selected = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True)

        # Commands registry: (id, label, shortcut, callback_name)
        self.commands = [
            ("file.new", "New File", "Ctrl+N", "new_file"),
            ("file.open", "Open File...", "", "open_file"),
            ("file.openFolder", "Open Folder...", "Ctrl+O", "open_folder"),
            ("file.save", "Save", "Ctrl+S", "save_file"),
            ("file.saveAll", "Save All", "Ctrl+Shift+S", "save_all"),
            ("file.close", "Close Editor", "Ctrl+W", "close_tab"),
            ("edit.undo", "Undo", "Ctrl+Z", "undo"),
            ("edit.redo", "Redo", "Ctrl+Shift+Z", "redo"),
            ("edit.find", "Find", "Ctrl+F", "find"),
            ("edit.replace", "Find and Replace", "Ctrl+H", "replace"),
            ("view.explorer", "Show Explorer", "Ctrl+Shift+E", "show_explorer"),
            ("view.search", "Show Search", "Ctrl+Shift+F", "show_search"),
            ("view.terminal", "Toggle Terminal", "Ctrl+`", "toggle_terminal"),
            ("view.settings", "Open Settings", "Ctrl+,", "show_settings"),
            ("view.git", "Show Git", "Ctrl+Shift+G", "show_git"),
            ("view.splitRight", "Split Editor Right", "Ctrl+\\", "split_right"),
            ("view.splitDown", "Split Editor Down", "Ctrl+Shift+\\", "split_down"),
            ("view.closeSplit", "Close Split Editor", "", "close_split"),
            ("ai.newChat", "New AI Chat", "", "new_chat"),
            ("ai.reconnect", "Reconnect AI", "", "reconnect_ai"),
            ("terminal.clear", "Clear Terminal", "Ctrl+L", "clear_terminal"),
            ("terminal.restart", "Restart Terminal", "", "restart_terminal"),
            ("git.commit", "Git: Commit", "", "git_commit"),
            ("git.push", "Git: Push", "", "git_push"),
            ("git.pull", "Git: Pull", "", "git_pull"),
            ("git.blame", "Git: Toggle Blame", "", "toggle_blame"),
            ("editor.foldAll", "Fold All", "Ctrl+K Ctrl+0", "fold_all"),
            ("editor.unfoldAll", "Unfold All", "Ctrl+K Ctrl+J", "unfold_all"),
        ]

        self._filtered_commands = self.commands.copy()

        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Main container with shadow effect
        container = QWidget()
        container.setFixedWidth(500)
        container.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 8px;
            }}
        """)
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(0)

        # Search input
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Type a command...")
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                background: {Theme.BG_MAIN};
                color: {Theme.TEXT_PRIMARY};
                border: none;
                border-bottom: 1px solid {Theme.BORDER};
                padding: 12px 16px;
                font-size: 14px;
            }}
        """)
        self.search_input.textChanged.connect(self._filter_commands)
        self.search_input.returnPressed.connect(self._execute_selected)
        self.search_input.installEventFilter(self)
        container_layout.addWidget(self.search_input)

        # Command list
        self.command_list = QListWidget()
        self.command_list.setStyleSheet(f"""
            QListWidget {{
                background: {Theme.BG_SECONDARY};
                border: none;
                padding: 4px;
            }}
            QListWidget::item {{
                background: transparent;
                color: {Theme.TEXT_PRIMARY};
                padding: 8px 12px;
                border-radius: 4px;
            }}
            QListWidget::item:selected {{
                background: {Theme.BG_SELECTED};
            }}
            QListWidget::item:hover {{
                background: {Theme.BG_HOVER};
            }}
        """)
        self.command_list.setMaximumHeight(300)
        self.command_list.itemClicked.connect(self._on_item_clicked)
        self.command_list.itemDoubleClicked.connect(self._on_item_double_clicked)
        container_layout.addWidget(self.command_list)

        layout.addWidget(container)

        self._populate_list()

    def _populate_list(self):
        """Populate the command list."""
        self.command_list.clear()
        for cmd_id, label, shortcut, callback in self._filtered_commands:
            item = QListWidgetItem()
            # Create custom widget for command item
            widget = QWidget()
            hlayout = QHBoxLayout(widget)
            hlayout.setContentsMargins(4, 2, 4, 2)

            label_widget = QLabel(label)
            label_widget.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 13px;")
            hlayout.addWidget(label_widget)

            hlayout.addStretch()

            if shortcut:
                shortcut_label = QLabel(shortcut)
                shortcut_label.setStyleSheet(f"""
                    color: {Theme.TEXT_MUTED};
                    font-size: 11px;
                    background: {Theme.BG_TERTIARY};
                    padding: 2px 6px;
                    border-radius: 3px;
                """)
                hlayout.addWidget(shortcut_label)

            item.setSizeHint(widget.sizeHint())
            item.setData(Qt.ItemDataRole.UserRole, callback)
            self.command_list.addItem(item)
            self.command_list.setItemWidget(item, widget)

        if self.command_list.count() > 0:
            self.command_list.setCurrentRow(0)

    def _filter_commands(self, text: str):
        """Filter commands based on search text."""
        if not text:
            self._filtered_commands = self.commands.copy()
        else:
            text_lower = text.lower()
            self._filtered_commands = [
                cmd for cmd in self.commands
                if text_lower in cmd[1].lower() or text_lower in cmd[0].lower()
            ]
        self._populate_list()

    def _execute_selected(self):
        """Execute the currently selected command."""
        current = self.command_list.currentItem()
        if current:
            callback = current.data(Qt.ItemDataRole.UserRole)
            self.command_selected.emit(callback)
            self.accept()

    def _on_item_clicked(self, item):
        """Handle item click."""
        self.command_list.setCurrentItem(item)

    def _on_item_double_clicked(self, item):
        """Handle item double-click - execute command."""
        callback = item.data(Qt.ItemDataRole.UserRole)
        self.command_selected.emit(callback)
        self.accept()

    def eventFilter(self, obj, event):
        """Handle keyboard navigation."""
        if obj == self.search_input and event.type() == event.Type.KeyPress:
            key = event.key()
            if key == Qt.Key.Key_Down:
                current = self.command_list.currentRow()
                if current < self.command_list.count() - 1:
                    self.command_list.setCurrentRow(current + 1)
                return True
            elif key == Qt.Key.Key_Up:
                current = self.command_list.currentRow()
                if current > 0:
                    self.command_list.setCurrentRow(current - 1)
                return True
            elif key == Qt.Key.Key_Escape:
                self.reject()
                return True
        return super().eventFilter(obj, event)

    def showEvent(self, event):
        """Center the palette and focus input."""
        super().showEvent(event)
        self.search_input.clear()
        self._filtered_commands = self.commands.copy()
        self._populate_list()
        self.search_input.setFocus()

        # Center on parent
        if self.parent():
            parent_rect = self.parent().geometry()
            self.move(
                parent_rect.center().x() - self.width() // 2,
                parent_rect.top() + 100
            )


# ============================================================================
# Quick Open (Ctrl+P)
# ============================================================================

class QuickOpenDialog(QDialog):
    """VS Code-style quick file open with fuzzy search."""

    file_selected = Signal(Path)

    def __init__(self, root_path: Path = None, parent=None):
        super().__init__(parent)
        self.root_path = root_path or Path.home()
        self._files: List[Path] = []
        self._filtered_files: List[Path] = []

        self.setWindowFlags(Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True)

        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Main container
        container = QWidget()
        container.setFixedWidth(600)
        container.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 8px;
            }}
        """)
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(0)

        # Search input
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search files by name...")
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                background: {Theme.BG_MAIN};
                color: {Theme.TEXT_PRIMARY};
                border: none;
                border-bottom: 1px solid {Theme.BORDER};
                padding: 12px 16px;
                font-size: 14px;
            }}
        """)
        self.search_input.textChanged.connect(self._filter_files)
        self.search_input.returnPressed.connect(self._open_selected)
        self.search_input.installEventFilter(self)
        container_layout.addWidget(self.search_input)

        # File list
        self.file_list = QListWidget()
        self.file_list.setStyleSheet(f"""
            QListWidget {{
                background: {Theme.BG_SECONDARY};
                border: none;
                padding: 4px;
            }}
            QListWidget::item {{
                background: transparent;
                color: {Theme.TEXT_PRIMARY};
                padding: 6px 12px;
                border-radius: 4px;
            }}
            QListWidget::item:selected {{
                background: {Theme.BG_SELECTED};
            }}
            QListWidget::item:hover {{
                background: {Theme.BG_HOVER};
            }}
        """)
        self.file_list.setMaximumHeight(400)
        self.file_list.itemDoubleClicked.connect(self._on_item_double_clicked)
        container_layout.addWidget(self.file_list)

        layout.addWidget(container)

    def set_root(self, path: str):
        """Set root path and scan for files."""
        self.root_path = Path(path)
        self._scan_files()

    def _scan_files(self):
        """Scan directory for files (runs in background)."""
        self._files = []
        try:
            # Walk directory but limit depth and exclude common folders
            exclude_dirs = {'.git', 'node_modules', '__pycache__', '.venv', 'venv', '.idea', '.vscode', 'dist', 'build'}
            for root, dirs, files in os.walk(self.root_path):
                # Filter out excluded directories
                dirs[:] = [d for d in dirs if d not in exclude_dirs and not d.startswith('.')]

                # Limit depth
                depth = len(Path(root).relative_to(self.root_path).parts)
                if depth > 10:
                    continue

                for f in files:
                    if not f.startswith('.'):
                        self._files.append(Path(root) / f)

                # Limit total files
                if len(self._files) > 5000:
                    break

        except Exception:
            pass

        self._filtered_files = self._files[:100]  # Show first 100 by default
        self._populate_list()

    def _populate_list(self):
        """Populate file list."""
        self.file_list.clear()
        for path in self._filtered_files[:50]:  # Limit display to 50
            try:
                rel_path = path.relative_to(self.root_path)
            except ValueError:
                rel_path = path

            item = QListWidgetItem()
            widget = QWidget()
            hlayout = QHBoxLayout(widget)
            hlayout.setContentsMargins(4, 2, 4, 2)

            # File icon
            icon_label = QLabel()
            icon_label.setPixmap(Icons.file(14, Theme.TEXT_MUTED).pixmap(14, 14))
            hlayout.addWidget(icon_label)

            # File name
            name_label = QLabel(path.name)
            name_label.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 13px;")
            hlayout.addWidget(name_label)

            hlayout.addStretch()

            # Relative path
            parent_path = str(rel_path.parent)
            if parent_path != '.':
                path_label = QLabel(parent_path)
                path_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
                hlayout.addWidget(path_label)

            item.setSizeHint(widget.sizeHint())
            item.setData(Qt.ItemDataRole.UserRole, path)
            self.file_list.addItem(item)
            self.file_list.setItemWidget(item, widget)

        if self.file_list.count() > 0:
            self.file_list.setCurrentRow(0)

    def _filter_files(self, text: str):
        """Filter files based on search text."""
        if not text:
            self._filtered_files = self._files[:100]
        else:
            text_lower = text.lower()
            self._filtered_files = [
                f for f in self._files
                if text_lower in f.name.lower() or text_lower in str(f).lower()
            ][:100]
        self._populate_list()

    def _open_selected(self):
        """Open the currently selected file."""
        current = self.file_list.currentItem()
        if current:
            path = current.data(Qt.ItemDataRole.UserRole)
            self.file_selected.emit(path)
            self.accept()

    def _on_item_double_clicked(self, item):
        """Handle double-click - open file."""
        path = item.data(Qt.ItemDataRole.UserRole)
        self.file_selected.emit(path)
        self.accept()

    def eventFilter(self, obj, event):
        """Handle keyboard navigation."""
        if obj == self.search_input and event.type() == event.Type.KeyPress:
            key = event.key()
            if key == Qt.Key.Key_Down:
                current = self.file_list.currentRow()
                if current < self.file_list.count() - 1:
                    self.file_list.setCurrentRow(current + 1)
                return True
            elif key == Qt.Key.Key_Up:
                current = self.file_list.currentRow()
                if current > 0:
                    self.file_list.setCurrentRow(current - 1)
                return True
            elif key == Qt.Key.Key_Escape:
                self.reject()
                return True
        return super().eventFilter(obj, event)

    def showEvent(self, event):
        """Setup on show."""
        super().showEvent(event)
        self.search_input.clear()
        self._filter_files("")
        self.search_input.setFocus()

        # Center on parent
        if self.parent():
            parent_rect = self.parent().geometry()
            self.move(
                parent_rect.center().x() - self.width() // 2,
                parent_rect.top() + 80
            )


# ============================================================================
# Search Panel
# ============================================================================

class SearchPanel(QWidget):
    """Search panel with system-wide search."""

    file_selected = Signal(Path)

    def __init__(self, root_path: str = None, parent=None):
        super().__init__(parent)
        self.root_path = Path(root_path) if root_path else Path.home()
        self.permissions = SearchPermissions()

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        header = QLabel("  SEARCH")
        header.setStyleSheet(f"""
            color: {Theme.TEXT_SECONDARY};
            font-size: 10px;
            font-weight: 600;
            letter-spacing: 0.5px;
            padding: 8px 0;
            background: {Theme.BG_SECONDARY};
        """)
        layout.addWidget(header)

        # Search inputs
        input_container = QWidget()
        input_container.setStyleSheet(f"background: {Theme.BG_SECONDARY};")
        input_layout = QVBoxLayout(input_container)
        input_layout.setContentsMargins(8, 4, 8, 8)
        input_layout.setSpacing(6)

        self.search_input = CompactLineEdit("Search")
        self.search_input.returnPressed.connect(self._do_search)
        input_layout.addWidget(self.search_input)

        self.filter_input = CompactLineEdit("Files to include (e.g. *.py)")
        input_layout.addWidget(self.filter_input)

        # Scope
        scope_row = QHBoxLayout()
        self.scope_project = QRadioButton("Workspace")
        self.scope_project.setChecked(True)
        self.scope_project.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 11px;")
        scope_row.addWidget(self.scope_project)

        self.scope_system = QRadioButton("Directories")
        self.scope_system.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 11px;")
        scope_row.addWidget(self.scope_system)
        scope_row.addStretch()
        input_layout.addLayout(scope_row)

        layout.addWidget(input_container)

        # Allowed directories
        self.dirs_container = QWidget()
        self.dirs_container.setStyleSheet(f"background: {Theme.BG_SECONDARY};")
        dirs_layout = QVBoxLayout(self.dirs_container)
        dirs_layout.setContentsMargins(8, 0, 8, 8)
        dirs_layout.setSpacing(4)

        dirs_header = QHBoxLayout()
        dirs_label = QLabel("Search Directories:")
        dirs_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 10px;")
        dirs_header.addWidget(dirs_label)
        dirs_header.addStretch()

        add_dir_btn = IconButton(Icons.plus, "Add Directory")
        add_dir_btn.clicked.connect(self._add_directory)
        dirs_header.addWidget(add_dir_btn)
        dirs_layout.addLayout(dirs_header)

        self.dirs_list = QListWidget()
        self.dirs_list.setMaximumHeight(80)
        self.dirs_list.setStyleSheet(f"""
            QListWidget {{
                background: {Theme.BG_TERTIARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 3px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 10px;
            }}
            QListWidget::item {{ padding: 3px 6px; }}
            QListWidget::item:hover {{ background: {Theme.BG_HOVER}; }}
        """)
        dirs_layout.addWidget(self.dirs_list)
        layout.addWidget(self.dirs_container)
        self.dirs_container.hide()

        self.scope_system.toggled.connect(lambda c: self.dirs_container.setVisible(c))
        self.scope_system.toggled.connect(self._refresh_dirs_list)

        # Results
        self.results_list = QListWidget()
        self.results_list.setStyleSheet(f"""
            QListWidget {{
                background: {Theme.BG_SECONDARY};
                border: none;
                color: {Theme.TEXT_PRIMARY};
                font-size: 11px;
            }}
            QListWidget::item {{
                padding: 4px 8px;
                border-bottom: 1px solid {Theme.BORDER};
            }}
            QListWidget::item:hover {{ background: {Theme.BG_HOVER}; }}
            QListWidget::item:selected {{ background: {Theme.BG_SELECTED}; }}
        """)
        self.results_list.itemDoubleClicked.connect(self._on_result_clicked)
        layout.addWidget(self.results_list)

        self.status_label = QLabel("")
        self.status_label.setStyleSheet(f"""
            color: {Theme.TEXT_MUTED};
            font-size: 10px;
            padding: 4px 8px;
            background: {Theme.BG_SECONDARY};
        """)
        layout.addWidget(self.status_label)

    def set_root(self, path: str):
        self.root_path = Path(path)

    def _refresh_dirs_list(self):
        self.dirs_list.clear()
        for d in self.permissions.allowed_dirs:
            self.dirs_list.addItem(str(d))

    def _add_directory(self):
        path = QFileDialog.getExistingDirectory(self, "Select Directory", str(Path.home()))
        if path:
            self.permissions.grant_access(Path(path))
            self._refresh_dirs_list()

    def _do_search(self):
        query = self.search_input.text().strip()
        if not query:
            return

        self.results_list.clear()
        file_filter = self.filter_input.text().strip() or "*"

        search_dirs = []
        if self.scope_project.isChecked():
            search_dirs = [self.root_path]
        else:
            search_dirs = list(self.permissions.allowed_dirs)
            if not search_dirs:
                self.status_label.setText("No directories added")
                return

        count = 0
        for search_dir in search_dirs:
            try:
                cmd = ["grep", "-rn", f"--include={file_filter}", query, str(search_dir)]
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
                lines = result.stdout.strip().split('\n') if result.stdout.strip() else []

                for line in lines[:50]:
                    if ':' in line:
                        parts = line.split(':', 2)
                        if len(parts) >= 3:
                            filepath, linenum, content = parts[0], parts[1], parts[2]
                            rel_path = Path(filepath)
                            try:
                                rel_path = rel_path.relative_to(search_dir)
                            except ValueError:
                                pass
                            item = QListWidgetItem(f"{rel_path}:{linenum}")
                            item.setToolTip(content.strip()[:100])
                            item.setData(Qt.ItemDataRole.UserRole, filepath)
                            self.results_list.addItem(item)
                            count += 1
            except subprocess.TimeoutExpired:
                self.status_label.setText("Search timed out")
                return
            except FileNotFoundError:
                self.status_label.setText("grep not found - install grep")
                return
            except Exception as e:
                logger.warning(f"Search error: {e}")
                self.status_label.setText(f"Search failed: {str(e)[:30]}")
                return

        self.status_label.setText(f"{count} results")

    def _on_result_clicked(self, item):
        filepath = item.data(Qt.ItemDataRole.UserRole)
        if filepath:
            self.file_selected.emit(Path(filepath))


# ============================================================================
# Git Panel
# ============================================================================

class DiffViewer(QDialog):
    """Diff viewer dialog."""

    def __init__(self, file_path: str, diff_text: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Diff: {Path(file_path).name}")
        self.setModal(True)
        self.setMinimumSize(700, 500)
        self.setStyleSheet(f"QDialog {{ background: {Theme.BG_SECONDARY}; }}")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        header = QLabel(file_path)
        header.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 12px; font-weight: 500;")
        layout.addWidget(header)

        diff_view = QPlainTextEdit()
        diff_view.setReadOnly(True)
        diff_view.setPlainText(diff_text)
        diff_view.setStyleSheet(f"""
            QPlainTextEdit {{
                background: {Theme.BG_DARK};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 3px;
                padding: 8px;
                font-family: 'Consolas', 'Monaco', monospace;
                font-size: 11px;
            }}
        """)
        layout.addWidget(diff_view)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        close_btn = SecondaryButton("Close")
        close_btn.clicked.connect(self.close)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)


class SettingRow(QWidget):
    """A single setting row with label, description, and control."""

    def __init__(self, label: str, description: str = "", parent=None):
        super().__init__(parent)
        self.setStyleSheet("background: transparent;")

        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 12, 0, 12)
        layout.setSpacing(16)

        # Left side: label and description
        text_layout = QVBoxLayout()
        text_layout.setSpacing(2)

        self.label = QLabel(label)
        self.label.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 13px;")
        text_layout.addWidget(self.label)

        if description:
            self.desc = QLabel(description)
            self.desc.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
            self.desc.setWordWrap(True)
            text_layout.addWidget(self.desc)

        layout.addLayout(text_layout, 1)

        # Right side: control placeholder
        self.control_layout = QHBoxLayout()
        self.control_layout.setSpacing(8)
        layout.addLayout(self.control_layout)

    def add_control(self, widget):
        """Add a control widget to the right side."""
        self.control_layout.addWidget(widget)


class ToggleSwitch(QPushButton):
    """iOS-style toggle switch."""

    toggled_signal = Signal(bool)

    def __init__(self, checked: bool = False, parent=None):
        super().__init__(parent)
        self._checked = checked
        self.setCheckable(True)
        self.setChecked(checked)
        self.setFixedSize(44, 24)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.clicked.connect(self._on_click)
        self._update_style()

    def _on_click(self):
        self._checked = self.isChecked()
        self._update_style()
        self.toggled_signal.emit(self._checked)

    def _update_style(self):
        if self._checked:
            self.setStyleSheet(f"""
                QPushButton {{
                    background: {Theme.ACCENT_BLUE};
                    border-radius: 12px;
                    border: none;
                }}
            """)
        else:
            self.setStyleSheet(f"""
                QPushButton {{
                    background: {Theme.BG_TERTIARY};
                    border-radius: 12px;
                    border: 1px solid {Theme.BORDER};
                }}
            """)

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # Draw the knob
        knob_size = 18
        knob_y = 3
        if self._checked:
            knob_x = self.width() - knob_size - 3
            painter.setBrush(QBrush(QColor("white")))
        else:
            knob_x = 3
            painter.setBrush(QBrush(QColor(Theme.TEXT_MUTED)))

        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawEllipse(knob_x, knob_y, knob_size, knob_size)


class AgentControlPanel(QWidget):
    """AI Agent session panel - shows live status, stats, and quick actions."""

    reconnect_requested = Signal()
    clear_chat_requested = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"background: {Theme.BG_MAIN};")
        self._start_time = None
        self._message_count = 0

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Header
        header = QWidget()
        header.setFixedHeight(40)
        header.setStyleSheet(f"background: {Theme.BG_SECONDARY}; border-bottom: 1px solid {Theme.BORDER};")
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(16, 0, 16, 0)

        header_icon = QLabel()
        header_icon.setPixmap(Icons.robot(16, Theme.ACCENT_BLUE).pixmap(16, 16))
        header_layout.addWidget(header_icon)

        header_title = QLabel("Session")
        header_title.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 12px; font-weight: 600;")
        header_layout.addWidget(header_title)
        header_layout.addStretch()

        layout.addWidget(header)

        # Scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setStyleSheet(f"""
            QScrollArea {{ background: {Theme.BG_MAIN}; border: none; }}
            QScrollBar:vertical {{ background: {Theme.BG_MAIN}; width: 10px; }}
            QScrollBar::handle:vertical {{ background: {Theme.BG_TERTIARY}; border-radius: 5px; min-height: 30px; }}
            QScrollBar::handle:vertical:hover {{ background: {Theme.BG_HOVER}; }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
        """)

        content = QWidget()
        content.setStyleSheet(f"background: {Theme.BG_MAIN};")
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(16, 16, 16, 16)
        content_layout.setSpacing(16)

        # =================================================================
        # Connection Status Card
        # =================================================================
        status_card = QWidget()
        status_card.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
            }}
        """)
        status_layout = QVBoxLayout(status_card)
        status_layout.setContentsMargins(16, 16, 16, 16)
        status_layout.setSpacing(12)

        # Status banner
        self.status_banner = QWidget()
        self.status_banner.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_TERTIARY};
                border-radius: 6px;
                border: none;
            }}
        """)
        banner_layout = QHBoxLayout(self.status_banner)
        banner_layout.setContentsMargins(12, 10, 12, 10)

        self.status_dot = StatusDot(Theme.TEXT_MUTED, 10)
        banner_layout.addWidget(self.status_dot)

        status_text = QVBoxLayout()
        status_text.setSpacing(2)
        self.status_label = QLabel("Not Connected")
        self.status_label.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 13px; font-weight: 600; border: none;")
        status_text.addWidget(self.status_label)

        self.project_label = QLabel("Open a project to connect")
        self.project_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px; border: none;")
        status_text.addWidget(self.project_label)
        banner_layout.addLayout(status_text, 1)

        self.reconnect_btn = QPushButton("Connect")
        self.reconnect_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.reconnect_btn.setFixedHeight(28)
        self.reconnect_btn.clicked.connect(self.reconnect_requested.emit)
        self.reconnect_btn.setStyleSheet(f"""
            QPushButton {{
                background: {Theme.ACCENT_BLUE};
                color: white;
                border: none;
                border-radius: 4px;
                padding: 0 16px;
                font-size: 11px;
                font-weight: 500;
            }}
            QPushButton:hover {{ background: #3b82f6; }}
        """)
        banner_layout.addWidget(self.reconnect_btn)

        status_layout.addWidget(self.status_banner)

        # Current model display (read-only)
        model_row = QHBoxLayout()
        model_label = QLabel("Model:")
        model_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px; border: none;")
        model_row.addWidget(model_label)

        self.current_model_label = QLabel("Not selected")
        self.current_model_label.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 11px; border: none;")
        model_row.addWidget(self.current_model_label)
        model_row.addStretch()

        status_layout.addLayout(model_row)

        content_layout.addWidget(status_card)

        # =================================================================
        # Session Statistics Card
        # =================================================================
        stats_card = QWidget()
        stats_card.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
            }}
        """)
        stats_layout = QVBoxLayout(stats_card)
        stats_layout.setContentsMargins(16, 12, 16, 12)
        stats_layout.setSpacing(12)

        stats_title = QLabel("Statistics")
        stats_title.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 12px; font-weight: 600; border: none;")
        stats_layout.addWidget(stats_title)

        # Stats grid
        stats_grid = QWidget()
        stats_grid.setStyleSheet("background: transparent; border: none;")
        grid_layout = QGridLayout(stats_grid)
        grid_layout.setContentsMargins(0, 0, 0, 0)
        grid_layout.setSpacing(12)

        # Tokens
        tokens_box, self.tokens_label = self._create_stat_box("0", "Tokens", Theme.ACCENT_BLUE)
        grid_layout.addWidget(tokens_box, 0, 0)

        # Cost
        cost_box, self.cost_label = self._create_stat_box("$0.00", "Cost", Theme.SUCCESS)
        grid_layout.addWidget(cost_box, 0, 1)

        # Messages
        messages_box, self.messages_label = self._create_stat_box("0", "Messages", Theme.ACCENT_PURPLE)
        grid_layout.addWidget(messages_box, 1, 0)

        # Duration
        duration_box, self.duration_label = self._create_stat_box("0:00", "Duration", Theme.ACCENT_ORANGE)
        grid_layout.addWidget(duration_box, 1, 1)

        stats_layout.addWidget(stats_grid)
        content_layout.addWidget(stats_card)

        # =================================================================
        # Quick Actions Card
        # =================================================================
        actions_card = QWidget()
        actions_card.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
            }}
        """)
        actions_layout = QVBoxLayout(actions_card)
        actions_layout.setContentsMargins(16, 12, 16, 12)
        actions_layout.setSpacing(12)

        actions_title = QLabel("Quick Actions")
        actions_title.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 12px; font-weight: 600; border: none;")
        actions_layout.addWidget(actions_title)

        # Action buttons
        btn_container = QWidget()
        btn_container.setStyleSheet("background: transparent; border: none;")
        btn_layout = QVBoxLayout(btn_container)
        btn_layout.setContentsMargins(0, 0, 0, 0)
        btn_layout.setSpacing(8)

        self.clear_btn = self._create_action_button("New Conversation", "Start fresh with a new conversation")
        self.clear_btn.clicked.connect(self.clear_chat_requested.emit)
        btn_layout.addWidget(self.clear_btn)

        self.settings_btn = self._create_action_button("Open Settings", "Configure AI provider and preferences")
        btn_layout.addWidget(self.settings_btn)

        actions_layout.addWidget(btn_container)
        content_layout.addWidget(actions_card)

        # =================================================================
        # Tips Card
        # =================================================================
        tips_card = QWidget()
        tips_card.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
            }}
        """)
        tips_layout = QVBoxLayout(tips_card)
        tips_layout.setContentsMargins(16, 12, 16, 12)
        tips_layout.setSpacing(8)

        tips_title = QLabel("Tips")
        tips_title.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 12px; font-weight: 600; border: none;")
        tips_layout.addWidget(tips_title)

        tips = [
            "Use Cmd+Enter to send messages",
            "Be specific about file paths and requirements",
            "Ask the AI to explain its changes",
            "Review suggested edits before accepting",
        ]

        for tip in tips:
            tip_row = QHBoxLayout()
            tip_row.setContentsMargins(0, 2, 0, 2)

            bullet = QLabel("-")
            bullet.setStyleSheet(f"color: {Theme.ACCENT_BLUE}; font-size: 11px; border: none;")
            bullet.setFixedWidth(12)
            tip_row.addWidget(bullet)

            tip_label = QLabel(tip)
            tip_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px; border: none;")
            tip_label.setWordWrap(True)
            tip_row.addWidget(tip_label, 1)

            tips_layout.addLayout(tip_row)

        content_layout.addWidget(tips_card)

        content_layout.addStretch()
        scroll.setWidget(content)
        layout.addWidget(scroll)

        # Timer for duration updates
        self._duration_timer = QTimer(self)
        self._duration_timer.timeout.connect(self._update_duration)

    def _create_stat_box(self, value: str, label: str, color: str) -> tuple:
        """Create a statistics display box. Returns (box_widget, value_label)."""
        box = QWidget()
        box.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_DARK};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
            }}
        """)
        box.setFixedHeight(60)
        layout = QVBoxLayout(box)
        layout.setContentsMargins(12, 8, 12, 8)
        layout.setSpacing(2)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        value_label = QLabel(value)
        value_label.setStyleSheet(f"color: {color}; font-size: 18px; font-weight: 600; border: none;")
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(value_label)

        desc_label = QLabel(label)
        desc_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 10px; border: none;")
        desc_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(desc_label)

        return box, value_label

    def _create_action_button(self, text: str, description: str) -> QPushButton:
        """Create an action button with description."""
        btn = QPushButton(text)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.setFixedHeight(36)
        btn.setToolTip(description)
        btn.setStyleSheet(f"""
            QPushButton {{
                background: {Theme.BG_TERTIARY};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 0 16px;
                font-size: 12px;
                text-align: left;
            }}
            QPushButton:hover {{ background: {Theme.BG_HOVER}; }}
        """)
        return btn

    def _update_duration(self):
        """Update session duration display."""
        if self._start_time and self.duration_label:
            elapsed = time.time() - self._start_time
            mins = int(elapsed // 60)
            secs = int(elapsed % 60)
            self.duration_label.setText(f"{mins}:{secs:02d}")

    def set_connected(self, connected: bool, project: str = None):
        """Update connection status display."""
        if connected:
            self.status_label.setText("Connected")
            self.status_label.setStyleSheet(f"color: {Theme.SUCCESS}; font-size: 13px; font-weight: 600; border: none;")
            self.status_dot.set_color(Theme.SUCCESS)
            self.reconnect_btn.setText("Reconnect")
            self.status_banner.setStyleSheet(f"""
                QWidget {{
                    background: rgba(35, 134, 54, 0.1);
                    border: 1px solid rgba(35, 134, 54, 0.3);
                    border-radius: 6px;
                }}
            """)
            # Start session timer
            if not self._start_time:
                self._start_time = time.time()
                self._duration_timer.start(1000)
        else:
            self.status_label.setText("Not Connected")
            self.status_label.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 13px; font-weight: 600; border: none;")
            self.status_dot.set_color(Theme.TEXT_MUTED)
            self.reconnect_btn.setText("Connect")
            self.status_banner.setStyleSheet(f"""
                QWidget {{
                    background: {Theme.BG_TERTIARY};
                    border-radius: 6px;
                    border: none;
                }}
            """)

        if project:
            self.project_label.setText(project)
            self.project_label.setStyleSheet(f"color: {Theme.TEXT_SECONDARY}; font-size: 11px; border: none;")
        else:
            self.project_label.setText("Open a project to connect")
            self.project_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px; border: none;")

    def set_model(self, model: str):
        """Update the current model display."""
        self.current_model_label.setText(model)

    def update_stats(self, tokens: int, cost: float):
        """Update token and cost display."""
        if self.tokens_label:
            self.tokens_label.setText(f"{tokens:,}")
        if self.cost_label:
            self.cost_label.setText(f"${cost:.4f}")

    def increment_messages(self):
        """Increment the message counter."""
        self._message_count += 1
        if self.messages_label:
            self.messages_label.setText(str(self._message_count))

    def reset_session(self):
        """Reset session statistics."""
        self._start_time = time.time()
        self._message_count = 0
        if self.tokens_label:
            self.tokens_label.setText("0")
        if self.cost_label:
            self.cost_label.setText("$0.00")
        if self.messages_label:
            self.messages_label.setText("0")
        if self.duration_label:
            self.duration_label.setText("0:00")


class GitPanel(QWidget):
    """VS Code-style Git source control panel."""

    # Signal emitted when git operations (commit, checkout, etc.) complete
    git_operation_completed = Signal()

    def __init__(self, root_path: str = None, parent=None):
        super().__init__(parent)
        self.root_path = Path(root_path) if root_path else Path.cwd()
        self.setStyleSheet(f"background: {Theme.BG_MAIN};")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Header with branch selector and refresh
        header = QWidget()
        header.setFixedHeight(40)
        header.setStyleSheet(f"background: {Theme.BG_SECONDARY}; border-bottom: 1px solid {Theme.BORDER};")
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(12, 0, 12, 0)

        # Git icon and branch
        git_icon = QLabel()
        git_icon.setPixmap(Icons.git(16, Theme.ACCENT_ORANGE).pixmap(16, 16))
        header_layout.addWidget(git_icon)

        self.branch_combo = QComboBox()
        self.branch_combo.setFixedWidth(140)
        self.branch_combo.setStyleSheet(f"""
            QComboBox {{
                background: {Theme.BG_TERTIARY};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 11px;
            }}
            QComboBox:hover {{ border-color: {Theme.ACCENT_BLUE}; }}
            QComboBox::drop-down {{ border: none; width: 20px; }}
            QComboBox QAbstractItemView {{
                background: {Theme.BG_SECONDARY};
                color: {Theme.TEXT_PRIMARY};
                selection-background-color: {Theme.BG_SELECTED};
                border: 1px solid {Theme.BORDER};
            }}
        """)
        header_layout.addWidget(self.branch_combo)

        header_layout.addStretch()

        refresh_btn = QPushButton()
        refresh_btn.setIcon(Icons.refresh(16, Theme.TEXT_MUTED))
        refresh_btn.setFixedSize(28, 28)
        refresh_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        refresh_btn.setToolTip("Refresh")
        refresh_btn.setStyleSheet(f"""
            QPushButton {{ background: transparent; border: none; border-radius: 4px; }}
            QPushButton:hover {{ background: {Theme.BG_HOVER}; }}
        """)
        refresh_btn.clicked.connect(self.refresh)
        header_layout.addWidget(refresh_btn)

        layout.addWidget(header)

        # Scroll area for content
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setStyleSheet(f"""
            QScrollArea {{ background: {Theme.BG_MAIN}; border: none; }}
            QScrollBar:vertical {{ background: {Theme.BG_MAIN}; width: 10px; }}
            QScrollBar::handle:vertical {{ background: {Theme.BG_TERTIARY}; border-radius: 5px; min-height: 30px; }}
            QScrollBar::handle:vertical:hover {{ background: {Theme.BG_HOVER}; }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
        """)

        content = QWidget()
        content.setStyleSheet(f"background: {Theme.BG_MAIN};")
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(12, 12, 12, 12)
        content_layout.setSpacing(16)

        # Commit Card
        commit_card = QWidget()
        commit_card.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
            }}
        """)
        commit_layout = QVBoxLayout(commit_card)
        commit_layout.setContentsMargins(12, 12, 12, 12)
        commit_layout.setSpacing(10)

        commit_title = QLabel("Commit")
        commit_title.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 12px; font-weight: 600; border: none;")
        commit_layout.addWidget(commit_title)

        # Commit message input
        self.commit_input = QPlainTextEdit()
        self.commit_input.setPlaceholderText("Enter commit message...")
        self.commit_input.setFixedHeight(60)
        self.commit_input.setStyleSheet(f"""
            QPlainTextEdit {{
                background: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 8px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 12px;
            }}
            QPlainTextEdit:focus {{ border-color: {Theme.ACCENT_BLUE}; }}
        """)
        commit_layout.addWidget(self.commit_input)

        # Commit buttons row
        commit_btns = QHBoxLayout()
        commit_btns.setSpacing(8)

        commit_btn = QPushButton("Commit")
        commit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        commit_btn.setFixedHeight(28)
        commit_btn.clicked.connect(self._do_commit)
        commit_btn.setStyleSheet(f"""
            QPushButton {{
                background: {Theme.ACCENT_BLUE};
                color: white;
                border: none;
                border-radius: 4px;
                padding: 0 16px;
                font-size: 11px;
                font-weight: 500;
            }}
            QPushButton:hover {{ background: #3b82f6; }}
            QPushButton:pressed {{ background: #1d4ed8; }}
        """)
        commit_btns.addWidget(commit_btn)

        stage_btn = QPushButton("Stage All")
        stage_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        stage_btn.setFixedHeight(28)
        stage_btn.clicked.connect(self._stage_all)
        stage_btn.setStyleSheet(f"""
            QPushButton {{
                background: {Theme.BG_TERTIARY};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 0 12px;
                font-size: 11px;
            }}
            QPushButton:hover {{ background: {Theme.BG_HOVER}; }}
        """)
        commit_btns.addWidget(stage_btn)
        commit_btns.addStretch()

        commit_layout.addLayout(commit_btns)
        content_layout.addWidget(commit_card)

        # Sync Card
        sync_card = QWidget()
        sync_card.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
            }}
        """)
        sync_layout = QVBoxLayout(sync_card)
        sync_layout.setContentsMargins(12, 12, 12, 12)
        sync_layout.setSpacing(10)

        sync_title = QLabel("Sync")
        sync_title.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 12px; font-weight: 600; border: none;")
        sync_layout.addWidget(sync_title)

        sync_btns = QHBoxLayout()
        sync_btns.setSpacing(8)

        pull_btn = QPushButton("Pull")
        pull_btn.setIcon(Icons.arrow_down(14, Theme.TEXT_PRIMARY))
        pull_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        pull_btn.setFixedHeight(32)
        pull_btn.clicked.connect(self._do_pull)
        pull_btn.setStyleSheet(f"""
            QPushButton {{
                background: {Theme.BG_TERTIARY};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 0 16px;
                font-size: 11px;
            }}
            QPushButton:hover {{ background: {Theme.BG_HOVER}; }}
        """)
        sync_btns.addWidget(pull_btn)

        push_btn = QPushButton("Push")
        push_btn.setIcon(Icons.arrow_up(14, Theme.TEXT_PRIMARY))
        push_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        push_btn.setFixedHeight(32)
        push_btn.clicked.connect(self._do_push)
        push_btn.setStyleSheet(pull_btn.styleSheet())
        sync_btns.addWidget(push_btn)

        sync_btns.addStretch()
        sync_layout.addLayout(sync_btns)
        content_layout.addWidget(sync_card)

        # Changes Card
        changes_card = QWidget()
        changes_card.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
            }}
        """)
        changes_layout = QVBoxLayout(changes_card)
        changes_layout.setContentsMargins(12, 12, 12, 12)
        changes_layout.setSpacing(8)

        changes_header = QHBoxLayout()
        changes_title = QLabel("Changes")
        changes_title.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 12px; font-weight: 600; border: none;")
        changes_header.addWidget(changes_title)

        self.changes_count = QLabel("0")
        self.changes_count.setStyleSheet(f"""
            background: {Theme.ACCENT_BLUE};
            color: white;
            font-size: 10px;
            padding: 2px 6px;
            border-radius: 8px;
            font-weight: 500;
        """)
        changes_header.addWidget(self.changes_count)
        changes_header.addStretch()
        changes_layout.addLayout(changes_header)

        self.changes_list = QListWidget()
        self.changes_list.setMinimumHeight(80)
        self.changes_list.setMaximumHeight(150)
        self.changes_list.setStyleSheet(f"""
            QListWidget {{
                background: {Theme.BG_DARK};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 11px;
                font-family: Menlo, monospace;
            }}
            QListWidget::item {{ padding: 4px 8px; }}
            QListWidget::item:hover {{ background: {Theme.BG_HOVER}; }}
            QListWidget::item:selected {{ background: {Theme.BG_SELECTED}; }}
        """)
        self.changes_list.itemDoubleClicked.connect(self._view_diff)
        changes_layout.addWidget(self.changes_list)

        content_layout.addWidget(changes_card)

        # History Card
        history_card = QWidget()
        history_card.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
            }}
        """)
        history_layout = QVBoxLayout(history_card)
        history_layout.setContentsMargins(12, 12, 12, 12)
        history_layout.setSpacing(8)

        history_title = QLabel("Recent Commits")
        history_title.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 12px; font-weight: 600; border: none;")
        history_layout.addWidget(history_title)

        self.history_list = QListWidget()
        self.history_list.setMinimumHeight(60)
        self.history_list.setMaximumHeight(120)
        self.history_list.setStyleSheet(f"""
            QListWidget {{
                background: {Theme.BG_DARK};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 11px;
                font-family: Menlo, monospace;
            }}
            QListWidget::item {{ padding: 4px 8px; }}
            QListWidget::item:hover {{ background: {Theme.BG_HOVER}; }}
        """)
        history_layout.addWidget(self.history_list)

        content_layout.addWidget(history_card)

        content_layout.addStretch()

        # Status bar at bottom
        self.status_label = QLabel("")
        self.status_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px; padding: 8px 0;")
        content_layout.addWidget(self.status_label)

        scroll.setWidget(content)
        layout.addWidget(scroll)

    def set_root(self, path: str):
        self.root_path = Path(path)
        self.refresh()

    def _run_git(self, *args) -> tuple:
        try:
            result = subprocess.run(
                ["git"] + list(args),
                cwd=self.root_path,
                capture_output=True,
                text=True,
                timeout=10
            )
            return result.returncode == 0, result.stdout.strip(), result.stderr.strip()
        except Exception as e:
            return False, "", str(e)

    def refresh(self):
        if not self.root_path or not self.root_path.exists():
            return

        # Branches
        self.branch_combo.clear()
        ok, branches, _ = self._run_git("branch", "-a")
        if ok:
            current = ""
            for line in branches.split('\n'):
                line = line.strip()
                if line.startswith('* '):
                    current = line[2:]
                    line = current
                if line and not line.startswith('remotes/'):
                    self.branch_combo.addItem(line)
            if current:
                idx = self.branch_combo.findText(current)
                if idx >= 0:
                    self.branch_combo.setCurrentIndex(idx)

        # Status
        self.changes_list.clear()
        change_count = 0
        ok, status, _ = self._run_git("status", "--porcelain")
        if ok and status:
            for line in status.split('\n'):
                if line:
                    change_count += 1
                    status_code = line[:2]
                    filename = line[3:]
                    color = Theme.ACCENT_GREEN if 'A' in status_code else (
                        Theme.ACCENT_RED if 'D' in status_code else Theme.ACCENT_ORANGE
                    )
                    item = QListWidgetItem(f"{status_code} {filename}")
                    item.setForeground(QColor(color))
                    item.setData(Qt.ItemDataRole.UserRole, filename)
                    self.changes_list.addItem(item)

        # Update changes count badge
        self.changes_count.setText(str(change_count))
        if change_count > 0:
            self.changes_count.setStyleSheet(f"""
                background: {Theme.ACCENT_BLUE};
                color: white;
                font-size: 10px;
                padding: 2px 6px;
                border-radius: 8px;
                font-weight: 500;
            """)
        else:
            self.changes_count.setStyleSheet(f"""
                background: {Theme.BG_TERTIARY};
                color: {Theme.TEXT_MUTED};
                font-size: 10px;
                padding: 2px 6px;
                border-radius: 8px;
            """)

        # History
        self.history_list.clear()
        ok, log, _ = self._run_git("log", "--oneline", "-8")
        if ok and log:
            for line in log.split('\n'):
                if line:
                    self.history_list.addItem(QListWidgetItem(line))

        # Emit signal so status bar can update
        self.git_operation_completed.emit()

    def _stage_all(self):
        ok, _, err = self._run_git("add", "-A")
        self.status_label.setText("Staged" if ok else f"Error: {err[:30]}")
        self.refresh()

    def _do_commit(self):
        message = self.commit_input.toPlainText().strip()
        if not message:
            self.status_label.setText("Enter message")
            return

        ok, _, err = self._run_git("commit", "-m", message)
        if ok:
            self.commit_input.clear()
            self.status_label.setText("Committed")
        else:
            self.status_label.setText(f"Error: {err[:30]}")
        self.refresh()

    def _do_pull(self):
        ok, _, err = self._run_git("pull")
        self.status_label.setText("Pulled" if ok else f"Error: {err[:30]}")
        self.refresh()

    def _do_push(self):
        ok, _, err = self._run_git("push")
        self.status_label.setText("Pushed" if ok else f"Error: {err[:30]}")
        self.refresh()  # Refresh to show updated status

    def _view_diff(self, item):
        filename = item.data(Qt.ItemDataRole.UserRole)
        if not filename:
            return

        ok, diff, _ = self._run_git("diff", filename)
        if not diff:
            ok, diff, _ = self._run_git("diff", "--staged", filename)

        if diff:
            dialog = DiffViewer(filename, diff, self)
            dialog.exec()


# ============================================================================
# Settings Panel
# ============================================================================

class SettingsPanel(QWidget):
    """VS Code-style settings panel with comprehensive options."""

    settings_changed = Signal(dict)
    provider_changed = Signal(str)  # Emits 'cisco' or 'anthropic'
    editor_settings_changed = Signal(dict)  # For editor-specific settings

    CISCO_MODELS = ["gpt-4.1", "gpt-4o", "gpt-4o-mini", "o4-mini", "o1"]
    CLAUDE_MODELS = ["claude-sonnet-4-20250514", "claude-opus-4-20250514",
                     "claude-3-5-sonnet-20241022", "claude-3-5-haiku-20241022"]

    def __init__(self, parent=None):
        super().__init__(parent)

        # Load saved settings, merge with defaults
        from circuit_agent.config import load_ui_settings
        saved = load_ui_settings()

        self.settings = {
            "model": "gpt-4.1",
            "auto_approve": False,
            "provider": saved.get("provider", "cisco"),
            "font_size": saved.get("font_size", 13),
            "font_family": saved.get("font_family", "Menlo"),
            "tab_size": saved.get("tab_size", 4),
            "word_wrap": saved.get("word_wrap", False),
            "line_numbers": saved.get("show_line_numbers", True),
            "minimap": saved.get("show_minimap", True),
            "auto_save": saved.get("auto_save", False),
            "extended_thinking": False,
        }

        self.setStyleSheet(f"background: {Theme.BG_MAIN};")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Search bar header
        search_container = QWidget()
        search_container.setFixedHeight(48)
        search_container.setStyleSheet(f"background: {Theme.BG_SECONDARY}; border-bottom: 1px solid {Theme.BORDER};")
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(16, 8, 16, 8)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search settings...")
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                background: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 6px 12px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 12px;
            }}
            QLineEdit:focus {{ border-color: {Theme.ACCENT_BLUE}; }}
        """)
        search_layout.addWidget(self.search_input)
        layout.addWidget(search_container)

        # Scroll area for settings
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setStyleSheet(f"""
            QScrollArea {{ background: {Theme.BG_MAIN}; border: none; }}
            QScrollBar:vertical {{
                background: {Theme.BG_MAIN}; width: 10px;
            }}
            QScrollBar::handle:vertical {{
                background: {Theme.BG_TERTIARY}; border-radius: 5px; min-height: 30px;
            }}
            QScrollBar::handle:vertical:hover {{ background: {Theme.BG_HOVER}; }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
        """)

        content = QWidget()
        content.setStyleSheet(f"background: {Theme.BG_MAIN};")
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(24, 16, 24, 24)
        content_layout.setSpacing(24)

        # =================================================================
        # AI Provider Section
        # =================================================================
        content_layout.addWidget(self._create_section_header("AI Provider"))

        provider_card = self._create_card()
        provider_layout = QVBoxLayout(provider_card)
        provider_layout.setContentsMargins(16, 12, 16, 12)
        provider_layout.setSpacing(12)

        # Provider selection
        provider_row = self._create_setting_row(
            "Provider",
            "Select the AI provider for code assistance"
        )
        self.provider_combo = QComboBox()
        self.provider_combo.addItem("Circuit AI (Cisco)", "cisco")
        self.provider_combo.addItem("Claude Code (Anthropic)", "anthropic")
        self.provider_combo.setFixedWidth(200)
        self.provider_combo.setStyleSheet(self._combo_style())
        self.provider_combo.currentIndexChanged.connect(self._on_provider_changed)
        provider_row.layout().addWidget(self.provider_combo)
        provider_layout.addWidget(provider_row)

        # Model selection
        model_row = self._create_setting_row(
            "Model",
            "Select the AI model to use"
        )
        self.model_combo = QComboBox()
        self.model_combo.setFixedWidth(200)
        self.model_combo.setStyleSheet(self._combo_style())
        self.model_combo.currentTextChanged.connect(self._on_model_changed)
        model_row.layout().addWidget(self.model_combo)
        provider_layout.addWidget(model_row)

        content_layout.addWidget(provider_card)

        # =================================================================
        # AI Behavior Section
        # =================================================================
        content_layout.addWidget(self._create_section_header("AI Behavior"))

        behavior_card = self._create_card()
        behavior_layout = QVBoxLayout(behavior_card)
        behavior_layout.setContentsMargins(16, 12, 16, 12)
        behavior_layout.setSpacing(12)

        # Auto-approve toggle
        auto_row = self._create_setting_row(
            "Auto-approve Safe Operations",
            "Automatically approve file reads and non-destructive operations"
        )
        self.auto_approve_toggle = ToggleSwitch(False)
        self.auto_approve_toggle.toggled_signal.connect(self._on_auto_approve_changed)
        auto_row.layout().addWidget(self.auto_approve_toggle)
        behavior_layout.addWidget(auto_row)

        # Extended thinking toggle
        thinking_row = self._create_setting_row(
            "Extended Thinking",
            "Enable deeper reasoning for complex problems (uses more tokens)"
        )
        self.thinking_toggle = ToggleSwitch(False)
        self.thinking_toggle.toggled_signal.connect(self._on_thinking_changed)
        thinking_row.layout().addWidget(self.thinking_toggle)
        behavior_layout.addWidget(thinking_row)

        content_layout.addWidget(behavior_card)

        # =================================================================
        # Editor Section
        # =================================================================
        content_layout.addWidget(self._create_section_header("Editor"))

        editor_card = self._create_card()
        editor_layout = QVBoxLayout(editor_card)
        editor_layout.setContentsMargins(16, 12, 16, 12)
        editor_layout.setSpacing(12)

        # Font family
        font_row = self._create_setting_row(
            "Font Family",
            "The font used in the code editor"
        )
        self.font_combo = QComboBox()
        self.font_combo.addItems(["Menlo", "Monaco", "SF Mono", "Consolas", "Fira Code", "JetBrains Mono", "Source Code Pro"])
        self.font_combo.setFixedWidth(160)
        self.font_combo.setStyleSheet(self._combo_style())
        self.font_combo.currentTextChanged.connect(self._on_font_changed)
        font_row.layout().addWidget(self.font_combo)
        editor_layout.addWidget(font_row)

        # Font size
        size_row = self._create_setting_row(
            "Font Size",
            "The font size in pixels"
        )
        self.font_size_spin = QSpinBox()
        self.font_size_spin.setRange(8, 32)
        self.font_size_spin.setValue(self.settings.get("font_size", 13))
        self.font_size_spin.setFixedWidth(80)
        self.font_size_spin.setStyleSheet(self._spin_style())
        self.font_size_spin.valueChanged.connect(self._on_font_size_changed)
        size_row.layout().addWidget(self.font_size_spin)
        editor_layout.addWidget(size_row)

        # Tab size
        tab_row = self._create_setting_row(
            "Tab Size",
            "The number of spaces per tab"
        )
        self.tab_size_spin = QSpinBox()
        self.tab_size_spin.setRange(2, 8)
        self.tab_size_spin.setValue(self.settings.get("tab_size", 4))
        self.tab_size_spin.setFixedWidth(80)
        self.tab_size_spin.setStyleSheet(self._spin_style())
        self.tab_size_spin.valueChanged.connect(self._on_tab_size_changed)
        tab_row.layout().addWidget(self.tab_size_spin)
        editor_layout.addWidget(tab_row)

        # Word wrap toggle
        wrap_row = self._create_setting_row(
            "Word Wrap",
            "Wrap long lines to fit the editor width"
        )
        self.word_wrap_toggle = ToggleSwitch(self.settings.get("word_wrap", False))
        self.word_wrap_toggle.toggled_signal.connect(self._on_word_wrap_changed)
        wrap_row.layout().addWidget(self.word_wrap_toggle)
        editor_layout.addWidget(wrap_row)

        # Line numbers toggle
        line_num_row = self._create_setting_row(
            "Line Numbers",
            "Show line numbers in the editor gutter"
        )
        self.line_numbers_toggle = ToggleSwitch(self.settings.get("line_numbers", True))
        self.line_numbers_toggle.toggled_signal.connect(self._on_line_numbers_changed)
        line_num_row.layout().addWidget(self.line_numbers_toggle)
        editor_layout.addWidget(line_num_row)

        # Minimap toggle
        minimap_row = self._create_setting_row(
            "Minimap",
            "Show a minimap of the code on the right side"
        )
        self.minimap_toggle = ToggleSwitch(self.settings.get("minimap", True))
        self.minimap_toggle.toggled_signal.connect(self._on_minimap_changed)
        minimap_row.layout().addWidget(self.minimap_toggle)
        editor_layout.addWidget(minimap_row)

        # Auto-save toggle
        autosave_row = self._create_setting_row(
            "Auto Save",
            "Automatically save files after changes"
        )
        self.autosave_toggle = ToggleSwitch(self.settings.get("auto_save", False))
        self.autosave_toggle.toggled_signal.connect(self._on_autosave_changed)
        autosave_row.layout().addWidget(self.autosave_toggle)
        editor_layout.addWidget(autosave_row)

        content_layout.addWidget(editor_card)

        # =================================================================
        # API Credentials Section
        # =================================================================
        content_layout.addWidget(self._create_section_header("API Credentials"))

        # Cisco credentials card
        cisco_card = self._create_card()
        cisco_layout = QVBoxLayout(cisco_card)
        cisco_layout.setContentsMargins(16, 12, 16, 12)
        cisco_layout.setSpacing(10)

        cisco_header = QLabel("Cisco AI")
        cisco_header.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 13px; font-weight: 600;")
        cisco_layout.addWidget(cisco_header)

        # Client ID
        self.client_id_input = self._create_credential_input("Client ID", "Enter Client ID...")
        cisco_layout.addWidget(self.client_id_input)

        # Client Secret
        self.client_secret_input = self._create_credential_input("Client Secret", "Enter Client Secret...", password=True)
        cisco_layout.addWidget(self.client_secret_input)

        # App Key
        self.app_key_input = self._create_credential_input("App Key", "Enter App Key...", password=True)
        cisco_layout.addWidget(self.app_key_input)

        # Cisco buttons
        cisco_btn_row = QHBoxLayout()
        cisco_btn_row.setSpacing(8)
        save_cisco_btn = self._create_button("Save", primary=True)
        save_cisco_btn.clicked.connect(self._save_credentials)
        cisco_btn_row.addWidget(save_cisco_btn)
        test_cisco_btn = self._create_button("Test Connection")
        test_cisco_btn.clicked.connect(self._test_connection)
        cisco_btn_row.addWidget(test_cisco_btn)
        cisco_btn_row.addStretch()
        cisco_layout.addLayout(cisco_btn_row)

        self.connection_status = QLabel("")
        self.connection_status.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
        cisco_layout.addWidget(self.connection_status)

        content_layout.addWidget(cisco_card)

        # Anthropic credentials card
        anthropic_card = self._create_card()
        anthropic_layout = QVBoxLayout(anthropic_card)
        anthropic_layout.setContentsMargins(16, 12, 16, 12)
        anthropic_layout.setSpacing(10)

        anthropic_header = QLabel("Anthropic (Claude)")
        anthropic_header.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 13px; font-weight: 600;")
        anthropic_layout.addWidget(anthropic_header)

        # API Key
        self.anthropic_key_input = self._create_credential_input("API Key", "sk-ant-...", password=True)
        anthropic_layout.addWidget(self.anthropic_key_input)

        # Anthropic buttons
        anthropic_btn_row = QHBoxLayout()
        anthropic_btn_row.setSpacing(8)
        save_anthropic_btn = self._create_button("Save", primary=True)
        save_anthropic_btn.clicked.connect(self._save_anthropic_key)
        anthropic_btn_row.addWidget(save_anthropic_btn)
        test_anthropic_btn = self._create_button("Test Connection")
        test_anthropic_btn.clicked.connect(self._test_anthropic)
        anthropic_btn_row.addWidget(test_anthropic_btn)
        anthropic_btn_row.addStretch()
        anthropic_layout.addLayout(anthropic_btn_row)

        self.anthropic_status = QLabel("")
        self.anthropic_status.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
        anthropic_layout.addWidget(self.anthropic_status)

        content_layout.addWidget(anthropic_card)

        # =================================================================
        # MCP Servers Section (GitHub Integration)
        # =================================================================
        content_layout.addWidget(self._create_section_header("MCP Servers"))

        mcp_card = self._create_card()
        mcp_layout = QVBoxLayout(mcp_card)
        mcp_layout.setContentsMargins(16, 12, 16, 12)
        mcp_layout.setSpacing(10)

        # GitHub MCP header with enable toggle
        github_header_row = QHBoxLayout()
        github_header = QLabel("GitHub Integration")
        github_header.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 13px; font-weight: 600;")
        github_header_row.addWidget(github_header)
        github_header_row.addStretch()
        self.github_mcp_toggle = ToggleSwitch(False)
        self.github_mcp_toggle.toggled_signal.connect(self._on_github_mcp_toggled)
        github_header_row.addWidget(self.github_mcp_toggle)
        mcp_layout.addLayout(github_header_row)

        github_desc = QLabel("Connect to GitHub MCP Server for repository, issues, PR, and Actions tools")
        github_desc.setWordWrap(True)
        github_desc.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
        mcp_layout.addWidget(github_desc)

        # GitHub PAT input
        self.github_pat_input = self._create_credential_input("Personal Access Token", "ghp_...", password=True)
        mcp_layout.addWidget(self.github_pat_input)

        # Toolsets selection
        toolsets_label = QLabel("Enabled Toolsets:")
        toolsets_label.setStyleSheet(f"color: {Theme.TEXT_SECONDARY}; font-size: 12px; margin-top: 8px;")
        mcp_layout.addWidget(toolsets_label)

        toolsets_grid = QGridLayout()
        toolsets_grid.setSpacing(8)

        self.github_toolset_checkboxes = {}
        toolsets = [
            ("repos", "Repositories"),
            ("issues", "Issues"),
            ("pull_requests", "Pull Requests"),
            ("actions", "Actions"),
            ("code_security", "Code Security"),
            ("discussions", "Discussions"),
        ]

        for i, (toolset_id, toolset_name) in enumerate(toolsets):
            checkbox = QCheckBox(toolset_name)
            checkbox.setStyleSheet(f"""
                QCheckBox {{
                    color: {Theme.TEXT_SECONDARY};
                    font-size: 11px;
                    spacing: 6px;
                }}
                QCheckBox::indicator {{
                    width: 14px;
                    height: 14px;
                    border-radius: 3px;
                    border: 1px solid {Theme.BORDER};
                    background: {Theme.BG_INPUT};
                }}
                QCheckBox::indicator:checked {{
                    background: {Theme.ACCENT_BLUE};
                    border-color: {Theme.ACCENT_BLUE};
                }}
            """)
            # Default: enable repos, issues, pull_requests, actions
            checkbox.setChecked(toolset_id in ["repos", "issues", "pull_requests", "actions"])
            self.github_toolset_checkboxes[toolset_id] = checkbox
            toolsets_grid.addWidget(checkbox, i // 3, i % 3)

        mcp_layout.addLayout(toolsets_grid)

        # MCP buttons
        mcp_btn_row = QHBoxLayout()
        mcp_btn_row.setSpacing(8)
        save_mcp_btn = self._create_button("Save GitHub Settings", primary=True)
        save_mcp_btn.clicked.connect(self._save_github_mcp)
        mcp_btn_row.addWidget(save_mcp_btn)
        test_mcp_btn = self._create_button("Test Connection")
        test_mcp_btn.clicked.connect(self._test_github_mcp)
        mcp_btn_row.addWidget(test_mcp_btn)
        mcp_btn_row.addStretch()
        mcp_layout.addLayout(mcp_btn_row)

        self.mcp_status = QLabel("")
        self.mcp_status.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
        mcp_layout.addWidget(self.mcp_status)

        content_layout.addWidget(mcp_card)

        # =================================================================
        # Keyboard Shortcuts Section
        # =================================================================
        content_layout.addWidget(self._create_section_header("Keyboard Shortcuts"))

        shortcuts_card = self._create_card()
        shortcuts_layout = QVBoxLayout(shortcuts_card)
        shortcuts_layout.setContentsMargins(16, 12, 16, 12)
        shortcuts_layout.setSpacing(6)

        shortcuts = [
            ("Cmd+N", "New File"),
            ("Cmd+O", "Open File"),
            ("Cmd+S", "Save File"),
            ("Cmd+K", "Open Folder"),
            ("Cmd+F", "Find in File"),
            ("Cmd+Shift+F", "Find in Project"),
            ("Cmd+`", "Toggle Terminal"),
            ("Cmd+B", "Toggle Sidebar"),
            ("Cmd+,", "Open Settings"),
            ("Cmd+Enter", "Send to AI"),
            ("Escape", "Cancel Request"),
        ]

        for key, desc in shortcuts:
            row = QHBoxLayout()
            row.setContentsMargins(0, 4, 0, 4)

            key_label = QLabel(key)
            key_label.setFixedWidth(120)
            key_label.setStyleSheet(f"""
                background: {Theme.BG_TERTIARY};
                color: {Theme.TEXT_PRIMARY};
                font-size: 11px;
                font-family: Menlo, monospace;
                padding: 4px 8px;
                border-radius: 4px;
            """)
            row.addWidget(key_label)

            desc_label = QLabel(desc)
            desc_label.setStyleSheet(f"color: {Theme.TEXT_SECONDARY}; font-size: 12px;")
            row.addWidget(desc_label)
            row.addStretch()

            shortcuts_layout.addLayout(row)

        content_layout.addWidget(shortcuts_card)

        # =================================================================
        # About Section
        # =================================================================
        content_layout.addWidget(self._create_section_header("About"))

        about_card = self._create_card()
        about_layout = QVBoxLayout(about_card)
        about_layout.setContentsMargins(16, 16, 16, 16)
        about_layout.setSpacing(8)

        app_name = QLabel("Circuit IDE")
        app_name.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 18px; font-weight: 600;")
        about_layout.addWidget(app_name)

        version = QLabel("Version 8.0.0")
        version.setStyleSheet(f"color: {Theme.TEXT_SECONDARY}; font-size: 12px;")
        about_layout.addWidget(version)

        desc = QLabel("Enterprise AI-powered code editor with integrated intelligent code assistance.")
        desc.setWordWrap(True)
        desc.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 12px; margin-top: 8px;")
        about_layout.addWidget(desc)

        content_layout.addWidget(about_card)

        content_layout.addStretch()
        scroll.setWidget(content)
        layout.addWidget(scroll)

        # Load existing settings
        self._load_credentials()
        self._load_anthropic_key()
        self._load_github_mcp_settings()
        self._load_provider_preference()
        self._update_model_list()

    # =========================================================================
    # Helper Methods for UI Creation
    # =========================================================================

    def _create_section_header(self, title: str) -> QLabel:
        """Create a section header label."""
        header = QLabel(title.upper())
        header.setStyleSheet(f"""
            color: {Theme.TEXT_SECONDARY};
            font-size: 11px;
            font-weight: 600;
            letter-spacing: 0.5px;
            margin-bottom: 4px;
        """)
        return header

    def _create_card(self) -> QWidget:
        """Create a settings card container."""
        card = QWidget()
        card.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 6px;
            }}
        """)
        return card

    def _create_setting_row(self, title: str, description: str) -> QWidget:
        """Create a setting row with title, description, and control area."""
        row = QWidget()
        row.setStyleSheet("background: transparent; border: none;")
        layout = QHBoxLayout(row)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        text_container = QVBoxLayout()
        text_container.setSpacing(2)

        title_label = QLabel(title)
        title_label.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 12px; font-weight: 500; border: none;")
        text_container.addWidget(title_label)

        desc_label = QLabel(description)
        desc_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px; border: none;")
        text_container.addWidget(desc_label)

        layout.addLayout(text_container, 1)
        return row

    def _create_credential_input(self, label: str, placeholder: str, password: bool = False) -> QWidget:
        """Create a labeled credential input field."""
        container = QWidget()
        container.setStyleSheet("background: transparent; border: none;")
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        lbl = QLabel(label + ":")
        lbl.setFixedWidth(100)
        lbl.setStyleSheet(f"color: {Theme.TEXT_SECONDARY}; font-size: 11px; border: none;")
        layout.addWidget(lbl)

        input_field = QLineEdit()
        input_field.setPlaceholderText(placeholder)
        if password:
            input_field.setEchoMode(QLineEdit.EchoMode.Password)
        input_field.setStyleSheet(f"""
            QLineEdit {{
                background: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 6px 10px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 12px;
            }}
            QLineEdit:focus {{ border-color: {Theme.ACCENT_BLUE}; }}
        """)
        layout.addWidget(input_field)

        if password:
            show_btn = QPushButton()
            show_btn.setIcon(Icons.eye(16, Theme.TEXT_MUTED))
            show_btn.setFixedSize(28, 28)
            show_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            show_btn.setStyleSheet(f"""
                QPushButton {{
                    background: transparent;
                    border: none;
                    border-radius: 4px;
                }}
                QPushButton:hover {{ background: {Theme.BG_HOVER}; }}
            """)
            show_btn.clicked.connect(lambda: self._toggle_echo(input_field))
            layout.addWidget(show_btn)

        # Store the input field as an attribute on the container for easy access
        container.input_field = input_field
        return container

    def _create_button(self, text: str, primary: bool = False) -> QPushButton:
        """Create a styled button."""
        btn = QPushButton(text)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.setFixedHeight(28)
        if primary:
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {Theme.ACCENT_BLUE};
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 0 16px;
                    font-size: 11px;
                    font-weight: 500;
                }}
                QPushButton:hover {{ background: #3b82f6; }}
                QPushButton:pressed {{ background: #1d4ed8; }}
            """)
        else:
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {Theme.BG_TERTIARY};
                    color: {Theme.TEXT_PRIMARY};
                    border: 1px solid {Theme.BORDER};
                    border-radius: 4px;
                    padding: 0 16px;
                    font-size: 11px;
                }}
                QPushButton:hover {{ background: {Theme.BG_HOVER}; }}
                QPushButton:pressed {{ background: {Theme.BG_SELECTED}; }}
            """)
        return btn

    def _combo_style(self) -> str:
        """Return stylesheet for combo boxes."""
        return f"""
            QComboBox {{
                background: {Theme.BG_INPUT};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
            }}
            QComboBox:hover {{ border-color: {Theme.ACCENT_BLUE}; }}
            QComboBox::drop-down {{ border: none; width: 24px; }}
            QComboBox QAbstractItemView {{
                background: {Theme.BG_SECONDARY};
                color: {Theme.TEXT_PRIMARY};
                selection-background-color: {Theme.BG_SELECTED};
                border: 1px solid {Theme.BORDER};
            }}
        """

    def _spin_style(self) -> str:
        """Return stylesheet for spin boxes."""
        return f"""
            QSpinBox {{
                background: {Theme.BG_INPUT};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 12px;
            }}
            QSpinBox:hover {{ border-color: {Theme.ACCENT_BLUE}; }}
            QSpinBox::up-button, QSpinBox::down-button {{
                width: 16px;
                background: {Theme.BG_TERTIARY};
                border: none;
            }}
        """

    def _toggle_echo(self, line_edit: QLineEdit):
        """Toggle password visibility for a line edit."""
        if line_edit.echoMode() == QLineEdit.EchoMode.Password:
            line_edit.setEchoMode(QLineEdit.EchoMode.Normal)
        else:
            line_edit.setEchoMode(QLineEdit.EchoMode.Password)

    # =========================================================================
    # Editor Settings Handlers
    # =========================================================================

    def _save_setting(self, key: str, value):
        """Save a single setting to persistent storage."""
        from circuit_agent.config import update_ui_setting
        update_ui_setting(key, value)

    def _on_font_changed(self, font: str):
        """Handle font family change."""
        self.settings["font_family"] = font
        self._save_setting("font_family", font)
        self.editor_settings_changed.emit({"font_family": font})

    def _on_font_size_changed(self, size: int):
        """Handle font size change."""
        self.settings["font_size"] = size
        self._save_setting("font_size", size)
        self.editor_settings_changed.emit({"font_size": size})

    def _on_tab_size_changed(self, size: int):
        """Handle tab size change."""
        self.settings["tab_size"] = size
        self._save_setting("tab_size", size)
        self.editor_settings_changed.emit({"tab_size": size})

    def _on_word_wrap_changed(self, enabled: bool):
        """Handle word wrap toggle."""
        self.settings["word_wrap"] = enabled
        self._save_setting("word_wrap", enabled)
        self.editor_settings_changed.emit({"word_wrap": enabled})

    def _on_line_numbers_changed(self, enabled: bool):
        """Handle line numbers toggle."""
        self.settings["line_numbers"] = enabled
        self._save_setting("show_line_numbers", enabled)
        self.editor_settings_changed.emit({"line_numbers": enabled})

    def _on_minimap_changed(self, enabled: bool):
        """Handle minimap toggle."""
        self.settings["minimap"] = enabled
        self._save_setting("show_minimap", enabled)
        self.editor_settings_changed.emit({"minimap": enabled})

    def _on_autosave_changed(self, enabled: bool):
        """Handle auto-save toggle."""
        self.settings["auto_save"] = enabled
        self._save_setting("auto_save", enabled)
        self.editor_settings_changed.emit({"auto_save": enabled})

    def _on_thinking_changed(self, enabled: bool):
        """Handle extended thinking toggle."""
        self.settings["extended_thinking"] = enabled
        self.settings_changed.emit(self.settings)

    # =========================================================================
    # Credential Methods
    # =========================================================================

    def _load_credentials(self):
        """Load saved credentials into the input fields."""
        try:
            from circuit_agent.config import load_credentials
            client_id, client_secret, app_key = load_credentials()
            if client_id:
                self.client_id_input.input_field.setText(client_id)
            if client_secret:
                self.client_secret_input.input_field.setText(client_secret)
            if app_key:
                self.app_key_input.input_field.setText(app_key)
            if client_id and client_secret and app_key:
                self.connection_status.setText("Credentials loaded")
                self.connection_status.setStyleSheet(f"color: {Theme.SUCCESS}; font-size: 11px;")
        except Exception as e:
            self.connection_status.setText(f"Could not load credentials")
            self.connection_status.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")

    def _save_credentials(self):
        """Save credentials to config."""
        client_id = self.client_id_input.input_field.text().strip()
        client_secret = self.client_secret_input.input_field.text().strip()
        app_key = self.app_key_input.input_field.text().strip()

        if not all([client_id, client_secret, app_key]):
            self.connection_status.setText("Please fill in all credential fields")
            self.connection_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 11px;")
            return

        try:
            from circuit_agent.config import save_credentials
            save_credentials(client_id, client_secret, app_key)
            self.connection_status.setText("Credentials saved successfully")
            self.connection_status.setStyleSheet(f"color: {Theme.SUCCESS}; font-size: 11px;")
        except Exception as e:
            self.connection_status.setText(f"Save failed: {str(e)[:50]}")
            self.connection_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 11px;")

    def _test_connection(self):
        """Test connection with the API."""
        client_id = self.client_id_input.input_field.text().strip()
        client_secret = self.client_secret_input.input_field.text().strip()
        app_key = self.app_key_input.input_field.text().strip()

        if not all([client_id, client_secret, app_key]):
            self.connection_status.setText("Please fill in all credential fields first")
            self.connection_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 10px;")
            return

        self.connection_status.setText("Testing connection...")
        self.connection_status.setStyleSheet(f"color: {Theme.WARNING}; font-size: 10px;")

        # Run test in background thread
        def test_thread():
            try:
                from circuit_agent.config import TOKEN_URL, ssl_config
                import httpx

                # Try to get an OAuth token
                token_data = {
                    "grant_type": "client_credentials",
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "scope": "openai",
                }

                with httpx.Client(verify=ssl_config.get_verify_param(), timeout=10.0) as client:
                    response = client.post(TOKEN_URL, data=token_data)
                    if response.status_code == 200:
                        return True, "Connection successful!"
                    else:
                        return False, f"Auth failed: {response.status_code}"
            except Exception as e:
                return False, f"Error: {str(e)[:50]}"

        import threading
        def run_test():
            success, message = test_thread()
            # Update UI on main thread
            QTimer.singleShot(0, lambda: self._show_connection_result(success, message))

        thread = threading.Thread(target=run_test, daemon=True)
        thread.start()

    def _show_connection_result(self, success: bool, message: str = ""):
        if success:
            self.connection_status.setText(message or "Connected")
            self.connection_status.setStyleSheet(f"color: {Theme.SUCCESS}; font-size: 10px;")
        else:
            self.connection_status.setText(message or "Connection failed")
            self.connection_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 10px;")

    # =========================================================================
    # Anthropic/Claude Methods
    # =========================================================================

    def _load_anthropic_key(self):
        """Load Anthropic API key from storage."""
        try:
            from circuit_agent.config import load_anthropic_key
            key = load_anthropic_key()
            if key:
                self.anthropic_key_input.input_field.setText(key)
                self.anthropic_status.setText("API key loaded")
                self.anthropic_status.setStyleSheet(f"color: {Theme.SUCCESS}; font-size: 11px;")
        except Exception:
            pass

    def _save_anthropic_key(self):
        """Save Anthropic API key."""
        api_key = self.anthropic_key_input.input_field.text().strip()

        if not api_key:
            self.anthropic_status.setText("Please enter an API key")
            self.anthropic_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 11px;")
            return

        try:
            from circuit_agent.config import save_anthropic_key
            success, method = save_anthropic_key(api_key)
            if success:
                self.anthropic_status.setText(f"API key saved ({method})")
                self.anthropic_status.setStyleSheet(f"color: {Theme.SUCCESS}; font-size: 11px;")
            else:
                self.anthropic_status.setText("Failed to save API key")
                self.anthropic_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 11px;")
        except Exception as e:
            self.anthropic_status.setText(f"Error: {str(e)[:40]}")
            self.anthropic_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 11px;")

    def _test_anthropic(self):
        """Test Anthropic API connection."""
        api_key = self.anthropic_key_input.input_field.text().strip()

        if not api_key:
            self.anthropic_status.setText("Please enter an API key first")
            self.anthropic_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 10px;")
            return

        self.anthropic_status.setText("Testing connection...")
        self.anthropic_status.setStyleSheet(f"color: {Theme.WARNING}; font-size: 10px;")

        def test_thread():
            try:
                import httpx

                with httpx.Client(timeout=10.0) as client:
                    response = client.get(
                        "https://api.anthropic.com/v1/models",
                        headers={
                            "x-api-key": api_key,
                            "anthropic-version": "2023-06-01",
                        }
                    )
                    if response.status_code == 200:
                        return True, "Connection successful!"
                    elif response.status_code == 401:
                        return False, "Invalid API key"
                    else:
                        return False, f"Error: {response.status_code}"
            except Exception as e:
                return False, f"Error: {str(e)[:40]}"

        import threading
        def run_test():
            success, message = test_thread()
            QTimer.singleShot(0, lambda: self._show_anthropic_result(success, message))

        thread = threading.Thread(target=run_test, daemon=True)
        thread.start()

    def _show_anthropic_result(self, success: bool, message: str = ""):
        if success:
            self.anthropic_status.setText(message or "Connected")
            self.anthropic_status.setStyleSheet(f"color: {Theme.SUCCESS}; font-size: 10px;")
        else:
            self.anthropic_status.setText(message or "Connection failed")
            self.anthropic_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 10px;")

    # =========================================================================
    # MCP/GitHub Methods
    # =========================================================================

    def _load_github_mcp_settings(self):
        """Load GitHub MCP settings from storage."""
        try:
            from circuit_agent.config import load_github_pat, load_github_mcp_config

            # Load PAT
            pat = load_github_pat()
            if pat:
                self.github_pat_input.input_field.setText(pat)

            # Load MCP config
            config = load_github_mcp_config()
            self.github_mcp_toggle.setChecked(config.get("enabled", False))

            # Load toolsets
            toolsets = config.get("toolsets", ["repos", "issues", "pull_requests", "actions"])
            for toolset_id, checkbox in self.github_toolset_checkboxes.items():
                checkbox.setChecked(toolset_id in toolsets)

            if pat and config.get("enabled"):
                self.mcp_status.setText("GitHub MCP configured")
                self.mcp_status.setStyleSheet(f"color: {Theme.SUCCESS}; font-size: 11px;")
        except Exception as e:
            pass

    def _on_github_mcp_toggled(self, enabled: bool):
        """Handle GitHub MCP toggle."""
        pass  # Settings are saved when Save button is clicked

    def _save_github_mcp(self):
        """Save GitHub MCP settings."""
        pat = self.github_pat_input.input_field.text().strip()
        enabled = self.github_mcp_toggle.isChecked()

        if enabled and not pat:
            self.mcp_status.setText("Please enter a GitHub Personal Access Token")
            self.mcp_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 11px;")
            return

        # Get enabled toolsets
        toolsets = [
            toolset_id for toolset_id, checkbox in self.github_toolset_checkboxes.items()
            if checkbox.isChecked()
        ]

        try:
            from circuit_agent.config import save_github_pat, save_github_mcp_config

            # Save PAT if provided
            if pat:
                success, method = save_github_pat(pat)
                if not success:
                    self.mcp_status.setText("Failed to save PAT")
                    self.mcp_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 11px;")
                    return

            # Save MCP config
            save_github_mcp_config(enabled=enabled, toolsets=toolsets)

            status_msg = "GitHub MCP settings saved"
            if enabled:
                status_msg += f" ({len(toolsets)} toolsets)"
            self.mcp_status.setText(status_msg)
            self.mcp_status.setStyleSheet(f"color: {Theme.SUCCESS}; font-size: 11px;")

            # Emit signal so AgentWorker can reinitialize MCP
            self.settings_changed.emit({"github_mcp_updated": True})

        except Exception as e:
            self.mcp_status.setText(f"Error: {str(e)[:40]}")
            self.mcp_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 11px;")

    def _test_github_mcp(self):
        """Test GitHub MCP connection."""
        pat = self.github_pat_input.input_field.text().strip()

        if not pat:
            self.mcp_status.setText("Please enter a GitHub PAT first")
            self.mcp_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 11px;")
            return

        # Basic PAT validation
        if not (pat.startswith("ghp_") or pat.startswith("github_pat_")):
            self.mcp_status.setText("Invalid PAT format (should start with ghp_ or github_pat_)")
            self.mcp_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 11px;")
            return

        self.mcp_status.setText("Testing GitHub connection...")
        self.mcp_status.setStyleSheet(f"color: {Theme.WARNING}; font-size: 11px;")

        def test_thread():
            try:
                from circuit_agent.config import ssl_config
                import httpx

                # Test GitHub API with the PAT
                with httpx.Client(verify=ssl_config.get_verify_param(), timeout=10.0) as client:
                    response = client.get(
                        "https://api.github.com/user",
                        headers={
                            "Authorization": f"Bearer {pat}",
                            "Accept": "application/vnd.github+json",
                        }
                    )
                    if response.status_code == 200:
                        user = response.json()
                        username = user.get("login", "Unknown")
                        return True, f"Connected as @{username}"
                    elif response.status_code == 401:
                        return False, "Invalid or expired token"
                    else:
                        return False, f"Error: {response.status_code}"
            except httpx.TimeoutException:
                return False, "Connection timed out"
            except httpx.ConnectError as e:
                return False, f"Connection failed: {str(e)[:30]}"
            except Exception as e:
                return False, f"Error: {str(e)[:40]}"

        import threading
        def run_test():
            success, message = test_thread()
            QTimer.singleShot(0, lambda: self._show_mcp_result(success, message))

        thread = threading.Thread(target=run_test, daemon=True)
        thread.start()

    def _show_mcp_result(self, success: bool, message: str = ""):
        if success:
            self.mcp_status.setText(message or "Connected")
            self.mcp_status.setStyleSheet(f"color: {Theme.SUCCESS}; font-size: 11px;")
        else:
            self.mcp_status.setText(message or "Connection failed")
            self.mcp_status.setStyleSheet(f"color: {Theme.ERROR}; font-size: 11px;")

    # =========================================================================
    # Provider Selection Methods
    # =========================================================================

    def _load_provider_preference(self):
        """Load saved provider preference."""
        try:
            from circuit_agent.config import load_provider_preference
            provider = load_provider_preference()
            index = 0 if provider == "cisco" else 1
            self.provider_combo.setCurrentIndex(index)
            self.settings["provider"] = provider
        except Exception:
            pass

    def _on_provider_changed(self, index):
        """Handle provider selection change."""
        provider = self.provider_combo.currentData()
        self.settings["provider"] = provider
        self._update_model_list()

        # Save preference
        try:
            from circuit_agent.config import save_provider_preference
            save_provider_preference(provider)
        except Exception:
            pass

        self.provider_changed.emit(provider)
        self.settings_changed.emit(self.settings)

    def _update_model_list(self):
        """Update model dropdown based on selected provider."""
        # Guard against being called before UI is fully set up
        if not hasattr(self, 'model_combo'):
            return

        provider = self.settings.get("provider", "cisco")
        current_model = self.model_combo.currentText()

        self.model_combo.blockSignals(True)
        self.model_combo.clear()

        if provider == "cisco":
            self.model_combo.addItems(self.CISCO_MODELS)
            # Try to restore selection or default
            if current_model in self.CISCO_MODELS:
                self.model_combo.setCurrentText(current_model)
            else:
                self.model_combo.setCurrentIndex(0)
                self.settings["model"] = self.CISCO_MODELS[0]
        else:
            self.model_combo.addItems(self.CLAUDE_MODELS)
            if current_model in self.CLAUDE_MODELS:
                self.model_combo.setCurrentText(current_model)
            else:
                self.model_combo.setCurrentIndex(0)
                self.settings["model"] = self.CLAUDE_MODELS[0]

        self.model_combo.blockSignals(False)

    def _on_model_changed(self, model):
        self.settings["model"] = model
        self.settings_changed.emit(self.settings)

    def _on_auto_approve_changed(self, enabled: bool):
        """Handle auto-approve toggle change."""
        self.settings["auto_approve"] = enabled
        self.settings_changed.emit(self.settings)


# ============================================================================
# Excel Viewer/Editor
# ============================================================================

class ExcelViewer(QWidget):
    """Excel file viewer and editor using QTableWidget."""

    modification_changed = Signal(bool)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_file: Optional[Path] = None
        self._is_modified = False
        self._workbook = None
        self._current_sheet = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Toolbar
        toolbar = QWidget()
        toolbar.setFixedHeight(32)
        toolbar.setStyleSheet(f"background: {Theme.BG_SECONDARY}; border-bottom: 1px solid {Theme.BORDER};")
        toolbar_layout = QHBoxLayout(toolbar)
        toolbar_layout.setContentsMargins(8, 0, 8, 0)
        toolbar_layout.setSpacing(8)

        # Sheet selector
        sheet_label = QLabel("Sheet:")
        sheet_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
        toolbar_layout.addWidget(sheet_label)

        self.sheet_combo = QComboBox()
        self.sheet_combo.setFixedWidth(150)
        self.sheet_combo.setStyleSheet(f"""
            QComboBox {{
                background: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 3px;
                padding: 2px 6px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 11px;
            }}
        """)
        self.sheet_combo.currentTextChanged.connect(self._on_sheet_changed)
        toolbar_layout.addWidget(self.sheet_combo)

        toolbar_layout.addStretch()

        # Add row/column buttons
        add_row_btn = SecondaryButton("+ Row")
        add_row_btn.clicked.connect(self._add_row)
        toolbar_layout.addWidget(add_row_btn)

        add_col_btn = SecondaryButton("+ Col")
        add_col_btn.clicked.connect(self._add_column)
        toolbar_layout.addWidget(add_col_btn)

        # Save button
        save_btn = CompactButton("Save")
        save_btn.clicked.connect(self.save_file)
        toolbar_layout.addWidget(save_btn)

        layout.addWidget(toolbar)

        # Table widget
        self.table = QTableWidget()
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background: {Theme.BG_MAIN};
                color: {Theme.TEXT_PRIMARY};
                border: none;
                gridline-color: {Theme.BORDER};
                font-size: 12px;
            }}
            QTableWidget::item {{
                padding: 4px 8px;
            }}
            QTableWidget::item:selected {{
                background: {Theme.BG_SELECTED};
            }}
            QHeaderView::section {{
                background: {Theme.BG_SECONDARY};
                color: {Theme.TEXT_SECONDARY};
                border: none;
                border-right: 1px solid {Theme.BORDER};
                border-bottom: 1px solid {Theme.BORDER};
                padding: 4px 8px;
                font-size: 11px;
                font-weight: 500;
            }}
        """)
        self.table.cellChanged.connect(self._on_cell_changed)
        layout.addWidget(self.table)

    def load_file(self, path: Path):
        """Load an Excel file."""
        self.current_file = path
        try:
            import openpyxl
            self._workbook = openpyxl.load_workbook(path)

            # Populate sheet selector
            self.sheet_combo.clear()
            self.sheet_combo.addItems(self._workbook.sheetnames)

            # Load first sheet
            if self._workbook.sheetnames:
                self._load_sheet(self._workbook.sheetnames[0])

            self._is_modified = False
        except ImportError:
            self._show_error("openpyxl not installed. Install with: pip install openpyxl")
        except Exception as e:
            self._show_error(f"Error loading file: {e}")

    def _show_error(self, message: str):
        """Show error in the table area."""
        self.table.clear()
        self.table.setRowCount(1)
        self.table.setColumnCount(1)
        self.table.setItem(0, 0, QTableWidgetItem(message))

    def _load_sheet(self, sheet_name: str):
        """Load a specific sheet into the table."""
        if not self._workbook or sheet_name not in self._workbook.sheetnames:
            return

        self._current_sheet = sheet_name
        sheet = self._workbook[sheet_name]

        # Block signals while loading
        self.table.blockSignals(True)

        # Set dimensions
        self.table.setRowCount(sheet.max_row or 1)
        self.table.setColumnCount(sheet.max_column or 1)

        # Set column headers (A, B, C, ...)
        headers = []
        for col in range(1, (sheet.max_column or 1) + 1):
            headers.append(self._col_to_letter(col))
        self.table.setHorizontalHeaderLabels(headers)

        # Load cell data
        for row in range(1, (sheet.max_row or 1) + 1):
            for col in range(1, (sheet.max_column or 1) + 1):
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                if value is not None:
                    item = QTableWidgetItem(str(value))
                    self.table.setItem(row - 1, col - 1, item)

        self.table.blockSignals(False)

    def _col_to_letter(self, col: int) -> str:
        """Convert column number to Excel letter (1=A, 2=B, etc.)."""
        result = ""
        while col > 0:
            col, remainder = divmod(col - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _on_sheet_changed(self, sheet_name: str):
        """Handle sheet selection change."""
        if sheet_name:
            self._load_sheet(sheet_name)

    def _on_cell_changed(self, row: int, col: int):
        """Handle cell edit."""
        if not self._workbook or not self._current_sheet:
            return

        item = self.table.item(row, col)
        value = item.text() if item else ""

        # Update workbook
        sheet = self._workbook[self._current_sheet]
        sheet.cell(row=row + 1, column=col + 1, value=value)

        self._is_modified = True
        self.modification_changed.emit(True)

    def _add_row(self):
        """Add a new row at the end."""
        self.table.insertRow(self.table.rowCount())
        self._is_modified = True
        self.modification_changed.emit(True)

    def _add_column(self):
        """Add a new column at the end."""
        col_count = self.table.columnCount()
        self.table.insertColumn(col_count)
        # Update header
        self.table.setHorizontalHeaderItem(col_count,
            QTableWidgetItem(self._col_to_letter(col_count + 1)))
        self._is_modified = True
        self.modification_changed.emit(True)

    def save_file(self):
        """Save the Excel file."""
        if not self._workbook or not self.current_file:
            return

        try:
            # Sync all data from table to workbook
            if self._current_sheet:
                sheet = self._workbook[self._current_sheet]
                for row in range(self.table.rowCount()):
                    for col in range(self.table.columnCount()):
                        item = self.table.item(row, col)
                        value = item.text() if item else None
                        sheet.cell(row=row + 1, column=col + 1, value=value)

            self._workbook.save(self.current_file)
            self._is_modified = False
            self.modification_changed.emit(False)
        except Exception as e:
            QMessageBox.warning(self, "Save Error", f"Could not save file: {e}")

    @property
    def is_modified(self) -> bool:
        return self._is_modified


# ============================================================================
# Code Editor
# ============================================================================

class FindReplaceWidget(QWidget):
    """Find and Replace widget for code editor."""

    find_next = Signal(str, bool)  # text, case_sensitive
    find_prev = Signal(str, bool)
    replace_one = Signal(str, str, bool)  # find, replace, case_sensitive
    replace_all = Signal(str, str, bool)
    closed = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(70)
        self._show_replace = False

        self.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border-bottom: 1px solid {Theme.BORDER};
            }}
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 6, 8, 6)
        layout.setSpacing(4)

        # Find row
        find_row = QHBoxLayout()
        find_row.setSpacing(6)

        self.find_input = CompactLineEdit("Find...")
        self.find_input.setFixedWidth(200)
        self.find_input.textChanged.connect(self._on_find_changed)
        self.find_input.returnPressed.connect(self._find_next)
        find_row.addWidget(self.find_input)

        self.match_label = QLabel("No results")
        self.match_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 10px;")
        self.match_label.setFixedWidth(80)
        find_row.addWidget(self.match_label)

        prev_btn = IconButton(Icons.arrow_up, "Previous (Shift+F3)")
        prev_btn.clicked.connect(self._find_prev)
        find_row.addWidget(prev_btn)

        next_btn = IconButton(Icons.arrow_down, "Next (F3)")
        next_btn.clicked.connect(self._find_next)
        find_row.addWidget(next_btn)

        self.case_btn = CompactCheckBox("Aa")
        self.case_btn.setToolTip("Match case")
        find_row.addWidget(self.case_btn)

        find_row.addStretch()

        self.toggle_replace_btn = IconButton(Icons.arrow_down, "Toggle Replace")
        self.toggle_replace_btn.clicked.connect(self._toggle_replace)
        find_row.addWidget(self.toggle_replace_btn)

        close_btn = IconButton(Icons.close, "Close (Esc)")
        close_btn.clicked.connect(self._close)
        find_row.addWidget(close_btn)

        layout.addLayout(find_row)

        # Replace row (initially hidden)
        self.replace_row = QWidget()
        replace_layout = QHBoxLayout(self.replace_row)
        replace_layout.setContentsMargins(0, 0, 0, 0)
        replace_layout.setSpacing(6)

        self.replace_input = CompactLineEdit("Replace with...")
        self.replace_input.setFixedWidth(200)
        self.replace_input.returnPressed.connect(self._replace_one)
        replace_layout.addWidget(self.replace_input)

        replace_btn = SecondaryButton("Replace")
        replace_btn.clicked.connect(self._replace_one)
        replace_layout.addWidget(replace_btn)

        replace_all_btn = SecondaryButton("Replace All")
        replace_all_btn.clicked.connect(self._replace_all)
        replace_layout.addWidget(replace_all_btn)

        replace_layout.addStretch()

        layout.addWidget(self.replace_row)
        self.replace_row.hide()
        self.setFixedHeight(35)

    def _on_find_changed(self, text: str):
        """Trigger search when text changes."""
        if text:
            self.find_next.emit(text, self.case_btn.isChecked())

    def _find_next(self):
        text = self.find_input.text()
        if text:
            self.find_next.emit(text, self.case_btn.isChecked())

    def _find_prev(self):
        text = self.find_input.text()
        if text:
            self.find_prev.emit(text, self.case_btn.isChecked())

    def _replace_one(self):
        find_text = self.find_input.text()
        replace_text = self.replace_input.text()
        if find_text:
            self.replace_one.emit(find_text, replace_text, self.case_btn.isChecked())

    def _replace_all(self):
        find_text = self.find_input.text()
        replace_text = self.replace_input.text()
        if find_text:
            self.replace_all.emit(find_text, replace_text, self.case_btn.isChecked())

    def _toggle_replace(self):
        self._show_replace = not self._show_replace
        self.replace_row.setVisible(self._show_replace)
        self.setFixedHeight(70 if self._show_replace else 35)

    def _close(self):
        self.hide()
        self.closed.emit()

    def update_match_count(self, current: int, total: int):
        if total == 0:
            self.match_label.setText("No results")
            self.match_label.setStyleSheet(f"color: {Theme.WARNING}; font-size: 10px;")
        else:
            self.match_label.setText(f"{current}/{total}")
            self.match_label.setStyleSheet(f"color: {Theme.TEXT_PRIMARY}; font-size: 10px;")

    def show_find(self):
        """Show find widget and focus input."""
        self._show_replace = False
        self.replace_row.hide()
        self.setFixedHeight(35)
        self.show()
        self.find_input.setFocus()
        self.find_input.selectAll()

    def show_replace(self):
        """Show find/replace widget and focus input."""
        self._show_replace = True
        self.replace_row.show()
        self.setFixedHeight(70)
        self.show()
        self.find_input.setFocus()
        self.find_input.selectAll()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Escape:
            self._close()
        else:
            super().keyPressEvent(event)


class LineNumberArea(QWidget):
    """Professional line number gutter with VS Code-style features."""

    def __init__(self, editor: 'CodeEditor'):
        super().__init__(editor)
        self._editor = editor
        self._relative_numbers = False  # Can be toggled for vim-style relative numbers
        self._git_changes: Dict[int, str] = {}  # line -> 'added'|'modified'|'deleted'
        self._breakpoints: set = set()  # Lines with breakpoints
        self._hovered_line = -1

        # Enable mouse tracking for hover effects
        self.setMouseTracking(True)

    def sizeHint(self):
        return QSize(self._calculate_width(), 0)

    def _calculate_width(self):
        """Calculate width needed for line numbers with extra space for indicators."""
        digits = len(str(max(1, self._editor.blockCount())))
        # Width: left padding (4) + git indicator (4) + number area + right padding (8)
        number_width = self._editor.fontMetrics().horizontalAdvance('9') * digits
        return 16 + number_width + 8

    def set_relative_numbers(self, enabled: bool):
        """Toggle relative line numbers (vim-style)."""
        self._relative_numbers = enabled
        self.update()

    def set_git_changes(self, changes: Dict[int, str]):
        """Set git diff indicators for lines."""
        self._git_changes = changes
        self.update()

    def toggle_breakpoint(self, line: int):
        """Toggle breakpoint on a line."""
        if line in self._breakpoints:
            self._breakpoints.remove(line)
        else:
            self._breakpoints.add(line)
        self.update()

    def paintEvent(self, event):
        """Paint line numbers with enhanced visual features."""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # Background with subtle gradient
        bg_color = QColor(Theme.BG_SECONDARY)
        painter.fillRect(event.rect(), bg_color)

        # Right border separator
        border_color = QColor(Theme.BORDER)
        painter.setPen(QPen(border_color, 1))
        painter.drawLine(self.width() - 1, event.rect().top(),
                        self.width() - 1, event.rect().bottom())

        block = self._editor.firstVisibleBlock()
        block_number = block.blockNumber()
        top = int(self._editor.blockBoundingGeometry(block).translated(
            self._editor.contentOffset()).top())
        bottom = top + int(self._editor.blockBoundingRect(block).height())

        # Get current line for highlighting
        current_line = self._editor.textCursor().blockNumber()
        font_height = self._editor.fontMetrics().height()

        # Prepare fonts
        normal_font = self._editor.font()
        bold_font = QFont(normal_font)
        bold_font.setWeight(QFont.Weight.Bold)

        while block.isValid() and top <= event.rect().bottom():
            if block.isVisible() and bottom >= event.rect().top():
                line_num = block_number + 1

                # Calculate display number (relative or absolute)
                if self._relative_numbers and block_number != current_line:
                    display_num = abs(block_number - current_line)
                else:
                    display_num = line_num

                number = str(display_num)

                # Draw current line highlight background
                if block_number == current_line:
                    highlight_color = QColor(Theme.BG_HOVER)
                    painter.fillRect(0, top, self.width() - 1, font_height, highlight_color)

                # Draw git change indicator (left edge)
                if line_num in self._git_changes:
                    change_type = self._git_changes[line_num]
                    if change_type == 'added':
                        indicator_color = QColor("#4EC9B0")  # Green
                    elif change_type == 'modified':
                        indicator_color = QColor("#569CD6")  # Blue
                    elif change_type == 'deleted':
                        indicator_color = QColor("#F14C4C")  # Red
                    else:
                        indicator_color = QColor(Theme.TEXT_MUTED)
                    painter.fillRect(0, top, 3, font_height, indicator_color)

                # Draw breakpoint indicator
                if line_num in self._breakpoints:
                    bp_color = QColor("#E51400")  # Red breakpoint
                    bp_radius = 4
                    bp_x = 8
                    bp_y = top + (font_height - bp_radius * 2) // 2 + bp_radius
                    painter.setBrush(bp_color)
                    painter.setPen(Qt.PenStyle.NoPen)
                    painter.drawEllipse(QPointF(bp_x, bp_y), bp_radius, bp_radius)

                # Draw hover highlight
                if block_number == self._hovered_line and block_number != current_line:
                    hover_color = QColor(Theme.BG_TERTIARY)
                    painter.fillRect(0, top, self.width() - 1, font_height, hover_color)

                # Set font and color for line number
                if block_number == current_line:
                    painter.setPen(QColor(Theme.TEXT_PRIMARY))
                    painter.setFont(bold_font)
                else:
                    painter.setPen(QColor(Theme.TEXT_MUTED))
                    painter.setFont(normal_font)

                # Draw line number (right-aligned with padding)
                painter.drawText(0, top, self.width() - 8,
                               font_height,
                               Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter,
                               number)

            block = block.next()
            top = bottom
            bottom = top + int(self._editor.blockBoundingRect(block).height())
            block_number += 1

    def mouseMoveEvent(self, event):
        """Track mouse for hover effects."""
        # Calculate which line is being hovered
        block = self._editor.firstVisibleBlock()
        top = int(self._editor.blockBoundingGeometry(block).translated(
            self._editor.contentOffset()).top())

        while block.isValid():
            bottom = top + int(self._editor.blockBoundingRect(block).height())
            if top <= event.position().y() < bottom:
                if self._hovered_line != block.blockNumber():
                    self._hovered_line = block.blockNumber()
                    self.update()
                return
            top = bottom
            block = block.next()

        if self._hovered_line != -1:
            self._hovered_line = -1
            self.update()

    def leaveEvent(self, event):
        """Clear hover when mouse leaves."""
        if self._hovered_line != -1:
            self._hovered_line = -1
            self.update()

    def mousePressEvent(self, event):
        """Handle clicks for breakpoints."""
        if event.button() == Qt.MouseButton.LeftButton:
            # Calculate which line was clicked
            block = self._editor.firstVisibleBlock()
            top = int(self._editor.blockBoundingGeometry(block).translated(
                self._editor.contentOffset()).top())

            while block.isValid():
                bottom = top + int(self._editor.blockBoundingRect(block).height())
                if top <= event.position().y() < bottom:
                    # Left area click = toggle breakpoint
                    if event.position().x() < 16:
                        self.toggle_breakpoint(block.blockNumber() + 1)
                    else:
                        # Click on line number = go to line
                        cursor = self._editor.textCursor()
                        cursor.setPosition(block.position())
                        self._editor.setTextCursor(cursor)
                    return
                top = bottom
                block = block.next()


class FoldingArea(QWidget):
    """Code folding margin showing fold markers for collapsible regions."""

    def __init__(self, editor: 'CodeEditor'):
        super().__init__(editor)
        self._editor = editor
        self._fold_regions: Dict[int, int] = {}  # start_line -> end_line
        self._collapsed: set = set()  # Set of collapsed start lines

        self.setFixedWidth(16)
        self.setCursor(Qt.CursorShape.PointingHandCursor)

        # Connect to editor signals
        self._editor.blockCountChanged.connect(self._update_folds)
        self._editor.updateRequest.connect(self._update_area)
        self._editor.textChanged.connect(self._update_folds)

    def _update_folds(self, block_count=None):
        """Detect foldable regions based on indentation."""
        self._fold_regions.clear()
        doc = self._editor.document()
        block = doc.begin()
        stack = []  # Stack of (start_line, indent_level)

        while block.isValid():
            text = block.text()
            line_num = block.blockNumber()

            if text.strip():
                # Calculate indent level
                indent = len(text) - len(text.lstrip())
                indent_level = indent // 4  # Assume 4-space tabs

                # Pop stack for regions that end
                while stack and stack[-1][1] >= indent_level:
                    start_line, _ = stack.pop()
                    if line_num > start_line + 1:
                        self._fold_regions[start_line] = line_num - 1

                # Check if this line starts a fold region (ends with : or { or has high indent)
                stripped = text.rstrip()
                if stripped.endswith(':') or stripped.endswith('{'):
                    stack.append((line_num, indent_level))

            block = block.next()

        # Close any remaining regions at end of file
        total_lines = doc.blockCount()
        while stack:
            start_line, _ = stack.pop()
            if total_lines > start_line + 1:
                self._fold_regions[start_line] = total_lines - 1

        self.update()

    def _update_area(self, rect, dy):
        """Update folding area when editor scrolls."""
        if dy:
            self.scroll(0, dy)
        else:
            self.update(0, rect.y(), self.width(), rect.height())

    def paintEvent(self, event):
        """Paint fold markers."""
        painter = QPainter(self)
        painter.fillRect(event.rect(), QColor(Theme.BG_SECONDARY))

        block = self._editor.firstVisibleBlock()
        block_number = block.blockNumber()
        top = int(self._editor.blockBoundingGeometry(block).translated(
            self._editor.contentOffset()).top())
        bottom = top + int(self._editor.blockBoundingRect(block).height())

        while block.isValid() and top <= event.rect().bottom():
            if block.isVisible() and bottom >= event.rect().top():
                line_num = block_number

                if line_num in self._fold_regions:
                    # Draw fold marker
                    is_collapsed = line_num in self._collapsed
                    self._draw_fold_marker(painter, top, is_collapsed)

            block = block.next()
            top = bottom
            bottom = top + int(self._editor.blockBoundingRect(block).height())
            block_number += 1

    def _draw_fold_marker(self, painter: QPainter, y: int, collapsed: bool):
        """Draw a fold/unfold marker."""
        size = 9
        x = (self.width() - size) // 2
        center_y = y + (self._editor.fontMetrics().height() - size) // 2

        # Draw box
        painter.setPen(QPen(QColor(Theme.TEXT_MUTED), 1))
        painter.setBrush(QColor(Theme.BG_TERTIARY))
        painter.drawRect(x, center_y, size, size)

        # Draw + or -
        painter.setPen(QPen(QColor(Theme.TEXT_PRIMARY), 1))
        mid = size // 2
        painter.drawLine(x + 2, center_y + mid, x + size - 2, center_y + mid)  # Horizontal
        if collapsed:
            painter.drawLine(x + mid, center_y + 2, x + mid, center_y + size - 2)  # Vertical

    def mousePressEvent(self, event):
        """Handle click to toggle fold."""
        if event.button() == Qt.MouseButton.LeftButton:
            # Find which line was clicked
            y = event.position().y()
            block = self._editor.firstVisibleBlock()
            top = int(self._editor.blockBoundingGeometry(block).translated(
                self._editor.contentOffset()).top())

            while block.isValid():
                height = int(self._editor.blockBoundingRect(block).height())
                if top <= y < top + height:
                    line_num = block.blockNumber()
                    if line_num in self._fold_regions:
                        self._toggle_fold(line_num)
                    break
                top += height
                block = block.next()

    def _toggle_fold(self, line_num: int):
        """Toggle fold state for a region."""
        if line_num in self._collapsed:
            self._unfold(line_num)
        else:
            self._fold(line_num)

    def _fold(self, start_line: int):
        """Collapse a fold region."""
        if start_line not in self._fold_regions:
            return

        end_line = self._fold_regions[start_line]
        self._collapsed.add(start_line)

        # Hide blocks
        doc = self._editor.document()
        block = doc.findBlockByNumber(start_line + 1)
        while block.isValid() and block.blockNumber() <= end_line:
            block.setVisible(False)
            block = block.next()

        self._editor.viewport().update()
        self.update()

    def _unfold(self, start_line: int):
        """Expand a fold region."""
        if start_line not in self._fold_regions:
            return

        end_line = self._fold_regions[start_line]
        self._collapsed.discard(start_line)

        # Show blocks
        doc = self._editor.document()
        block = doc.findBlockByNumber(start_line + 1)
        while block.isValid() and block.blockNumber() <= end_line:
            block.setVisible(True)
            block = block.next()

        self._editor.viewport().update()
        self.update()

    def fold_all(self):
        """Collapse all fold regions."""
        for start_line in self._fold_regions:
            self._fold(start_line)

    def unfold_all(self):
        """Expand all fold regions."""
        for start_line in list(self._collapsed):
            self._unfold(start_line)


class GitBlameGutter(QWidget):
    """Gutter widget showing git blame information for each line."""

    def __init__(self, editor: 'CodeEditor'):
        super().__init__(editor)
        self._editor = editor
        self._blame_data: Dict[int, dict] = {}  # line -> {author, date, commit}
        self._visible = False
        self._loading = False

        self.setFixedWidth(0)  # Hidden by default

        # Connect to editor signals
        self._editor.blockCountChanged.connect(self._update_width)
        self._editor.updateRequest.connect(self._update_area)

    def set_visible(self, visible: bool):
        """Show or hide the blame gutter."""
        self._visible = visible
        if visible:
            self._update_width()
        else:
            self.setFixedWidth(0)
        self.update()

    def load_blame(self, file_path: Path):
        """Load git blame data for the file."""
        if not file_path or not file_path.exists():
            return

        self._loading = True
        self._blame_data.clear()
        self.update()

        try:
            result = subprocess.run(
                ["git", "blame", "--line-porcelain", str(file_path)],
                cwd=str(file_path.parent),
                capture_output=True,
                text=True,
                timeout=10
            )

            if result.returncode == 0:
                self._parse_blame(result.stdout)
        except Exception:
            pass

        self._loading = False
        self._update_width()
        self.update()

    def _parse_blame(self, output: str):
        """Parse git blame --line-porcelain output."""
        lines = output.split('\n')
        current_line = 0
        current_data = {}

        for line in lines:
            if line.startswith('\t'):
                # Content line - save the data
                current_line += 1
                self._blame_data[current_line] = current_data.copy()
            elif line.startswith('author '):
                current_data['author'] = line[7:]
            elif line.startswith('author-time '):
                try:
                    timestamp = int(line[12:])
                    from datetime import datetime
                    dt = datetime.fromtimestamp(timestamp)
                    current_data['date'] = dt.strftime('%Y-%m-%d')
                except (ValueError, OSError):
                    current_data['date'] = ''
            elif len(line) >= 40 and line[:40].replace('0', '').replace('a', '').replace('b', '').replace('c', '').replace('d', '').replace('e', '').replace('f', '') == '':
                # Commit hash
                current_data['commit'] = line[:8]

    def _update_width(self, block_count=None):
        """Update gutter width based on content."""
        if not self._visible:
            self.setFixedWidth(0)
            return

        if self._blame_data:
            # Calculate width needed for blame info
            sample = next(iter(self._blame_data.values()), {})
            author = sample.get('author', '')[:12]
            date = sample.get('date', '')
            text = f"{author} {date}"
            width = self.fontMetrics().horizontalAdvance(text) + 20
            self.setFixedWidth(max(width, 150))
        else:
            self.setFixedWidth(100)

    def _update_area(self, rect, dy):
        """Update gutter area when editor scrolls."""
        if dy:
            self.scroll(0, dy)
        else:
            self.update(0, rect.y(), self.width(), rect.height())

    def paintEvent(self, event):
        """Paint the blame gutter."""
        if not self._visible:
            return

        painter = QPainter(self)
        painter.fillRect(event.rect(), QColor(Theme.BG_SECONDARY))

        if self._loading:
            painter.setPen(QColor(Theme.TEXT_MUTED))
            painter.drawText(event.rect(), Qt.AlignmentFlag.AlignCenter, "Loading...")
            return

        # Set font
        font = self._editor.font()
        font.setPointSize(font.pointSize() - 1)
        painter.setFont(font)

        block = self._editor.firstVisibleBlock()
        block_number = block.blockNumber()
        top = int(self._editor.blockBoundingGeometry(block).translated(
            self._editor.contentOffset()).top())
        bottom = top + int(self._editor.blockBoundingRect(block).height())

        last_commit = None

        while block.isValid() and top <= event.rect().bottom():
            if block.isVisible() and bottom >= event.rect().top():
                line_num = block_number + 1
                data = self._blame_data.get(line_num, {})

                if data:
                    commit = data.get('commit', '')
                    author = data.get('author', '')[:12]
                    date = data.get('date', '')

                    # Only show if different from previous line (cleaner look)
                    if commit != last_commit:
                        painter.setPen(QColor(Theme.TEXT_MUTED))
                        text = f"{author:<12} {date}"
                        painter.drawText(5, top, self.width() - 10,
                                        self._editor.fontMetrics().height(),
                                        Qt.AlignmentFlag.AlignLeft, text)
                        last_commit = commit

            block = block.next()
            top = bottom
            bottom = top + int(self._editor.blockBoundingRect(block).height())
            block_number += 1


class CodeEditor(QPlainTextEdit):
    """Code editor with syntax highlighting and find/replace."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_file: Optional[Path] = None
        self._find_matches: List[int] = []  # List of match positions
        self._current_match = -1
        self._blame_visible = False

        font = QFont("Consolas", 12)
        font.setStyleHint(QFont.StyleHint.Monospace)
        self.setFont(font)

        self.setStyleSheet(f"""
            QPlainTextEdit {{
                background: {Theme.BG_MAIN};
                color: {Theme.TEXT_PRIMARY};
                border: none;
                selection-background-color: {Theme.BG_SELECTED};
            }}
        """)

        self.setTabStopDistance(self.fontMetrics().horizontalAdvance(' ') * 4)
        self.setLineWrapMode(QPlainTextEdit.LineWrapMode.NoWrap)

        # Use Pygments highlighter if available, otherwise fallback
        if PYGMENTS_AVAILABLE:
            self.highlighter = PygmentsHighlighter(self.document())
        else:
            self.highlighter = PythonHighlighter(self.document())

        # Line number area
        self.line_number_area = LineNumberArea(self)

        # Code folding area
        self.folding_area = FoldingArea(self)

        # Git blame gutter
        self.blame_gutter = GitBlameGutter(self)

        # Update margins and geometry
        self._update_margins()

        # Connect signals for updates
        self.updateRequest.connect(self._update_gutter_geometry)
        self.blockCountChanged.connect(self._update_line_number_width)
        self.cursorPositionChanged.connect(self._highlight_current_line)

    def _update_line_number_width(self, *args):
        """Update line number area width when block count changes."""
        self._update_margins()

    def _highlight_current_line(self):
        """Update line number area when cursor moves."""
        self.line_number_area.update()

    def _update_gutter_geometry(self, rect, dy):
        """Update the position of gutter widgets."""
        if dy:
            self.line_number_area.scroll(0, dy)
            self.folding_area.scroll(0, dy)
        else:
            self.line_number_area.update(0, rect.y(), self.line_number_area.width(), rect.height())
            self.folding_area.update(0, rect.y(), self.folding_area.width(), rect.height())

    def resizeEvent(self, event):
        """Handle resize to update gutter geometry."""
        super().resizeEvent(event)
        cr = self.contentsRect()
        line_num_width = self.line_number_area._calculate_width()
        fold_width = self.folding_area.width()

        # Line numbers on the left
        self.line_number_area.setGeometry(cr.left(), cr.top(), line_num_width, cr.height())
        # Folding area next to line numbers
        self.folding_area.setGeometry(cr.left() + line_num_width, cr.top(), fold_width, cr.height())

    def load_file(self, path: Path):
        try:
            # File size protection
            file_size = path.stat().st_size
            size_mb = file_size / (1024 * 1024)

            if size_mb > 50:
                self.setPlainText(f"File too large to open ({size_mb:.1f} MB)\n\nMaximum file size: 50 MB")
                return

            if size_mb > 5:
                # Show warning for large files
                from PySide6.QtWidgets import QMessageBox
                reply = QMessageBox.warning(
                    self,
                    "Large File Warning",
                    f"This file is {size_mb:.1f} MB.\n\nLarge files may slow down the editor.\n\nDo you want to continue?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )
                if reply != QMessageBox.StandardButton.Yes:
                    return

            content = path.read_text()
            self.setPlainText(content)
            self.current_file = path

            # Update syntax highlighter for file type
            if PYGMENTS_AVAILABLE and hasattr(self.highlighter, 'set_language_from_file'):
                self.highlighter.set_language_from_file(str(path))
        except Exception as e:
            self.setPlainText(f"Error loading file: {e}")

    def save_file(self):
        if self.current_file:
            self.current_file.write_text(self.toPlainText())

    def toggle_blame(self):
        """Toggle git blame gutter visibility."""
        self._blame_visible = not self._blame_visible
        self.blame_gutter.set_visible(self._blame_visible)

        if self._blame_visible and self.current_file:
            self.blame_gutter.load_blame(self.current_file)

        self._update_margins()

    def _update_margins(self):
        """Update viewport margins for line numbers, folding area and blame gutter."""
        left_margin = self.line_number_area._calculate_width() + self.folding_area.width()
        if self._blame_visible:
            left_margin += self.blame_gutter.width()
        self.setViewportMargins(left_margin, 0, 0, 0)

    def is_blame_visible(self) -> bool:
        """Check if git blame is currently visible."""
        return self._blame_visible

    def find_text(self, text: str, case_sensitive: bool = False, forward: bool = True) -> tuple:
        """Find text and return (current_index, total_matches)."""
        if not text:
            self._find_matches = []
            self._current_match = -1
            return (0, 0)

        # Build regex flags
        flags = 0 if case_sensitive else re.IGNORECASE

        # Find all matches
        content = self.toPlainText()
        self._find_matches = [m.start() for m in re.finditer(re.escape(text), content, flags)]

        if not self._find_matches:
            self._current_match = -1
            return (0, 0)

        # Find next match from cursor
        cursor_pos = self.textCursor().position()

        if forward:
            # Find first match after cursor
            for i, pos in enumerate(self._find_matches):
                if pos >= cursor_pos:
                    self._current_match = i
                    break
            else:
                self._current_match = 0  # Wrap to start
        else:
            # Find last match before cursor
            for i in range(len(self._find_matches) - 1, -1, -1):
                if self._find_matches[i] < cursor_pos:
                    self._current_match = i
                    break
            else:
                self._current_match = len(self._find_matches) - 1  # Wrap to end

        # Select the match
        self._select_match(text)
        return (self._current_match + 1, len(self._find_matches))

    def find_next(self, text: str, case_sensitive: bool = False) -> tuple:
        """Find next occurrence."""
        if not self._find_matches or not text:
            return self.find_text(text, case_sensitive, forward=True)

        self._current_match = (self._current_match + 1) % len(self._find_matches)
        self._select_match(text)
        return (self._current_match + 1, len(self._find_matches))

    def find_prev(self, text: str, case_sensitive: bool = False) -> tuple:
        """Find previous occurrence."""
        if not self._find_matches or not text:
            return self.find_text(text, case_sensitive, forward=False)

        self._current_match = (self._current_match - 1) % len(self._find_matches)
        self._select_match(text)
        return (self._current_match + 1, len(self._find_matches))

    def _select_match(self, text: str):
        """Select the current match."""
        if self._current_match >= 0 and self._current_match < len(self._find_matches):
            pos = self._find_matches[self._current_match]
            cursor = self.textCursor()
            cursor.setPosition(pos)
            cursor.setPosition(pos + len(text), QTextCursor.MoveMode.KeepAnchor)
            self.setTextCursor(cursor)
            self.centerCursor()

    def replace_current(self, find_text: str, replace_text: str, case_sensitive: bool = False) -> tuple:
        """Replace current selection if it matches."""
        cursor = self.textCursor()
        if cursor.hasSelection():
            selected = cursor.selectedText()
            match = selected == find_text if case_sensitive else selected.lower() == find_text.lower()
            if match:
                cursor.insertText(replace_text)
                # Re-find to update matches
                return self.find_text(find_text, case_sensitive, forward=True)
        return (self._current_match + 1 if self._current_match >= 0 else 0, len(self._find_matches))

    def replace_all(self, find_text: str, replace_text: str, case_sensitive: bool = False) -> int:
        """Replace all occurrences. Returns count of replacements."""
        if not find_text:
            return 0

        flags = 0 if case_sensitive else re.IGNORECASE
        content = self.toPlainText()
        new_content, count = re.subn(re.escape(find_text), replace_text, content, flags=flags)

        if count > 0:
            cursor = self.textCursor()
            cursor.beginEditBlock()
            self.setPlainText(new_content)
            cursor.endEditBlock()
            self._find_matches = []
            self._current_match = -1

        return count


class Minimap(QWidget):
    """Professional VS Code-style minimap with syntax highlighting and enhanced features."""

    clicked = Signal(int)  # line number clicked

    # Syntax color mapping for minimap (simplified, based on token type)
    SYNTAX_COLORS = {
        'keyword': '#569CD6',      # Blue - keywords
        'string': '#CE9178',       # Orange - strings
        'comment': '#6A9955',      # Green - comments
        'number': '#B5CEA8',       # Light green - numbers
        'function': '#DCDCAA',     # Yellow - functions
        'class': '#4EC9B0',        # Cyan - classes
        'operator': '#D4D4D4',     # White - operators
        'default': '#9CDCFE',      # Light blue - default
    }

    def __init__(self, editor: QPlainTextEdit, parent=None):
        super().__init__(parent)
        self._editor = editor
        self._char_width = 0.8  # Scale factor for character width
        self._line_height = 2   # Height of each line in minimap
        self._viewport_top = 0
        self._viewport_bottom = 0
        self._dragging = False
        self._hover_line = -1
        self._search_highlights: List[int] = []  # Lines with search matches
        self._git_changes: Dict[int, str] = {}   # line -> 'added'|'modified'|'deleted'
        self._diagnostics: Dict[int, str] = {}   # line -> 'error'|'warning'|'info'
        self._selection_start = -1
        self._selection_end = -1
        self._cached_pixmap = None
        self._cache_valid = False

        self.setFixedWidth(90)
        self.setMinimumHeight(100)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setMouseTracking(True)

        # Connect to editor signals
        self._editor.textChanged.connect(self._invalidate_cache)
        self._editor.verticalScrollBar().valueChanged.connect(self._update_viewport)
        self._editor.cursorPositionChanged.connect(self._on_cursor_changed)
        self._editor.selectionChanged.connect(self._on_selection_changed)

    def _invalidate_cache(self):
        """Invalidate the cached rendering."""
        self._cache_valid = False
        self.update()

    def _on_cursor_changed(self):
        """Handle cursor position change."""
        self.update()

    def _on_selection_changed(self):
        """Track selection for highlighting."""
        cursor = self._editor.textCursor()
        if cursor.hasSelection():
            start = cursor.document().findBlock(cursor.selectionStart()).blockNumber()
            end = cursor.document().findBlock(cursor.selectionEnd()).blockNumber()
            self._selection_start = start
            self._selection_end = end
        else:
            self._selection_start = -1
            self._selection_end = -1
        self.update()

    def set_search_highlights(self, lines: List[int]):
        """Set lines that contain search matches."""
        self._search_highlights = lines
        self._invalidate_cache()

    def set_git_changes(self, changes: Dict[int, str]):
        """Set git diff indicators."""
        self._git_changes = changes
        self._invalidate_cache()

    def set_diagnostics(self, diagnostics: Dict[int, str]):
        """Set error/warning indicators. diagnostics maps line -> 'error'|'warning'|'info'"""
        self._diagnostics = diagnostics
        self._invalidate_cache()

    def _update_viewport(self):
        """Update the viewport indicator position."""
        scrollbar = self._editor.verticalScrollBar()
        total_lines = self._editor.document().blockCount()
        if total_lines == 0:
            self._viewport_top = 0
            self._viewport_bottom = self.height()
        else:
            visible_lines = max(1, self._editor.viewport().height() // max(1, self._editor.fontMetrics().height()))
            first_visible = scrollbar.value()

            # Calculate viewport position in minimap coordinates
            max_y = self._get_content_height()
            if max_y > 0:
                self._viewport_top = int((first_visible / max(total_lines, 1)) * max_y)
                self._viewport_bottom = int(((first_visible + visible_lines) / max(total_lines, 1)) * max_y)
                self._viewport_bottom = min(self._viewport_bottom, max_y)
            else:
                self._viewport_top = 0
                self._viewport_bottom = self.height()

        self.update()

    def _get_content_height(self) -> int:
        """Get the total content height in minimap coordinates."""
        return self._editor.document().blockCount() * (self._line_height + 1)

    def _classify_token(self, text: str, pos: int) -> str:
        """Simple token classification for syntax coloring."""
        # Check if we're in a string
        in_string = False
        string_char = None
        for i, char in enumerate(text[:pos]):
            if char in ('"', "'") and (i == 0 or text[i-1] != '\\'):
                if not in_string:
                    in_string = True
                    string_char = char
                elif char == string_char:
                    in_string = False
        if in_string:
            return 'string'

        # Check for comment
        stripped = text.lstrip()
        if stripped.startswith('#') or stripped.startswith('//'):
            return 'comment'

        # Extract the word at position
        start = pos
        while start > 0 and (text[start-1].isalnum() or text[start-1] == '_'):
            start -= 1
        end = pos
        while end < len(text) and (text[end].isalnum() or text[end] == '_'):
            end += 1
        word = text[start:end]

        # Python/JS keywords
        keywords = {'def', 'class', 'if', 'else', 'elif', 'for', 'while', 'try', 'except',
                   'finally', 'with', 'as', 'import', 'from', 'return', 'yield', 'raise',
                   'break', 'continue', 'pass', 'lambda', 'and', 'or', 'not', 'in', 'is',
                   'function', 'const', 'let', 'var', 'async', 'await', 'export', 'default'}
        if word in keywords:
            return 'keyword'

        # Check if it looks like a function call
        if end < len(text) and text[end] == '(':
            return 'function'

        # Check if it looks like a class name (starts with uppercase)
        if word and word[0].isupper():
            return 'class'

        # Check for numbers
        if word.isdigit() or (word.startswith('0x') and len(word) > 2):
            return 'number'

        return 'default'

    def paintEvent(self, event):
        """Paint the enhanced minimap."""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # Background with subtle left border
        painter.fillRect(self.rect(), QColor(Theme.BG_DARK))

        # Left border separator
        painter.setPen(QPen(QColor(Theme.BORDER), 1))
        painter.drawLine(0, 0, 0, self.height())

        # Draw code representation
        doc = self._editor.document()
        block = doc.begin()
        y = 2  # Top padding

        cursor_line = self._editor.textCursor().blockNumber()
        total_lines = doc.blockCount()

        # Calculate scale factor to fit content
        content_height = total_lines * (self._line_height + 1)
        scale = min(1.0, (self.height() - 4) / max(content_height, 1))

        while block.isValid() and y < self.height():
            line_num = block.blockNumber()
            text = block.text()

            # Git change indicator on left edge
            if line_num + 1 in self._git_changes:
                change_type = self._git_changes[line_num + 1]
                if change_type == 'added':
                    git_color = QColor("#4EC9B0")
                elif change_type == 'modified':
                    git_color = QColor("#569CD6")
                else:
                    git_color = QColor("#F14C4C")
                painter.fillRect(1, int(y), 2, self._line_height, git_color)

            # Diagnostic indicator on right edge (errors=red, warnings=yellow, info=blue)
            if line_num + 1 in self._diagnostics:
                diag_type = self._diagnostics[line_num + 1]
                if diag_type == 'error':
                    diag_color = QColor("#F14C4C")  # Red for errors
                elif diag_type == 'warning':
                    diag_color = QColor("#CCA700")  # Yellow/orange for warnings
                else:
                    diag_color = QColor("#3794FF")  # Blue for info/hints
                # Draw on right edge as a small rectangle
                painter.fillRect(self.width() - 6, int(y), 4, self._line_height, diag_color)

            # Search highlight (full width yellow background)
            if line_num + 1 in self._search_highlights:
                painter.fillRect(4, int(y), self.width() - 8, self._line_height,
                               QColor(255, 200, 0, 60))

            # Selection highlight
            if self._selection_start >= 0 and self._selection_start <= line_num <= self._selection_end:
                painter.fillRect(4, int(y), self.width() - 8, self._line_height,
                               QColor(Theme.ACCENT_BLUE).lighter(150))

            # Current line highlight
            if line_num == cursor_line:
                painter.fillRect(4, int(y) - 1, self.width() - 8, self._line_height + 2,
                               QColor(Theme.ACCENT_BLUE))

            # Hover line highlight
            if line_num == self._hover_line and line_num != cursor_line:
                painter.fillRect(4, int(y), self.width() - 8, self._line_height,
                               QColor(255, 255, 255, 20))

            # Draw code with syntax coloring
            if text.strip():
                indent = len(text) - len(text.lstrip())
                x = 6 + int(indent * self._char_width)

                # Analyze text and draw colored segments
                segments = self._tokenize_line(text.strip())
                for seg_text, token_type in segments:
                    color = QColor(self.SYNTAX_COLORS.get(token_type, self.SYNTAX_COLORS['default']))
                    if line_num != cursor_line:
                        color.setAlpha(180)  # Slightly transparent for non-current lines
                    seg_width = int(len(seg_text) * self._char_width)
                    if seg_width > 0 and x < self.width() - 4:
                        painter.fillRect(int(x), int(y), min(seg_width, self.width() - x - 4),
                                       self._line_height, color)
                    x += seg_width

            y += self._line_height + 1
            block = block.next()

        # Draw viewport indicator
        max_y = y
        if total_lines > 0:
            scrollbar = self._editor.verticalScrollBar()
            visible_lines = max(1, self._editor.viewport().height() // max(1, self._editor.fontMetrics().height()))
            first_visible = scrollbar.value()

            viewport_top = int((first_visible / max(total_lines, 1)) * max_y)
            viewport_height = int((visible_lines / max(total_lines, 1)) * max_y)
            viewport_height = max(viewport_height, 20)  # Minimum height
            viewport_top = min(viewport_top, max_y - viewport_height)

            # Viewport background with gradient
            viewport_rect = QRect(2, viewport_top, self.width() - 4, viewport_height)

            # Create gradient for viewport
            gradient = QLinearGradient(viewport_rect.left(), 0, viewport_rect.right(), 0)
            gradient.setColorAt(0.0, QColor(255, 255, 255, 15))
            gradient.setColorAt(0.5, QColor(255, 255, 255, 25))
            gradient.setColorAt(1.0, QColor(255, 255, 255, 15))
            painter.fillRect(viewport_rect, gradient)

            # Viewport border with accent color
            painter.setPen(QPen(QColor(Theme.ACCENT_BLUE), 1))
            painter.drawRect(viewport_rect.adjusted(0, 0, -1, -1))

            # Store for mouse handling
            self._viewport_top = viewport_top
            self._viewport_bottom = viewport_top + viewport_height

    def _tokenize_line(self, text: str) -> List[tuple]:
        """Simple tokenization for syntax coloring."""
        if not text:
            return []

        segments = []

        # Check for full-line comment
        if text.startswith('#') or text.startswith('//'):
            return [(text, 'comment')]

        # Simple tokenization
        i = 0
        current_segment = ""
        current_type = "default"

        while i < len(text):
            char = text[i]

            # String detection
            if char in ('"', "'"):
                if current_segment:
                    segments.append((current_segment, current_type))
                quote = char
                string_content = char
                i += 1
                while i < len(text):
                    string_content += text[i]
                    if text[i] == quote and (len(string_content) < 2 or text[i-1] != '\\'):
                        i += 1
                        break
                    i += 1
                segments.append((string_content, 'string'))
                current_segment = ""
                current_type = "default"
                continue

            # Word characters
            if char.isalnum() or char == '_':
                current_segment += char
            else:
                if current_segment:
                    # Classify the word
                    word_type = self._classify_word(current_segment, text, i - len(current_segment))
                    segments.append((current_segment, word_type))
                    current_segment = ""

                # Add the non-word character
                if char.strip():
                    segments.append((char, 'operator'))

            i += 1

        # Add remaining segment
        if current_segment:
            word_type = self._classify_word(current_segment, text, len(text) - len(current_segment))
            segments.append((current_segment, word_type))

        return segments

    def _classify_word(self, word: str, line: str, pos: int) -> str:
        """Classify a word for syntax coloring."""
        keywords = {'def', 'class', 'if', 'else', 'elif', 'for', 'while', 'try', 'except',
                   'finally', 'with', 'as', 'import', 'from', 'return', 'yield', 'raise',
                   'break', 'continue', 'pass', 'lambda', 'and', 'or', 'not', 'in', 'is',
                   'True', 'False', 'None', 'self', 'function', 'const', 'let', 'var',
                   'async', 'await', 'export', 'default', 'new', 'this'}

        if word in keywords:
            return 'keyword'

        # Check if followed by ( -> function
        after_word = line[pos + len(word):].lstrip()
        if after_word.startswith('('):
            return 'function'

        # Check for class names (PascalCase)
        if word and word[0].isupper() and not word.isupper():
            return 'class'

        # Numbers
        if word.isdigit() or word.replace('.', '').replace('e', '').replace('-', '').isdigit():
            return 'number'

        return 'default'

    def mousePressEvent(self, event):
        """Handle mouse press - start dragging or jump to position."""
        if event.button() == Qt.MouseButton.LeftButton:
            self._dragging = True
            self._scroll_to_position(event.position().y())

    def mouseMoveEvent(self, event):
        """Handle mouse move - drag scrolling or hover."""
        if self._dragging:
            self._scroll_to_position(event.position().y())
        else:
            # Update hover line
            total_lines = self._editor.document().blockCount()
            if total_lines > 0:
                content_height = total_lines * (self._line_height + 1) + 4
                effective_height = min(content_height, self.height())
                y = max(0, min(event.position().y(), effective_height))
                line = int((y / effective_height) * total_lines)
                line = max(0, min(line, total_lines - 1))
                if line != self._hover_line:
                    self._hover_line = line
                    self.update()

    def mouseReleaseEvent(self, event):
        """Handle mouse release - stop dragging."""
        if event.button() == Qt.MouseButton.LeftButton:
            self._dragging = False

    def leaveEvent(self, event):
        """Clear hover when mouse leaves."""
        if self._hover_line != -1:
            self._hover_line = -1
            self.update()

    def _scroll_to_position(self, y: float):
        """Scroll the editor to the position corresponding to y coordinate."""
        total_lines = self._editor.document().blockCount()
        if total_lines == 0:
            return

        # Calculate the actual minimap content height
        content_height = total_lines * (self._line_height + 1) + 4  # +4 for padding

        # Use the smaller of content height or widget height for ratio calculation
        effective_height = min(content_height, self.height())

        # Clamp y to valid range
        y = max(0, min(y, effective_height))

        # Calculate target line - direct linear mapping
        target_line = int((y / effective_height) * total_lines)
        target_line = max(0, min(target_line, total_lines - 1))

        # Calculate how many lines are visible in the editor
        visible_lines = max(1, self._editor.viewport().height() // max(1, self._editor.fontMetrics().height()))

        # Center the target line in the viewport
        scroll_target = max(0, target_line - visible_lines // 2)
        scroll_target = min(scroll_target, max(0, total_lines - visible_lines))

        # Scroll editor immediately
        scrollbar = self._editor.verticalScrollBar()
        scrollbar.setValue(scroll_target)

    def resizeEvent(self, event):
        """Handle resize - update viewport."""
        super().resizeEvent(event)
        self._invalidate_cache()
        self._update_viewport()


class LinterWorker(QThread):
    """Background worker for running linters on Python files."""

    diagnostics_ready = Signal(dict)  # line -> 'error'|'warning'|'info'

    # Linter preference order (pylint = most thorough but slower)
    LINTERS = ['pylint', 'ruff', 'flake8']

    def __init__(self, file_path: str, content: str = None, parent=None):
        super().__init__(parent)
        self._file_path = file_path
        self._content = content  # Optional: lint from content instead of file
        self._linter = None

    def _detect_linter(self) -> Optional[str]:
        """Detect available linter."""
        import shutil
        for linter in self.LINTERS:
            if shutil.which(linter):
                return linter
        return None

    def run(self):
        """Run linter and emit diagnostics."""
        if not self._file_path.endswith('.py'):
            self.diagnostics_ready.emit({})
            return

        linter = self._detect_linter()
        if not linter:
            self.diagnostics_ready.emit({})
            return

        self._linter = linter
        diagnostics = {}

        try:
            if linter == 'ruff':
                diagnostics = self._run_ruff()
            elif linter == 'flake8':
                diagnostics = self._run_flake8()
            elif linter == 'pylint':
                diagnostics = self._run_pylint()
        except Exception:
            pass

        self.diagnostics_ready.emit(diagnostics)

    def _run_ruff(self) -> Dict[int, str]:
        """Run ruff linter (fastest, Rust-based)."""
        import subprocess
        diagnostics = {}

        try:
            cmd = ['ruff', 'check', '--output-format', 'json', self._file_path]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)

            if result.stdout:
                issues = json.loads(result.stdout)
                for issue in issues:
                    line = issue.get('location', {}).get('row', 0)
                    code = issue.get('code', '')
                    # E/W prefix = error/warning, others = info
                    if code.startswith('E') or code.startswith('F'):
                        severity = 'error'
                    elif code.startswith('W'):
                        severity = 'warning'
                    else:
                        severity = 'info'
                    # Keep highest severity per line
                    if line not in diagnostics or self._severity_rank(severity) > self._severity_rank(diagnostics[line]):
                        diagnostics[line] = severity
        except Exception:
            pass

        return diagnostics

    def _run_flake8(self) -> Dict[int, str]:
        """Run flake8 linter."""
        import subprocess
        diagnostics = {}

        try:
            cmd = ['flake8', '--format', '%(row)d:%(col)d:%(code)s:%(text)s', self._file_path]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)

            for line_text in result.stdout.strip().split('\n'):
                if ':' in line_text:
                    parts = line_text.split(':')
                    if len(parts) >= 3:
                        try:
                            line = int(parts[0])
                            code = parts[2] if len(parts) > 2 else ''
                            # E = error, W = warning, C/F = convention/fatal
                            if code.startswith('E') or code.startswith('F'):
                                severity = 'error'
                            elif code.startswith('W'):
                                severity = 'warning'
                            else:
                                severity = 'info'
                            if line not in diagnostics or self._severity_rank(severity) > self._severity_rank(diagnostics[line]):
                                diagnostics[line] = severity
                        except ValueError:
                            pass
        except Exception:
            pass

        return diagnostics

    def _run_pylint(self) -> Dict[int, str]:
        """Run pylint linter."""
        import subprocess
        diagnostics = {}

        try:
            cmd = ['pylint', '--output-format', 'json', '--score', 'n', self._file_path]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)

            if result.stdout:
                try:
                    issues = json.loads(result.stdout)
                    for issue in issues:
                        line = issue.get('line', 0)
                        msg_type = issue.get('type', '').lower()
                        # Map pylint types to our severity
                        if msg_type in ('error', 'fatal'):
                            severity = 'error'
                        elif msg_type == 'warning':
                            severity = 'warning'
                        else:
                            severity = 'info'
                        if line not in diagnostics or self._severity_rank(severity) > self._severity_rank(diagnostics[line]):
                            diagnostics[line] = severity
                except json.JSONDecodeError:
                    pass
        except Exception:
            pass

        return diagnostics

    def _severity_rank(self, severity: str) -> int:
        """Get numeric rank for severity (higher = more severe)."""
        return {'info': 1, 'warning': 2, 'error': 3}.get(severity, 0)


class EditorWithFindReplace(QWidget):
    """Wrapper widget that combines code editor with find/replace functionality."""

    cursor_position_changed = Signal(int, int)  # line, column
    modification_changed = Signal(bool)  # is_modified

    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_file: Optional[Path] = None
        self._is_modified = False
        self._linter_worker: Optional[LinterWorker] = None
        self._lint_timer: Optional[QTimer] = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Find/Replace widget (initially hidden)
        self.find_widget = FindReplaceWidget()
        self.find_widget.hide()
        layout.addWidget(self.find_widget)

        # Editor area with minimap
        editor_container = QWidget()
        editor_layout = QHBoxLayout(editor_container)
        editor_layout.setContentsMargins(0, 0, 0, 0)
        editor_layout.setSpacing(0)

        # Code editor
        self.editor = CodeEditor()
        editor_layout.addWidget(self.editor)

        # Minimap
        self.minimap = Minimap(self.editor)
        editor_layout.addWidget(self.minimap)

        layout.addWidget(editor_container)

        # Connect find/replace signals
        self.find_widget.find_next.connect(self._on_find_next)
        self.find_widget.find_prev.connect(self._on_find_prev)
        self.find_widget.replace_one.connect(self._on_replace_one)
        self.find_widget.replace_all.connect(self._on_replace_all)
        self.find_widget.closed.connect(lambda: self.editor.setFocus())

        # Connect cursor position change
        self.editor.cursorPositionChanged.connect(self._on_cursor_changed)

        # Connect document modification tracking
        self.editor.document().modificationChanged.connect(self._on_modification_changed)

        # Setup debounced linting on text change
        self._lint_timer = QTimer()
        self._lint_timer.setSingleShot(True)
        self._lint_timer.timeout.connect(self._run_linter)
        self.editor.textChanged.connect(self._schedule_lint)

        # Install shortcut handler
        self.editor.installEventFilter(self)

    def _on_modification_changed(self, modified: bool):
        """Track document modification state."""
        self._is_modified = modified
        self.modification_changed.emit(modified)

    @property
    def is_modified(self) -> bool:
        """Return whether the document has unsaved changes."""
        return self._is_modified

    def _on_cursor_changed(self):
        """Emit cursor position when it changes."""
        cursor = self.editor.textCursor()
        line = cursor.blockNumber() + 1
        col = cursor.columnNumber() + 1
        self.cursor_position_changed.emit(line, col)

    def _schedule_lint(self):
        """Schedule linting with debounce (wait for user to stop typing)."""
        if self._lint_timer:
            self._lint_timer.stop()
            self._lint_timer.start(1500)  # 1.5 second debounce

    def _run_linter(self):
        """Run linter on current file."""
        if not self.current_file or not str(self.current_file).endswith('.py'):
            return

        # Cancel any running linter
        if self._linter_worker and self._linter_worker.isRunning():
            return  # Don't interrupt running linter

        # Save file first if modified (linter needs saved file)
        if self._is_modified:
            self.save_file()

        # Start linter worker
        self._linter_worker = LinterWorker(str(self.current_file))
        self._linter_worker.diagnostics_ready.connect(self._on_diagnostics_ready)
        self._linter_worker.start()

    def _on_diagnostics_ready(self, diagnostics: Dict[int, str]):
        """Handle linter diagnostics and update minimap."""
        self.minimap.set_diagnostics(diagnostics)

        # Also update line number area if it supports diagnostics
        if hasattr(self.editor, 'line_number_area'):
            # Could extend LineNumberArea to show diagnostics too
            pass

    def run_linter_now(self):
        """Manually trigger linting immediately."""
        if self._lint_timer:
            self._lint_timer.stop()
        self._run_linter()

    def eventFilter(self, obj, event):
        """Handle keyboard shortcuts."""
        if obj == self.editor and event.type() == event.Type.KeyPress:
            modifiers = event.modifiers()
            key = event.key()

            # Ctrl+F - Find
            if modifiers == Qt.KeyboardModifier.ControlModifier and key == Qt.Key.Key_F:
                self.show_find()
                return True

            # Ctrl+H - Replace
            if modifiers == Qt.KeyboardModifier.ControlModifier and key == Qt.Key.Key_H:
                self.show_replace()
                return True

            # F3 - Find Next
            if key == Qt.Key.Key_F3:
                if modifiers == Qt.KeyboardModifier.ShiftModifier:
                    self._on_find_prev(self.find_widget.find_input.text(),
                                       self.find_widget.case_btn.isChecked())
                else:
                    self._on_find_next(self.find_widget.find_input.text(),
                                       self.find_widget.case_btn.isChecked())
                return True

            # Escape - Close find widget
            if key == Qt.Key.Key_Escape and self.find_widget.isVisible():
                self.find_widget.hide()
                self.editor.setFocus()
                return True

        return super().eventFilter(obj, event)

    def show_find(self):
        """Show find widget."""
        self.find_widget.show_find()
        # Pre-populate with selected text if any
        cursor = self.editor.textCursor()
        if cursor.hasSelection():
            self.find_widget.find_input.setText(cursor.selectedText())

    def show_replace(self):
        """Show replace widget."""
        self.find_widget.show_replace()
        cursor = self.editor.textCursor()
        if cursor.hasSelection():
            self.find_widget.find_input.setText(cursor.selectedText())

    def _on_find_next(self, text: str, case_sensitive: bool):
        current, total = self.editor.find_next(text, case_sensitive)
        self.find_widget.update_match_count(current, total)

    def _on_find_prev(self, text: str, case_sensitive: bool):
        current, total = self.editor.find_prev(text, case_sensitive)
        self.find_widget.update_match_count(current, total)

    def _on_replace_one(self, find_text: str, replace_text: str, case_sensitive: bool):
        current, total = self.editor.replace_current(find_text, replace_text, case_sensitive)
        self.find_widget.update_match_count(current, total)

    def _on_replace_all(self, find_text: str, replace_text: str, case_sensitive: bool):
        count = self.editor.replace_all(find_text, replace_text, case_sensitive)
        self.find_widget.update_match_count(0, 0)
        if count > 0:
            self.find_widget.match_label.setText(f"{count} replaced")
            self.find_widget.match_label.setStyleSheet(f"color: {Theme.SUCCESS}; font-size: 10px;")

    # Proxy methods to underlying editor
    def load_file(self, path: Path):
        self.editor.load_file(path)
        self.current_file = path
        # Reset modification state after loading
        self.editor.document().setModified(False)
        self._is_modified = False
        # Run linter on file open
        QTimer.singleShot(500, self._run_linter)

    def save_file(self):
        self.editor.save_file()
        # Reset modification state after saving
        self.editor.document().setModified(False)
        self._is_modified = False
        # Run linter after save
        QTimer.singleShot(100, self._run_linter)

    def toPlainText(self):
        return self.editor.toPlainText()

    def setPlainText(self, text: str):
        self.editor.setPlainText(text)

    def toggle_blame(self):
        """Toggle git blame view."""
        self.editor.toggle_blame()

    def is_blame_visible(self) -> bool:
        """Check if git blame is visible."""
        return self.editor.is_blame_visible()

    def fold_all(self):
        """Fold all collapsible regions."""
        self.editor.folding_area.fold_all()

    def unfold_all(self):
        """Unfold all collapsed regions."""
        self.editor.folding_area.unfold_all()


# ============================================================================
# Editor Tabs
# ============================================================================

class EditorArea(QWidget):
    """Split editor area supporting horizontal/vertical splits."""

    file_opened = Signal(Path)
    cursor_position_changed = Signal(int, int)  # line, col

    def __init__(self, parent=None):
        super().__init__(parent)
        self._open_files: Dict[str, int] = {}
        self._original_names: Dict[int, str] = {}

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Main splitter for editor panes
        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        self.splitter.setHandleWidth(2)
        self.splitter.setStyleSheet(f"""
            QSplitter::handle {{
                background: {Theme.BORDER};
            }}
            QSplitter::handle:hover {{
                background: {Theme.ACCENT_BLUE};
            }}
        """)

        # Create first editor pane
        self._panes: List[EditorTabs] = []
        self._active_pane: Optional[EditorTabs] = None
        self._add_pane()

        layout.addWidget(self.splitter)

        # Context menu for splitting
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)

    def _add_pane(self, orientation: Qt.Orientation = None) -> 'EditorTabs':
        """Add a new editor pane."""
        pane = EditorTabs()
        pane.currentChanged.connect(lambda: self._on_pane_activated(pane))
        pane.tabCloseRequested.connect(lambda idx: self._on_tab_close_requested(pane, idx))

        # Connect cursor position signal
        if hasattr(pane, 'cursor_position_changed'):
            pane.cursor_position_changed.connect(self.cursor_position_changed.emit)

        self._panes.append(pane)
        self.splitter.addWidget(pane)

        if not self._active_pane:
            self._active_pane = pane

        return pane

    def _on_pane_activated(self, pane: 'EditorTabs'):
        """Handle pane activation."""
        self._active_pane = pane

    def _on_tab_close_requested(self, pane: 'EditorTabs', index: int):
        """Handle tab close in a pane."""
        pane._on_close_requested(index)

        # If pane is empty and we have more than one pane, remove it
        if pane.count() == 0 and len(self._panes) > 1:
            self._remove_pane(pane)

    def _remove_pane(self, pane: 'EditorTabs'):
        """Remove an empty pane."""
        if pane in self._panes:
            self._panes.remove(pane)
            pane.setParent(None)
            pane.deleteLater()

            if self._active_pane == pane and self._panes:
                self._active_pane = self._panes[0]

    def split_horizontal(self):
        """Split the editor horizontally (side by side)."""
        self.splitter.setOrientation(Qt.Orientation.Horizontal)
        self._add_pane()
        self._balance_splits()

    def split_vertical(self):
        """Split the editor vertically (top/bottom)."""
        self.splitter.setOrientation(Qt.Orientation.Vertical)
        self._add_pane()
        self._balance_splits()

    def _balance_splits(self):
        """Balance split sizes equally."""
        count = len(self._panes)
        if count > 0:
            size = self.splitter.width() if self.splitter.orientation() == Qt.Orientation.Horizontal else self.splitter.height()
            sizes = [size // count] * count
            self.splitter.setSizes(sizes)

    def close_split(self):
        """Close the current split pane."""
        if len(self._panes) > 1 and self._active_pane:
            self._remove_pane(self._active_pane)

    def _show_context_menu(self, pos):
        """Show context menu for split options."""
        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu {{
                background: {Theme.BG_SECONDARY};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                padding: 4px;
            }}
            QMenu::item {{
                padding: 6px 20px;
            }}
            QMenu::item:selected {{
                background: {Theme.BG_SELECTED};
            }}
        """)

        split_h = menu.addAction("Split Editor Right")
        split_h.triggered.connect(self.split_horizontal)

        split_v = menu.addAction("Split Editor Down")
        split_v.triggered.connect(self.split_vertical)

        if len(self._panes) > 1:
            menu.addSeparator()
            close_split = menu.addAction("Close Split")
            close_split.triggered.connect(self.close_split)

        menu.exec(self.mapToGlobal(pos))

    # Delegate methods to active pane
    def open_file(self, path: Path):
        """Open file in active pane."""
        if self._active_pane:
            self._active_pane.open_file(path)

    def show_welcome(self, welcome_widget):
        """Show welcome screen."""
        if self._active_pane:
            self._active_pane.show_welcome(welcome_widget)

    def save_current(self):
        """Save current file."""
        if self._active_pane:
            self._active_pane.save_current()

    def has_unsaved_changes(self) -> bool:
        """Check for unsaved changes across all panes."""
        return any(pane.has_unsaved_changes() for pane in self._panes)

    def clear(self):
        """Clear all panes."""
        for pane in self._panes:
            pane.clear()

    def count(self):
        """Total tab count across all panes."""
        return sum(pane.count() for pane in self._panes)

    def currentWidget(self):
        """Get current widget from active pane."""
        if self._active_pane:
            return self._active_pane.currentWidget()
        return None

    def currentIndex(self):
        """Get current index from active pane."""
        if self._active_pane:
            return self._active_pane.currentIndex()
        return -1

    def widget(self, index):
        """Get widget at index from active pane."""
        if self._active_pane:
            return self._active_pane.widget(index)
        return None

    def close_tab(self, index):
        """Close tab at index in active pane."""
        if self._active_pane:
            self._active_pane._close_tab(index)

    def addTab(self, widget, label):
        """Add tab to active pane."""
        if self._active_pane:
            return self._active_pane.addTab(widget, label)
        return -1

    def setCurrentIndex(self, index):
        """Set current index in active pane."""
        if self._active_pane:
            self._active_pane.setCurrentIndex(index)

    def setTabIcon(self, index, icon):
        """Set tab icon in active pane."""
        if self._active_pane:
            self._active_pane.setTabIcon(index, icon)

    def removeTab(self, index):
        """Remove tab at index from active pane."""
        if self._active_pane:
            self._active_pane.removeTab(index)


class EditorTabs(QTabWidget):
    """Tabbed editor container with modification indicators."""

    MODIFIED_DOT = "*"  # Asterisk for modified indicator (VS Code style)
    cursor_position_changed = Signal(int, int)  # line, col

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setTabsClosable(True)
        self.setMovable(True)
        self.setDocumentMode(True)

        self.setStyleSheet(f"""
            QTabWidget::pane {{ border: none; background: {Theme.BG_MAIN}; }}
            QTabBar {{ background: {Theme.BG_SECONDARY}; }}
            QTabBar::tab {{
                background: {Theme.BG_SECONDARY};
                color: {Theme.TEXT_SECONDARY};
                padding: 6px 24px 6px 14px;
                border: none;
                border-bottom: 1px solid transparent;
                font-size: 11px;
            }}
            QTabBar::tab:selected {{
                background: {Theme.BG_MAIN};
                color: {Theme.TEXT_PRIMARY};
                border-bottom: 1px solid {Theme.ACCENT_BLUE};
            }}
            QTabBar::tab:hover {{ background: {Theme.BG_TERTIARY}; }}
            QTabBar::close-button {{
                subcontrol-position: right;
                margin-right: 4px;
                width: 12px;
                height: 12px;
                border-radius: 2px;
                background: transparent;
            }}
            QTabBar::close-button:hover {{
                background: rgba(255, 255, 255, 0.1);
            }}
        """)

        self._open_files: Dict[str, int] = {}
        self._original_names: Dict[int, str] = {}  # index -> original filename
        self._close_icon = Icons.close(12, Theme.TEXT_MUTED)

        # Set custom close button icons for each tab
        self._setup_close_buttons()

        self.tabCloseRequested.connect(self._on_close_requested)

        # Enable middle-click to close tabs
        self.tabBar().installEventFilter(self)

    def eventFilter(self, obj, event):
        """Handle middle-click to close tabs and context menu."""
        if obj == self.tabBar():
            if event.type() == event.Type.MouseButtonPress:
                if event.button() == Qt.MouseButton.MiddleButton:
                    tab_index = self.tabBar().tabAt(event.pos())
                    if tab_index >= 0:
                        self._on_close_requested(tab_index)
                        return True
            elif event.type() == event.Type.ContextMenu:
                tab_index = self.tabBar().tabAt(event.pos())
                if tab_index >= 0:
                    self._show_tab_context_menu(tab_index, event.globalPos())
                    return True
        return super().eventFilter(obj, event)

    def _show_tab_context_menu(self, index: int, pos):
        """Show context menu for tab."""
        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu {{
                background: {Theme.BG_SECONDARY};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                padding: 4px;
            }}
            QMenu::item {{
                padding: 6px 24px 6px 12px;
            }}
            QMenu::item:selected {{
                background: {Theme.BG_SELECTED};
            }}
            QMenu::separator {{
                height: 1px;
                background: {Theme.BORDER};
                margin: 4px 8px;
            }}
        """)

        widget = self.widget(index)
        file_path = getattr(widget, 'current_file', None)
        file_name = self._original_names.get(index, 'File')

        # Close actions
        close_action = menu.addAction("Close")
        close_action.triggered.connect(lambda: self._on_close_requested(index))

        close_others_action = menu.addAction("Close Others")
        close_others_action.triggered.connect(lambda: self._close_others(index))
        close_others_action.setEnabled(self.count() > 1)

        close_right_action = menu.addAction("Close to the Right")
        close_right_action.triggered.connect(lambda: self._close_to_right(index))
        close_right_action.setEnabled(index < self.count() - 1)

        close_all_action = menu.addAction("Close All")
        close_all_action.triggered.connect(self._close_all)

        menu.addSeparator()

        # File actions
        if file_path:
            copy_path_action = menu.addAction("Copy Path")
            copy_path_action.triggered.connect(lambda: self._copy_path(file_path))

            copy_name_action = menu.addAction("Copy File Name")
            copy_name_action.triggered.connect(lambda: QApplication.clipboard().setText(file_name))

            menu.addSeparator()

            reveal_action = menu.addAction("Reveal in File Explorer")
            reveal_action.triggered.connect(lambda: self._reveal_in_explorer(file_path))

        menu.exec(pos)

    def _close_others(self, keep_index: int):
        """Close all tabs except the specified one."""
        for i in range(self.count() - 1, -1, -1):
            if i != keep_index:
                self._on_close_requested(i)

    def _close_to_right(self, index: int):
        """Close all tabs to the right of the specified one."""
        for i in range(self.count() - 1, index, -1):
            self._on_close_requested(i)

    def _close_all(self):
        """Close all tabs."""
        for i in range(self.count() - 1, -1, -1):
            self._on_close_requested(i)

    def _copy_path(self, path: Path):
        """Copy file path to clipboard."""
        QApplication.clipboard().setText(str(path))
        ToastNotification.success("Path copied to clipboard", self.window())

    def _reveal_in_explorer(self, path: Path):
        """Reveal file in system file explorer."""
        import subprocess
        if sys.platform == "darwin":
            subprocess.run(["open", "-R", str(path)])
        elif sys.platform == "win32":
            subprocess.run(["explorer", "/select,", str(path)])
        else:
            subprocess.run(["xdg-open", str(path.parent)])

    def _setup_close_buttons(self):
        """Setup close buttons for existing tabs."""
        for i in range(self.count()):
            self._set_tab_close_button(i)

    def addTab(self, widget, label):
        """Override addTab to set close button on new tabs."""
        index = super().addTab(widget, label)
        self._set_tab_close_button(index)
        return index

    def _set_tab_close_button(self, index: int):
        """Set the close button for a specific tab."""
        close_btn = QPushButton()
        close_btn.setFixedSize(16, 16)
        close_btn.setIcon(self._close_icon)
        close_btn.setIconSize(QSize(12, 12))
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                border: none;
                border-radius: 3px;
                padding: 2px;
            }}
            QPushButton:hover {{
                background: rgba(255, 255, 255, 0.1);
            }}
        """)
        close_btn.clicked.connect(self._on_close_button_clicked)
        self.tabBar().setTabButton(index, QTabBar.ButtonPosition.RightSide, close_btn)

    def _on_close_button_clicked(self):
        """Handle close button click - find the correct tab index."""
        btn = self.sender()
        for i in range(self.count()):
            if self.tabBar().tabButton(i, QTabBar.ButtonPosition.RightSide) == btn:
                self.tabCloseRequested.emit(i)
                break

    def _on_close_requested(self, index: int):
        """Handle tab close with unsaved changes warning."""
        widget = self.widget(index)
        if isinstance(widget, (EditorWithFindReplace, ExcelViewer)) and widget.is_modified:
            # Show confirmation dialog
            reply = QMessageBox.question(
                self, "Unsaved Changes",
                f"'{self._original_names.get(index, 'File')}' has unsaved changes.\n\nDo you want to save before closing?",
                QMessageBox.StandardButton.Save |
                QMessageBox.StandardButton.Discard |
                QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Save
            )
            if reply == QMessageBox.StandardButton.Save:
                widget.save_file()
                self._close_tab(index)
            elif reply == QMessageBox.StandardButton.Discard:
                self._close_tab(index)
            # Cancel: do nothing
        else:
            self._close_tab(index)

    def _close_tab(self, index: int):
        """Actually close the tab and clean up tracking."""
        # Find and remove from open files dict
        path_to_remove = None
        for path, idx in self._open_files.items():
            if idx == index:
                path_to_remove = path
                break
        if path_to_remove:
            del self._open_files[path_to_remove]

        # Remove from original names
        if index in self._original_names:
            del self._original_names[index]

        # Update indices for tabs after this one
        new_open_files = {}
        for path, idx in self._open_files.items():
            if idx > index:
                new_open_files[path] = idx - 1
            else:
                new_open_files[path] = idx
        self._open_files = new_open_files

        new_original_names = {}
        for idx, name in self._original_names.items():
            if idx > index:
                new_original_names[idx - 1] = name
            else:
                new_original_names[idx] = name
        self._original_names = new_original_names

        self.removeTab(index)

    def _on_modification_changed(self, modified: bool):
        """Update tab title when modification state changes."""
        editor = self.sender()
        if not editor:
            return

        # Find the tab index for this editor
        for i in range(self.count()):
            if self.widget(i) == editor:
                original_name = self._original_names.get(i, "")
                if modified:
                    # Add dot indicator
                    self.setTabText(i, f"{self.MODIFIED_DOT} {original_name}")
                    # Change tab text color to indicate unsaved
                    self.tabBar().setTabTextColor(i, QColor(Theme.WARNING))
                else:
                    # Remove dot indicator
                    self.setTabText(i, original_name)
                    # Reset color
                    self.tabBar().setTabTextColor(i, QColor(Theme.TEXT_SECONDARY))
                break

    def open_file(self, path: Path):
        path_str = str(path)

        if path_str in self._open_files:
            self.setCurrentIndex(self._open_files[path_str])
            return

        # Check file extension to determine which viewer to use
        suffix = path.suffix.lower()

        if suffix in ('.xlsx', '.xls', '.xlsm'):
            # Use Excel viewer for spreadsheet files
            viewer = ExcelViewer()
            viewer.load_file(path)
            viewer.modification_changed.connect(self._on_modification_changed)
            index = self.addTab(viewer, path.name)
        else:
            # Use EditorWithFindReplace for text files
            editor = EditorWithFindReplace()
            editor.load_file(path)
            editor.modification_changed.connect(self._on_modification_changed)
            index = self.addTab(editor, path.name)

        self._open_files[path_str] = index
        self._original_names[index] = path.name
        self.setCurrentIndex(index)

    def show_welcome(self, welcome_widget):
        """Show welcome screen in editor area."""
        self.clear()
        self._open_files.clear()
        self._original_names.clear()
        self.addTab(welcome_widget, "Welcome")

    def save_current(self):
        """Save the current editor's file."""
        current = self.currentWidget()
        if isinstance(current, EditorWithFindReplace):
            current.save_file()
        elif isinstance(current, ExcelViewer):
            current.save_file()

    def has_unsaved_changes(self) -> bool:
        """Check if any open file has unsaved changes."""
        for i in range(self.count()):
            widget = self.widget(i)
            if isinstance(widget, (EditorWithFindReplace, ExcelViewer)) and widget.is_modified:
                return True
        return False


# ============================================================================
# Terminal Panel
# ============================================================================

class AnsiColorParser:
    """Parse ANSI escape codes and convert to Qt formatting."""

    # ANSI color codes to Qt colors
    COLORS = {
        30: "#1e1e1e", 31: "#f14c4c", 32: "#4ec9b0", 33: "#dcdcaa",
        34: "#569cd6", 35: "#c586c0", 36: "#4fc1ff", 37: "#d4d4d4",
        90: "#808080", 91: "#f14c4c", 92: "#4ec9b0", 93: "#dcdcaa",
        94: "#569cd6", 95: "#c586c0", 96: "#4fc1ff", 97: "#ffffff",
    }

    @classmethod
    def parse(cls, text: str) -> List[tuple]:
        """Parse ANSI text into (text, color) tuples."""
        import re
        result = []
        current_color = None
        pattern = re.compile(r'\x1b\[([0-9;]*)m')

        last_end = 0
        for match in pattern.finditer(text):
            # Add text before this escape sequence
            if match.start() > last_end:
                result.append((text[last_end:match.start()], current_color))

            # Parse the escape code
            codes = match.group(1).split(';') if match.group(1) else ['0']
            for code in codes:
                code_int = int(code) if code else 0
                if code_int == 0:
                    current_color = None
                elif code_int in cls.COLORS:
                    current_color = cls.COLORS[code_int]

            last_end = match.end()

        # Add remaining text
        if last_end < len(text):
            result.append((text[last_end:], current_color))

        return result


class TerminalPanel(QWidget):
    """Professional integrated terminal with persistent shell session."""

    command_executed = Signal(str, int)  # command, exit_code
    MAX_BUFFER_LINES = 10000  # Prevent memory issues with large output

    def __init__(self, working_dir: str = None, parent=None):
        super().__init__(parent)
        self.working_dir = Path(working_dir) if working_dir else Path.home()
        self.history: List[str] = []
        self.history_index = -1
        self._shell_ready = False

        # Persistent shell process
        self.shell = QProcess(self)
        self.shell.setProcessChannelMode(QProcess.ProcessChannelMode.MergedChannels)
        self.shell.readyReadStandardOutput.connect(self._on_output)
        self.shell.finished.connect(self._on_shell_finished)
        self.shell.errorOccurred.connect(self._on_shell_error)

        self._setup_ui()
        self._start_shell()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Terminal header with tabs
        header = QWidget()
        header.setFixedHeight(32)
        header.setStyleSheet(f"background: {Theme.BG_SECONDARY}; border-top: 1px solid {Theme.BORDER};")
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(8, 0, 8, 0)
        header_layout.setSpacing(4)

        # Terminal tab with shell indicator
        self.terminal_btn = QPushButton("zsh")
        self.terminal_btn.setCheckable(True)
        self.terminal_btn.setChecked(True)
        self.terminal_btn.setStyleSheet(f"""
            QPushButton {{
                background: {Theme.BG_TERTIARY};
                color: {Theme.TEXT_PRIMARY};
                border: none;
                padding: 4px 12px;
                font-size: 11px;
                border-radius: 4px 4px 0 0;
            }}
            QPushButton:checked {{
                background: {Theme.BG_DARK};
                border-bottom: 2px solid {Theme.ACCENT_BLUE};
            }}
        """)
        header_layout.addWidget(self.terminal_btn)

        # Working directory label
        self.cwd_label = QLabel(str(self.working_dir.name or "/"))
        self.cwd_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 10px; padding: 0 8px;")
        header_layout.addWidget(self.cwd_label)

        header_layout.addStretch()

        # Kill process button
        self.kill_btn = IconButton(Icons.close, "Kill Process (Ctrl+C)")
        self.kill_btn.clicked.connect(self._kill_process)
        self.kill_btn.setEnabled(False)
        header_layout.addWidget(self.kill_btn)

        # Clear button
        clear_btn = IconButton(Icons.clear, "Clear (Ctrl+L)")
        clear_btn.clicked.connect(self._clear_terminal)
        header_layout.addWidget(clear_btn)

        # Restart shell button
        restart_btn = IconButton(Icons.refresh, "Restart Shell")
        restart_btn.clicked.connect(self._restart_shell)
        header_layout.addWidget(restart_btn)

        layout.addWidget(header)

        # Terminal output area - larger font, better styling
        self.output = QPlainTextEdit()
        self.output.setReadOnly(True)
        self.output.setFont(QFont("Menlo", 12) if sys.platform == "darwin" else QFont("Consolas", 11))
        self.output.setStyleSheet(f"""
            QPlainTextEdit {{
                background: #1a1a1a;
                color: #f0f0f0;
                border: none;
                padding: 8px;
                selection-background-color: #3a3a5a;
            }}
            QScrollBar:vertical {{
                background: #1a1a1a;
                width: 10px;
            }}
            QScrollBar::handle:vertical {{
                background: #3a3a3a;
                border-radius: 5px;
                min-height: 30px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: #4a4a4a;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0;
            }}
        """)
        # Allow clicking in output to focus input
        self.output.mousePressEvent = lambda e: self.input.setFocus()
        layout.addWidget(self.output)

        # Input line with better styling
        input_container = QWidget()
        input_container.setFixedHeight(36)
        input_container.setStyleSheet(f"background: #1a1a1a; border-top: 1px solid {Theme.BORDER};")
        input_layout = QHBoxLayout(input_container)
        input_layout.setContentsMargins(8, 4, 8, 4)
        input_layout.setSpacing(8)

        # Prompt shows current directory
        self.prompt = QLabel("$")
        self.prompt.setStyleSheet(f"""
            color: {Theme.ACCENT_GREEN};
            font-weight: bold;
            font-family: {'Menlo' if sys.platform == 'darwin' else 'Consolas'};
            font-size: 12px;
        """)
        input_layout.addWidget(self.prompt)

        self.input = QLineEdit()
        self.input.setStyleSheet(f"""
            QLineEdit {{
                background: transparent;
                border: none;
                color: #f0f0f0;
                font-family: {'Menlo' if sys.platform == 'darwin' else 'Consolas'};
                font-size: 12px;
                padding: 2px;
            }}
        """)
        self.input.setPlaceholderText("Enter command...")
        self.input.returnPressed.connect(self._send_command)
        self.input.installEventFilter(self)
        input_layout.addWidget(self.input)

        layout.addWidget(input_container)

    def _start_shell(self):
        """Start the interactive shell process."""
        import platform

        # Set environment for better shell experience
        env = QProcess.systemEnvironment()
        env.append("TERM=xterm-256color")
        env.append("CLICOLOR=1")
        env.append("CLICOLOR_FORCE=1")
        self.shell.setEnvironment(env)

        self.shell.setWorkingDirectory(str(self.working_dir))

        # Show welcome banner first (before shell output)
        self._show_welcome_banner()

        if platform.system() == "Windows":
            self.shell.start("cmd.exe", [])
        else:
            # Use zsh (macOS default) in non-login mode to suppress startup messages
            # -f flag skips reading startup files for cleaner output
            self.shell.start("/bin/zsh", ["-f"])

    def _restart_shell(self):
        """Restart the shell process."""
        if self.shell.state() != QProcess.ProcessState.NotRunning:
            self.shell.kill()
            self.shell.waitForFinished(1000)

        self._clear_terminal()
        self._start_shell()

    def _kill_process(self):
        """Send interrupt signal to running process."""
        if self.shell.state() == QProcess.ProcessState.Running:
            # Send Ctrl+C (SIGINT) to the shell
            self.shell.write(b'\x03')
            self._append_output("^C\n", Theme.ACCENT_RED)

    def eventFilter(self, obj, event):
        """Handle key events for history navigation and Ctrl+C."""
        if obj == self.input and event.type() == event.Type.KeyPress:
            key = event.key()
            modifiers = event.modifiers()

            if key == Qt.Key.Key_Up:
                self._history_prev()
                return True
            elif key == Qt.Key.Key_Down:
                self._history_next()
                return True
            elif key == Qt.Key.Key_C and modifiers == Qt.KeyboardModifier.ControlModifier:
                self._kill_process()
                return True
            elif key == Qt.Key.Key_L and modifiers == Qt.KeyboardModifier.ControlModifier:
                self._clear_terminal()
                return True
            elif key == Qt.Key.Key_Tab:
                # Basic tab completion - could be enhanced
                self._tab_complete()
                return True

        return super().eventFilter(obj, event)

    def _tab_complete(self):
        """Basic tab completion for file paths."""
        text = self.input.text()
        if not text:
            return

        # Get the last word (potential path)
        parts = text.rsplit(' ', 1)
        prefix = parts[-1] if parts else ""

        if not prefix:
            return

        # Try to complete the path
        try:
            if prefix.startswith('~'):
                base_path = Path.home()
                prefix = prefix[1:].lstrip('/')
            elif prefix.startswith('/'):
                base_path = Path('/')
                prefix = prefix[1:]
            else:
                base_path = self.working_dir

            # Find matching files/dirs
            search_dir = base_path / Path(prefix).parent if '/' in prefix else base_path
            search_prefix = Path(prefix).name if '/' in prefix else prefix

            if search_dir.exists():
                matches = [p.name for p in search_dir.iterdir()
                          if p.name.startswith(search_prefix)]

                if len(matches) == 1:
                    # Complete with the single match
                    completion = matches[0]
                    if (search_dir / completion).is_dir():
                        completion += "/"
                    new_text = text[:-len(search_prefix)] + completion if search_prefix else text + completion
                    self.input.setText(new_text)
                elif len(matches) > 1:
                    # Show all matches
                    self._append_output("\n" + "  ".join(sorted(matches)) + "\n", Theme.TEXT_MUTED)

        except Exception:
            pass

    def _history_prev(self):
        """Navigate to previous command in history."""
        if self.history and self.history_index > 0:
            self.history_index -= 1
            self.input.setText(self.history[self.history_index])
            self.input.setCursorPosition(len(self.input.text()))

    def _history_next(self):
        """Navigate to next command in history."""
        if self.history_index < len(self.history) - 1:
            self.history_index += 1
            self.input.setText(self.history[self.history_index])
        else:
            self.history_index = len(self.history)
            self.input.clear()

    def _send_command(self):
        """Send command to the shell."""
        command = self.input.text()
        self.input.clear()

        if not command.strip():
            # Just send newline for empty command
            self.shell.write(b'\n')
            return

        # Add to history
        if command.strip():
            self.history.append(command)
            self.history_index = len(self.history)

        # Handle local clear command
        if command.strip() == "clear":
            self._clear_terminal()
            return

        # Send command to shell (with newline)
        self.shell.write((command + '\n').encode('utf-8'))

        # Update working directory if cd command
        if command.strip().startswith("cd "):
            path = command.strip()[3:].strip()
            QTimer.singleShot(100, lambda: self._update_cwd_after_cd(path))

    def _update_cwd_after_cd(self, path: str):
        """Update working directory after cd command."""
        try:
            if path == "~" or path == "":
                new_path = Path.home()
            elif path.startswith("~"):
                new_path = Path.home() / path[2:]
            elif path.startswith("/"):
                new_path = Path(path)
            else:
                new_path = (self.working_dir / path).resolve()

            if new_path.is_dir():
                self.working_dir = new_path
                self.cwd_label.setText(new_path.name or "/")
                self.shell.setWorkingDirectory(str(new_path))
        except Exception:
            pass

    def _on_output(self):
        """Handle output from shell."""
        data = self.shell.readAllStandardOutput()
        text = bytes(data).decode('utf-8', errors='replace')

        if text:
            # Parse ANSI colors and append
            self._append_ansi_text(text)

    def _append_ansi_text(self, text: str):
        """Append text with ANSI color support."""
        # Strip some control sequences that don't render well
        import re
        # Remove cursor movement, clear screen, etc.
        text = re.sub(r'\x1b\[\?[0-9;]*[a-zA-Z]', '', text)
        text = re.sub(r'\x1b\[[0-9;]*[ABCDJKH]', '', text)
        text = re.sub(r'\x1b\]0;[^\x07]*\x07', '', text)  # Window title
        text = re.sub(r'\x1b\[[\?]?[0-9;]*[hlm]', '', text)  # Mode changes

        # Parse remaining ANSI colors
        segments = AnsiColorParser.parse(text)

        cursor = self.output.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)

        for segment_text, color in segments:
            if segment_text:
                fmt = QTextCharFormat()
                if color:
                    fmt.setForeground(QColor(color))
                else:
                    fmt.setForeground(QColor("#f0f0f0"))
                cursor.insertText(segment_text, fmt)

        self.output.setTextCursor(cursor)
        self.output.ensureCursorVisible()

        # Limit buffer size to prevent memory issues
        self._prune_buffer()

    def _prune_buffer(self):
        """Remove oldest lines if buffer exceeds maximum."""
        doc = self.output.document()
        line_count = doc.blockCount()
        if line_count > self.MAX_BUFFER_LINES:
            # Remove oldest lines
            cursor = self.output.textCursor()
            cursor.movePosition(QTextCursor.MoveOperation.Start)
            lines_to_remove = line_count - self.MAX_BUFFER_LINES
            for _ in range(lines_to_remove):
                cursor.movePosition(QTextCursor.MoveOperation.Down, QTextCursor.MoveMode.KeepAnchor)
            cursor.removeSelectedText()
            cursor.deleteChar()  # Remove trailing newline

    def _on_shell_finished(self, exit_code: int, exit_status):
        """Handle shell process exit."""
        self._append_output(f"\n[Process exited with code {exit_code}]\n", Theme.WARNING)
        self.kill_btn.setEnabled(False)
        # Restart shell automatically
        QTimer.singleShot(500, self._start_shell)

    def _on_shell_error(self, error):
        """Handle shell process errors."""
        error_messages = {
            QProcess.ProcessError.FailedToStart: "Failed to start shell",
            QProcess.ProcessError.Crashed: "Shell crashed",
            QProcess.ProcessError.Timedout: "Shell timed out",
            QProcess.ProcessError.WriteError: "Write error",
            QProcess.ProcessError.ReadError: "Read error",
            QProcess.ProcessError.UnknownError: "Unknown error",
        }
        msg = error_messages.get(error, "Unknown error")
        self._append_output(f"\n[{msg}]\n", Theme.ACCENT_RED)

    def _append_output(self, text: str, color: str):
        """Append plain text to output with specified color."""
        cursor = self.output.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)

        fmt = QTextCharFormat()
        fmt.setForeground(QColor(color))
        cursor.insertText(text, fmt)

        self.output.setTextCursor(cursor)
        self.output.ensureCursorVisible()

    def _clear_terminal(self):
        """Clear terminal output and show welcome banner."""
        self.output.clear()
        self._show_welcome_banner()

    def _show_welcome_banner(self):
        """Display a nice welcome banner."""
        banner = f'''<span style="color: {Theme.ACCENT_BLUE};">
  
                                           
     <span style="color: {Theme.ACCENT_GREEN};">Circuit IDE Terminal</span>                 
                                           
     <span style="color: {Theme.TEXT_MUTED};">Type commands below</span>                  
     <span style="color: {Theme.TEXT_MUTED};">Ctrl+L to clear  Ctrl+C to cancel</span>    
                                           
  
</span>'''
        self.output.appendHtml(f'<pre style="font-family: Menlo, Consolas, monospace; line-height: 1.4;">{banner}</pre>')

    def set_working_dir(self, path: str):
        """Set the working directory and cd the shell to it."""
        self.working_dir = Path(path)
        self.cwd_label.setText(self.working_dir.name or "/")

        # CD the shell to the new directory
        if self.shell.state() == QProcess.ProcessState.Running:
            self.shell.write(f'cd "{path}"\n'.encode('utf-8'))
            self.shell.setWorkingDirectory(path)

    def cleanup(self):
        """Clean up resources before destruction."""
        if self.shell.state() != QProcess.ProcessState.NotRunning:
            self.shell.kill()
            self.shell.waitForFinished(1000)

    def closeEvent(self, event):
        """Handle widget close - cleanup resources."""
        self.cleanup()
        super().closeEvent(event)


# ============================================================================
# Chat Panel
# ============================================================================

class TokenTracker(QFrame):
    """Compact token usage display."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"QFrame {{ background: {Theme.BG_TERTIARY}; border-radius: 3px; }}")

        layout = QHBoxLayout(self)
        layout.setContentsMargins(8, 4, 8, 4)
        layout.setSpacing(12)

        self.token_label = QLabel("0 tokens")
        self.token_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 10px;")
        layout.addWidget(self.token_label)

        self.cost_label = QLabel("$0.00")
        self.cost_label.setStyleSheet(f"color: {Theme.SUCCESS}; font-size: 10px;")
        layout.addWidget(self.cost_label)

        layout.addStretch()

    def update(self, tokens: int, cost: float):
        self.token_label.setText(f"{tokens:,} tokens")
        self.cost_label.setText(f"${cost:.4f}")


class OutputEntry(QFrame):
    """Claude Code-style output entry - clean, terminal-like with provider-specific colors."""

    # Class variable to track current provider
    _current_provider = "circuit"  # Default to circuit

    @classmethod
    def set_provider(cls, provider: str):
        """Set the current provider for message coloring."""
        cls._current_provider = provider.lower()

    @classmethod
    def get_provider(cls) -> str:
        return cls._current_provider

    def __init__(self, role: str, content: str, provider: str = None, parent=None):
        super().__init__(parent)
        self.role = role
        is_user = role == "user"

        # Use passed provider or class default
        self._provider = provider or self._current_provider

        # Determine color based on provider
        if self._provider == "claude":
            provider_color = Theme.CLAUDE_COLOR
        else:
            provider_color = Theme.CIRCUIT_COLOR

        self.setStyleSheet(f"""
            QFrame {{
                background: transparent;
                margin: 0;
                padding: 0;
            }}
        """)

        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(16, 6, 24, 6)  # More right margin
        main_layout.setSpacing(10)
        main_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        # Marker - use provider color for both user and assistant
        marker = QLabel()
        if is_user:
            marker.setPixmap(Icons.send(12, provider_color).pixmap(12, 12))
        else:
            marker.setPixmap(Icons.robot(12, provider_color).pixmap(12, 12))
        marker.setFixedSize(16, 16)
        marker.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignHCenter)
        main_layout.addWidget(marker)

        # Content - use provider color
        self.content_label = QLabel(content)
        self.content_label.setWordWrap(True)
        self.content_label.setTextInteractionFlags(
            Qt.TextInteractionFlag.TextSelectableByMouse |
            Qt.TextInteractionFlag.LinksAccessibleByMouse
        )
        self.content_label.setStyleSheet(f"""
            QLabel {{
                color: {provider_color};
                font-size: 12px;
                line-height: 1.5;
                padding-right: 8px;
            }}
        """)
        self.content_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.content_label.setMinimumWidth(100)
        main_layout.addWidget(self.content_label, 1)

    def update_content(self, content: str):
        self.content_label.setText(content)


# Keep ChatMessage as alias for compatibility
ChatMessage = OutputEntry


class AnimatedSpinner(QLabel):
    """Animated spinner for loading states."""

    FRAMES = ["|", "/", "-", "\\"]  # ASCII spinner frames

    def __init__(self, parent=None):
        super().__init__(parent)
        self._frame = 0
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._update_frame)
        self.setStyleSheet(f"color: {Theme.WARNING}; font-size: 12px; font-family: monospace;")
        self.setText(self.FRAMES[0])

    def start(self):
        """Start the animation."""
        self._timer.start(150)

    def stop(self):
        """Stop the animation."""
        self._timer.stop()

    def _update_frame(self):
        """Update to next frame."""
        self._frame = (self._frame + 1) % len(self.FRAMES)
        self.setText(self.FRAMES[self._frame])


class ToolInvocation(QFrame):
    """Claude Code-style tool invocation display.

    Shows tool calls with expandable results section.
    """

    def __init__(self, name: str, params: dict = None, status: str = "running", parent=None):
        super().__init__(parent)
        self.status = status
        self._expanded = False
        self._result_text = ""
        self.params = params or {}

        self.setStyleSheet(f"""
            QFrame {{
                background: transparent;
                margin: 0;
                padding: 0;
            }}
        """)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 4, 24, 4)  # Match OutputEntry margins
        main_layout.setSpacing(0)

        # Tool call row
        call_row = QHBoxLayout()
        call_row.setSpacing(8)
        call_row.setAlignment(Qt.AlignmentFlag.AlignTop)

        # Status marker/spinner
        self.status_widget = QWidget()
        self.status_widget.setFixedWidth(16)
        status_layout = QHBoxLayout(self.status_widget)
        status_layout.setContentsMargins(0, 0, 0, 0)

        if status == "running":
            self.spinner = AnimatedSpinner()
            self.spinner.start()
            status_layout.addWidget(self.spinner)
        else:
            self._create_status_icon(status, status_layout)

        call_row.addWidget(self.status_widget)

        # Tool name and parameters - function call style
        call_text = self._format_call(name, self.params)
        self.call_label = QLabel(call_text)
        self.call_label.setWordWrap(True)
        self.call_label.setStyleSheet(f"""
            QLabel {{
                color: {Theme.ACCENT_CYAN};
                font-size: 12px;
                font-family: 'Consolas', 'Monaco', monospace;
                padding-right: 8px;
            }}
        """)
        self.call_label.setMinimumWidth(100)
        call_row.addWidget(self.call_label, 1)

        main_layout.addLayout(call_row)

        # Expandable results section (initially hidden)
        self.result_container = QFrame()
        self.result_container.setStyleSheet(f"""
            QFrame {{
                background: {Theme.BG_TERTIARY};
                border-radius: 4px;
                margin-left: 26px;
                margin-top: 4px;
                margin-right: 8px;
            }}
        """)
        result_layout = QVBoxLayout(self.result_container)
        result_layout.setContentsMargins(8, 6, 8, 6)

        self.result_label = QLabel("")
        self.result_label.setWordWrap(True)
        self.result_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        self.result_label.setStyleSheet(f"""
            QLabel {{
                color: {Theme.TEXT_SECONDARY};
                font-size: 11px;
                font-family: 'Consolas', 'Monaco', monospace;
            }}
        """)
        result_layout.addWidget(self.result_label)

        self.result_container.hide()
        main_layout.addWidget(self.result_container)

        # Make the call row clickable for expand/collapse
        self.call_label.setCursor(Qt.CursorShape.PointingHandCursor)
        self.call_label.mousePressEvent = self._toggle_expand

    def _format_call(self, name: str, params: dict) -> str:
        """Format tool call like: Read(file: "path/to/file.py")"""
        if not params:
            return f"{name}()"

        # Format parameters
        param_strs = []
        for key, value in params.items():
            if isinstance(value, str):
                # Truncate long strings
                display_val = value if len(value) < 50 else value[:47] + "..."
                param_strs.append(f'{key}: "{display_val}"')
            else:
                param_strs.append(f"{key}: {value}")

        return f"{name}({', '.join(param_strs)})"

    def _create_status_icon(self, status: str, layout: QHBoxLayout):
        """Create appropriate status icon using VS Code icons."""
        icon_map = {
            "done": (Icons.check, Theme.SUCCESS),
            "success": (Icons.check, Theme.SUCCESS),
            "error": (Icons.error, Theme.ERROR),
            "pending": (Icons.info, Theme.TEXT_MUTED),
        }

        icon_method, color = icon_map.get(status, (Icons.info, Theme.TEXT_MUTED))

        label = QLabel()
        label.setPixmap(icon_method(12, color).pixmap(12, 12))
        label.setFixedSize(12, 12)
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(label)

    def _toggle_expand(self, event):
        """Toggle result expansion."""
        if self._result_text:
            self._expanded = not self._expanded
            self.result_container.setVisible(self._expanded)

    def set_status(self, status: str):
        """Update the status of the tool call."""
        self.status = status

        # Stop spinner if exists
        if hasattr(self, 'spinner') and status != "running":
            self.spinner.stop()
            self.spinner.deleteLater()

            # Replace with status icon
            layout = self.status_widget.layout()
            while layout.count():
                item = layout.takeAt(0)
                if w := item.widget():
                    w.deleteLater()
            self._create_status_icon(status, layout)

    def set_result(self, result: str, lines_changed: int = None):
        """Set the result text and optionally show lines changed."""
        self._result_text = result

        if lines_changed is not None:
            summary = f"({lines_changed:+d} lines)" if lines_changed != 0 else "(no changes)"
            self.call_label.setText(f"{self.call_label.text()} {summary}")

        # Truncate very long results
        display = result[:2000] + "..." if len(result) > 2000 else result
        self.result_label.setText(display)


# Keep ToolCallWidget as alias for compatibility
ToolCallWidget = ToolInvocation


class ChatPanel(QWidget):
    """Claude Code-style output panel - terminal-like interface."""

    message_sent = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._streaming_message = None
        self._streaming_content = ""
        self._tool_widgets = {}  # Track tool widgets by ID

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Minimal header bar
        header = QWidget()
        header.setFixedHeight(32)
        header.setStyleSheet(f"background: {Theme.BG_SECONDARY}; border-bottom: 1px solid {Theme.BORDER};")
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(12, 0, 12, 0)

        self.status_dot = StatusDot(Theme.TEXT_MUTED, 6)
        header_layout.addWidget(self.status_dot)

        self.status_label = QLabel("Disconnected")
        self.status_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
        header_layout.addWidget(self.status_label)

        header_layout.addStretch()

        # Token tracker inline in header
        self.token_tracker = TokenTracker()
        header_layout.addWidget(self.token_tracker)

        layout.addWidget(header)

        # Output area - terminal style
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll.setStyleSheet(f"""
            QScrollArea {{ background: {Theme.BG_DARK}; border: none; }}
            QScrollBar:vertical {{
                background: {Theme.BG_DARK}; width: 8px;
            }}
            QScrollBar::handle:vertical {{
                background: {Theme.BG_TERTIARY}; border-radius: 4px; min-height: 30px;
            }}
            QScrollBar::handle:vertical:hover {{ background: {Theme.BG_HOVER}; }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
        """)

        self.messages_container = QWidget()
        self.messages_container.setStyleSheet(f"background: {Theme.BG_DARK};")
        self.messages_layout = QVBoxLayout(self.messages_container)
        self.messages_layout.setContentsMargins(0, 8, 12, 8)  # Right margin for scrollbar
        self.messages_layout.setSpacing(12)  # Gap between user query and agent response
        self.messages_layout.addStretch()

        self.scroll.setWidget(self.messages_container)
        layout.addWidget(self.scroll)

        # Input area - clean prompt style
        input_container = QWidget()
        input_container.setStyleSheet(f"background: {Theme.BG_SECONDARY}; border-top: 1px solid {Theme.BORDER};")
        input_layout = QHBoxLayout(input_container)
        input_layout.setContentsMargins(12, 10, 12, 10)
        input_layout.setSpacing(10)

        # Prompt marker
        prompt = QLabel(">")
        prompt.setStyleSheet(f"color: {Theme.CIRCUIT_COLOR}; font-size: 14px; font-weight: bold;")
        prompt.setFixedWidth(16)
        input_layout.addWidget(prompt)

        self.input_field = CompactLineEdit("")
        self.input_field.setPlaceholderText("Ask a question or describe a task...")
        self.input_field.returnPressed.connect(self._send_message)
        self.input_field.setStyleSheet(f"""
            QLineEdit {{
                background: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 8px 12px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 13px;
            }}
            QLineEdit:focus {{ border-color: {Theme.CIRCUIT_COLOR}; }}
        """)
        input_layout.addWidget(self.input_field)

        self.send_btn = CompactButton("Send")
        self.send_btn.setFixedSize(60, 32)
        self.send_btn.clicked.connect(self._send_message)
        self.send_btn.setStyleSheet(f"""
            QPushButton {{
                background: {Theme.CIRCUIT_COLOR};
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: 500;
            }}
            QPushButton:hover {{ background: #5AA3D4; }}
            QPushButton:pressed {{ background: #4A93C4; }}
        """)
        input_layout.addWidget(self.send_btn)

        layout.addWidget(input_container)

    def set_connected(self, connected: bool):
        if connected:
            self.status_label.setText("Connected")
            self.status_label.setStyleSheet(f"color: {Theme.CIRCUIT_COLOR}; font-size: 11px; font-weight: bold;")
            self.status_dot.set_color(Theme.CIRCUIT_COLOR)
        else:
            self.status_label.setText("Disconnected")
            self.status_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 10px;")
            self.status_dot.set_color(Theme.TEXT_MUTED)

    def update_tokens(self, tokens: int, cost: float):
        self.token_tracker.update(tokens, cost)

    def _send_message(self):
        text = self.input_field.text().strip()
        if text:
            self.add_message("user", text)
            self.input_field.clear()
            self.message_sent.emit(text)

    def add_message(self, role: str, content: str):
        msg = ChatMessage(role, content)
        self.messages_layout.insertWidget(self.messages_layout.count() - 1, msg)
        QTimer.singleShot(50, self._scroll_bottom)

        if role == "assistant":
            self._streaming_message = msg
            self._streaming_content = content

    def start_assistant_message(self):
        self._streaming_content = ""
        msg = ChatMessage("assistant", "")
        self._streaming_message = msg
        self.messages_layout.insertWidget(self.messages_layout.count() - 1, msg)

    def add_streaming_content(self, chunk: str):
        if self._streaming_message:
            self._streaming_content += chunk
            self._streaming_message.update_content(self._streaming_content)
            QTimer.singleShot(10, self._scroll_bottom)

    def finish_streaming(self):
        self._streaming_message = None
        self._streaming_content = ""

    def add_tool_call(self, name: str, params: dict = None, status: str = "running", tool_id: str = None):
        """Add a tool invocation to the output.

        Args:
            name: Tool name (e.g., "Read", "Search", "Update")
            params: Tool parameters dict (e.g., {"file": "main.py"})
            status: One of "running", "done", "error"
            tool_id: Optional ID to track this tool for later updates
        """
        # Handle legacy string params
        if isinstance(params, str):
            params = {"detail": params} if params else None

        tool = ToolInvocation(name, params, status)
        self.messages_layout.insertWidget(self.messages_layout.count() - 1, tool)

        if tool_id:
            self._tool_widgets[tool_id] = tool

        QTimer.singleShot(50, self._scroll_bottom)
        return tool

    def update_tool_status(self, tool_id: str, status: str, result: str = None, lines_changed: int = None):
        """Update an existing tool's status and optionally set result."""
        if tool_id in self._tool_widgets:
            tool = self._tool_widgets[tool_id]
            tool.set_status(status)
            if result:
                tool.set_result(result, lines_changed)

    def _scroll_bottom(self):
        self.scroll.verticalScrollBar().setValue(self.scroll.verticalScrollBar().maximum())

    def clear(self):
        while self.messages_layout.count() > 1:
            item = self.messages_layout.takeAt(0)
            if w := item.widget():
                w.deleteLater()
        self._streaming_message = None
        self._streaming_content = ""
        self._tool_widgets.clear()

    def set_enabled(self, enabled: bool):
        self.input_field.setEnabled(enabled)
        self.send_btn.setEnabled(enabled)


# ============================================================================
# Claude Code Panel - Headless Mode with Session Continuity
# ============================================================================

class ClaudeCodeWorker(QThread):
    """Worker thread for Claude Code CLI subprocess using headless mode."""

    # Signals for streaming output
    text_output = Signal(str)       # assistant text content
    tool_started = Signal(dict)     # tool_use_begin events
    tool_input = Signal(dict)       # tool input with full details (for diffs)
    tool_completed = Signal(dict)   # tool_result events
    usage_updated = Signal(int, int)  # input_tokens, output_tokens
    finished = Signal(int, str, float)  # exit_code, session_id, cost_usd
    error = Signal(str)

    def __init__(self, message: str, working_dir: str = None, continue_session: bool = True, parent=None):
        super().__init__(parent)
        self._message = message
        self._working_dir = working_dir
        self._continue = continue_session
        self._process = None
        self._session_id = ""
        self._cancelled = False
        self._total_cost = 0.0
        # Tool tracking
        self._current_tool_id = None
        self._current_tool_name = None
        self._current_tool_input = ""

    def run(self):
        """Execute Claude Code in headless mode and stream results."""
        try:
            # Build command with streaming JSON output
            # Note: stream-json requires --verbose flag
            # Pre-approve common tools so Claude can work without interactive prompts
            allowed_tools = "Read,Edit,Write,Bash,Glob,Grep,Task,TodoWrite,WebFetch,WebSearch,NotebookEdit"
            cmd = [
                "claude", "-p", self._message,
                "--output-format", "stream-json",
                "--verbose",
                "--allowedTools", allowed_tools
            ]
            if self._continue:
                cmd.append("--continue")

            self._process = subprocess.Popen(
                cmd,
                cwd=self._working_dir,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1
            )

            # Parse streaming JSON line by line
            for line in iter(self._process.stdout.readline, ''):
                if self._cancelled:
                    break
                line = line.strip()
                if line:
                    try:
                        event = json.loads(line)
                        self._handle_event(event)
                    except json.JSONDecodeError:
                        # Non-JSON output (rare), emit as text
                        self.text_output.emit(line + "\n")

            exit_code = self._process.wait()
            if not self._cancelled:
                self.finished.emit(exit_code, self._session_id, self._total_cost)

        except FileNotFoundError:
            self.error.emit("Claude Code CLI not found. Please install it: npm install -g @anthropic-ai/claude-code")
        except Exception as e:
            self.error.emit(f"Error: {str(e)}")

    def _handle_event(self, event: dict):
        """Route streaming JSON event to appropriate signal."""
        event_type = event.get("type", "")

        if event_type == "assistant":
            # Assistant message with content blocks
            message = event.get("message", {})
            content = message.get("content", [])
            for block in content:
                if block.get("type") == "text":
                    self.text_output.emit(block.get("text", ""))
            # Emit usage info
            usage = message.get("usage", {})
            if usage:
                input_tokens = usage.get("input_tokens", 0) + usage.get("cache_read_input_tokens", 0)
                output_tokens = usage.get("output_tokens", 0)
                self.usage_updated.emit(input_tokens, output_tokens)
        elif event_type == "content_block_start":
            # Content block starting (tool or text)
            content_block = event.get("content_block", {})
            if content_block.get("type") == "tool_use":
                self._current_tool_id = content_block.get("id", "")
                self._current_tool_name = content_block.get("name", "Unknown")
                self._current_tool_input = ""
                self.tool_started.emit({
                    "tool_name": self._current_tool_name,
                    "tool_id": self._current_tool_id
                })
        elif event_type == "content_block_delta":
            # Streaming delta (text or tool input)
            delta = event.get("delta", {})
            if delta.get("type") == "text_delta":
                self.text_output.emit(delta.get("text", ""))
            elif delta.get("type") == "input_json_delta":
                # Accumulate tool input JSON
                self._current_tool_input += delta.get("partial_json", "")
        elif event_type == "content_block_stop":
            # Content block finished - if it was a tool, emit the full input
            if self._current_tool_id and self._current_tool_name:
                try:
                    input_data = json.loads(self._current_tool_input) if self._current_tool_input else {}
                except json.JSONDecodeError:
                    input_data = {}
                self.tool_input.emit({
                    "tool_name": self._current_tool_name,
                    "tool_id": self._current_tool_id,
                    "input": input_data
                })
                self._current_tool_id = None
                self._current_tool_name = None
                self._current_tool_input = ""
        elif event_type == "result":
            # Final result with session_id and cost
            self._session_id = event.get("session_id", "")
            self._total_cost = event.get("total_cost_usd", 0.0)
        elif event_type == "error":
            # Error event
            error_msg = event.get("error", {})
            self.error.emit(error_msg.get("message", "Unknown error"))

    def cancel(self):
        """Cancel the running process."""
        self._cancelled = True
        if self._process:
            try:
                self._process.terminate()
            except Exception:
                pass


class ClaudeCodePanel(QWidget):
    """Claude Code panel using headless mode with session continuity."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._worker: Optional[ClaudeCodeWorker] = None
        self._working_dir: Optional[str] = None
        self._session_id: Optional[str] = None
        self._is_first_message = True

        # Status tracking
        self._start_time: Optional[float] = None
        self._input_tokens = 0
        self._output_tokens = 0
        self._spinner_frames = ["|", "/", "-", "\\"]  # ASCII spinner frames
        self._spinner_index = 0

        # Text accumulation for markdown formatting
        self._text_buffer = ""
        self._in_code_block = False
        self._code_block_lang = ""
        self._code_blocks: List[str] = []  # Store code block contents for copy
        self._current_code_block = ""  # Current code block being accumulated

        # Timer for status updates
        self._status_timer = QTimer(self)
        self._status_timer.timeout.connect(self._update_status_display)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Header
        header = QWidget()
        header.setFixedHeight(32)
        header.setStyleSheet(f"background: {Theme.BG_SECONDARY}; border-bottom: 1px solid {Theme.BORDER};")
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(12, 0, 12, 0)

        self.status_dot = StatusDot(Theme.TEXT_MUTED, 6)
        header_layout.addWidget(self.status_dot)

        self.status_label = QLabel("Claude Code")
        self.status_label.setStyleSheet(f"color: {Theme.CLAUDE_COLOR}; font-size: 11px; font-weight: bold;")
        header_layout.addWidget(self.status_label)

        header_layout.addStretch()

        # Restart button
        self.restart_btn = IconButton(Icons.refresh, "Restart Claude")
        self.restart_btn.clicked.connect(self._restart_claude)
        header_layout.addWidget(self.restart_btn)

        layout.addWidget(header)

        # Terminal output area
        self.output = QTextEdit()
        self.output.setReadOnly(True)
        self.output.setFont(QFont("Menlo", 12))
        self.output.setStyleSheet(f"""
            QTextEdit {{
                background: {Theme.BG_DARK};
                color: {Theme.TEXT_PRIMARY};
                border: none;
                padding: 12px;
                selection-background-color: {Theme.BG_SELECTED};
            }}
        """)
        self.output.setAcceptRichText(True)
        self.output.setMouseTracking(True)
        self.output.mousePressEvent = self._on_output_click
        layout.addWidget(self.output)

        # Status bar (visible while processing)
        self.status_bar = QWidget()
        self.status_bar.setFixedHeight(28)
        self.status_bar.setStyleSheet(f"background: {Theme.BG_TERTIARY}; border-top: 1px solid {Theme.BORDER};")
        status_layout = QHBoxLayout(self.status_bar)
        status_layout.setContentsMargins(12, 0, 12, 0)
        status_layout.setSpacing(8)

        self.spinner_label = QLabel("|")
        self.spinner_label.setStyleSheet(f"color: {Theme.CLAUDE_COLOR}; font-size: 14px; font-family: monospace;")
        self.spinner_label.setFixedWidth(16)
        status_layout.addWidget(self.spinner_label)

        self.activity_label = QLabel("Thinking...")
        self.activity_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
        status_layout.addWidget(self.activity_label)

        status_layout.addStretch()

        self.tokens_label = QLabel("")
        self.tokens_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
        status_layout.addWidget(self.tokens_label)

        self.time_label = QLabel("")
        self.time_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
        status_layout.addWidget(self.time_label)

        self.status_bar.hide()  # Hidden by default
        layout.addWidget(self.status_bar)

        # Input area
        input_container = QWidget()
        input_container.setStyleSheet(f"background: {Theme.BG_SECONDARY}; border-top: 1px solid {Theme.BORDER};")
        input_layout = QHBoxLayout(input_container)
        input_layout.setContentsMargins(12, 10, 12, 10)
        input_layout.setSpacing(10)

        # Prompt marker
        prompt = QLabel(">")
        prompt.setStyleSheet(f"color: {Theme.CLAUDE_COLOR}; font-size: 14px; font-weight: bold;")
        prompt.setFixedWidth(16)
        input_layout.addWidget(prompt)

        self.input_field = CompactLineEdit("")
        self.input_field.setPlaceholderText("Message Claude Code...")
        self.input_field.returnPressed.connect(self._send_input)
        self.input_field.setStyleSheet(f"""
            QLineEdit {{
                background: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 8px 12px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 13px;
                font-family: Menlo, monospace;
            }}
            QLineEdit:focus {{ border-color: {Theme.CLAUDE_COLOR}; }}
        """)
        input_layout.addWidget(self.input_field)

        self.send_btn = CompactButton("Send")
        self.send_btn.setFixedSize(60, 32)
        self.send_btn.clicked.connect(self._send_input)
        self.send_btn.setStyleSheet(f"""
            QPushButton {{
                background: {Theme.CLAUDE_COLOR};
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: 500;
            }}
            QPushButton:hover {{ background: #D4956B; }}
            QPushButton:pressed {{ background: #C08358; }}
        """)
        input_layout.addWidget(self.send_btn)

        layout.addWidget(input_container)

    def set_working_dir(self, path: str):
        """Set the working directory for Claude Code."""
        self._working_dir = path

    def start_claude(self):
        """Initialize Claude Code panel (ready for messages)."""
        self.output.clear()
        self._code_blocks = []  # Reset code blocks for copy
        self._is_first_message = True
        self.status_label.setText("Claude Code - Ready")
        self.status_dot.set_color(Theme.SUCCESS)
        self._append_output("Claude Code ready. Type a message to begin.\n", Theme.TEXT_MUTED)

    def stop_claude(self):
        """Stop any running Claude Code process."""
        if self._worker:
            self._worker.cancel()
            self._worker.wait(2000)
            self._worker = None

    def _restart_claude(self):
        """Start a new Claude Code session."""
        self.stop_claude()
        self._session_id = None
        self._is_first_message = True
        self.output.clear()
        self._code_blocks = []  # Reset code blocks for copy
        self.status_label.setText("Claude Code - Ready")
        self.status_dot.set_color(Theme.SUCCESS)
        self._append_output("New session started. Type a message to begin.\n", Theme.TEXT_MUTED)

    def _send_input(self):
        """Send message to Claude Code."""
        text = self.input_field.text().strip()
        if not text:
            return

        # Don't send if already processing
        if self._worker and self._worker.isRunning():
            return

        self.input_field.clear()

        # Show user message with gap before response
        self._append_output(f"\n> {text}\n\n", Theme.CLAUDE_COLOR)

        # Update status
        self.status_label.setText("Claude Code - Thinking...")
        self.status_dot.set_color(Theme.WARNING)
        self.send_btn.setEnabled(False)
        self.input_field.setEnabled(False)

        # Start status tracking
        self._start_time = time.time()
        self._input_tokens = 0
        self._output_tokens = 0
        self._spinner_index = 0
        self._text_buffer = ""
        self._in_code_block = False
        self.activity_label.setText("Thinking...")
        self.tokens_label.setText("")
        self.time_label.setText("0s")
        self.status_bar.show()
        self._status_timer.start(100)  # Update every 100ms

        # Create and start worker
        # Use --continue only if not the first message (to maintain context)
        self._worker = ClaudeCodeWorker(
            message=text,
            working_dir=self._working_dir,
            continue_session=not self._is_first_message,
            parent=self
        )
        self._worker.text_output.connect(self._on_text)
        self._worker.tool_started.connect(self._on_tool_start)
        self._worker.tool_input.connect(self._on_tool_input)
        self._worker.tool_completed.connect(self._on_tool_complete)
        self._worker.usage_updated.connect(self._on_usage_updated)
        self._worker.finished.connect(self._on_finished)
        self._worker.error.connect(self._on_error)
        self._worker.start()

        self._is_first_message = False

    @Slot(str)
    def _on_text(self, text: str):
        """Handle streaming text output with markdown formatting."""
        self._text_buffer += text

        # Process complete lines from buffer
        while '\n' in self._text_buffer:
            line, self._text_buffer = self._text_buffer.split('\n', 1)
            self._process_markdown_line(line + '\n')

    def _process_markdown_line(self, line: str):
        """Process a single line with markdown formatting."""
        stripped = line.rstrip('\n')

        # Handle code block start/end
        if stripped.startswith('```'):
            if self._in_code_block:
                # End code block - save content and add to list
                self._code_blocks.append(self._current_code_block)
                self._current_code_block = ""
                self._in_code_block = False
                self._code_block_lang = ""
                self._append_raw_html(f'</pre></div>')
            else:
                # Start code block
                self._in_code_block = True
                self._current_code_block = ""
                self._code_block_lang = stripped[3:].strip()
                lang_label = f'<span style="color: {Theme.TEXT_MUTED}; font-size: 10px;">{self._code_block_lang}</span>' if self._code_block_lang else ''
                code_block_id = len(self._code_blocks)
                copy_btn = f'<a href="copy://{code_block_id}" style="color: {Theme.ACCENT_BLUE}; text-decoration: none; font-size: 10px;">Copy</a>'
                self._append_raw_html(
                    f'<div style="background: {Theme.BG_DARK}; border: 1px solid {Theme.BORDER}; '
                    f'border-radius: 4px; margin: 8px 0; padding: 0;">'
                    f'<div style="background: {Theme.BG_TERTIARY}; padding: 4px 8px; '
                    f'border-bottom: 1px solid {Theme.BORDER}; font-size: 11px; display: flex; justify-content: space-between;">'
                    f'{lang_label}<span style="margin-left: auto;">{copy_btn}</span></div>'
                    f'<pre style="margin: 0; padding: 8px; overflow-x: auto; font-family: Menlo, monospace; font-size: 12px;">'
                )
            return

        # Inside code block - accumulate content and append with syntax coloring
        if self._in_code_block:
            self._current_code_block += line
            escaped = line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            self._append_raw_html(f'<span style="color: {Theme.TEXT_PRIMARY};">{escaped}</span>')
            return

        # Format markdown elements
        html = self._format_markdown(stripped)
        if html:
            self._append_raw_html(html + '<br>')

    def _format_markdown(self, text: str) -> str:
        """Convert markdown text to HTML."""
        if not text:
            return ''

        # Escape HTML first
        text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        # Headers
        if text.startswith('### '):
            return f'<span style="color: {Theme.TEXT_PRIMARY}; font-size: 14px; font-weight: bold;">{text[4:]}</span>'
        if text.startswith('## '):
            return f'<span style="color: {Theme.TEXT_PRIMARY}; font-size: 15px; font-weight: bold;">{text[3:]}</span>'
        if text.startswith('# '):
            return f'<span style="color: {Theme.TEXT_PRIMARY}; font-size: 16px; font-weight: bold;">{text[2:]}</span>'

        # Bullet points
        if text.startswith('- ') or text.startswith('* '):
            return f'<span style="color: {Theme.CLAUDE_COLOR};">-</span> {self._format_inline(text[2:])}'
        if text.startswith('  - ') or text.startswith('  * '):
            return f'&nbsp;&nbsp;<span style="color: {Theme.TEXT_MUTED};">-</span> {self._format_inline(text[4:])}'

        # Numbered lists (supports multi-digit numbers like 10, 11, etc.)
        import re
        num_match = re.match(r'^(\d+)\.\s', text)
        if num_match:
            num = num_match.group(1)
            rest = text[num_match.end():]
            return f'<span style="color: {Theme.ACCENT_BLUE};">{num}.</span> {self._format_inline(rest)}'

        # Blockquotes
        if text.startswith('> '):
            return f'<span style="color: {Theme.TEXT_MUTED}; border-left: 2px solid {Theme.CLAUDE_COLOR}; padding-left: 8px; display: inline-block;">{self._format_inline(text[2:])}</span>'

        # Regular text with inline formatting
        return self._format_inline(text)

    def _format_inline(self, text: str) -> str:
        """Format inline markdown elements (bold, italic, code)."""
        # Inline code `code`
        result = ""
        i = 0
        while i < len(text):
            if text[i] == '`' and i + 1 < len(text):
                # Find closing backtick
                end = text.find('`', i + 1)
                if end != -1:
                    code = text[i+1:end]
                    result += f'<span style="background: {Theme.BG_TERTIARY}; color: {Theme.ACCENT_ORANGE}; padding: 1px 4px; border-radius: 3px; font-family: Menlo, monospace; font-size: 11px;">{code}</span>'
                    i = end + 1
                    continue
            result += text[i]
            i += 1
        text = result

        # Bold **text**
        parts = text.split('**')
        if len(parts) > 1:
            result = parts[0]
            for i, part in enumerate(parts[1:], 1):
                if i % 2 == 1:
                    result += f'<b style="color: {Theme.TEXT_PRIMARY};">{part}</b>'
                else:
                    result += part
            text = result

        # Italic *text* (but not **)
        parts = re.split(r'(?<!\*)\*(?!\*)', text)
        if len(parts) > 1:
            result = parts[0]
            for i, part in enumerate(parts[1:], 1):
                if i % 2 == 1:
                    result += f'<i>{part}</i>'
                else:
                    result += part
            text = result

        return f'<span style="color: {Theme.TEXT_PRIMARY};">{text}</span>'

    def _append_raw_html(self, html: str):
        """Append raw HTML to output."""
        cursor = self.output.textCursor()
        cursor.movePosition(cursor.MoveOperation.End)
        cursor.insertHtml(html)
        self.output.setTextCursor(cursor)
        self.output.ensureCursorVisible()

    def _on_output_click(self, event):
        """Handle clicks on output area, including copy links."""
        cursor = self.output.cursorForPosition(event.pos())
        anchor = cursor.charFormat().anchorHref()
        if anchor and anchor.startswith("copy://"):
            try:
                block_id = int(anchor.replace("copy://", ""))
                if 0 <= block_id < len(self._code_blocks):
                    code = self._code_blocks[block_id]
                    QApplication.clipboard().setText(code)
                    # Show toast notification
                    ToastNotification.success("Code copied to clipboard", self.window())
            except (ValueError, IndexError):
                pass
        # Call original handler for selection etc.
        QTextEdit.mousePressEvent(self.output, event)

    @Slot(dict)
    def _on_tool_start(self, event: dict):
        """Show tool execution starting."""
        tool_name = event.get("tool_name", "Unknown")
        self.activity_label.setText(f"Using {tool_name}...")

    @Slot(dict)
    def _on_tool_input(self, event: dict):
        """Handle tool input - show diffs for Edit/Write tools."""
        tool_name = event.get("tool_name", "")
        input_data = event.get("input", {})

        if tool_name == "Edit":
            self._display_edit_diff(input_data)
        elif tool_name == "Write":
            self._display_write_info(input_data)
        elif tool_name == "Read":
            self._display_read_info(input_data)
        elif tool_name == "Bash":
            self._display_bash_info(input_data)
        else:
            # Generic tool display
            file_path = input_data.get("file_path", input_data.get("path", ""))
            if file_path:
                self._append_output(f"\n> {tool_name}({file_path})\n", Theme.CLAUDE_COLOR)
            else:
                self._append_output(f"\n> {tool_name}\n", Theme.CLAUDE_COLOR)

    def _display_edit_diff(self, input_data: dict):
        """Display a diff view for Edit tool."""
        file_path = input_data.get("file_path", "unknown")
        old_string = input_data.get("old_string", "")
        new_string = input_data.get("new_string", "")

        # Count lines changed
        old_lines = old_string.split('\n') if old_string else []
        new_lines = new_string.split('\n') if new_string else []
        removed = len(old_lines)
        added = len(new_lines)

        # Header
        self._append_output(f"\n> Edit({file_path})\n", Theme.CLAUDE_COLOR)
        self._append_output(f"  | Added {added} lines, removed {removed} lines\n", Theme.TEXT_MUTED)

        # Show diff (limited to avoid huge outputs)
        max_context_lines = 10
        if old_lines or new_lines:
            # Show removed lines (red with -)
            for i, line in enumerate(old_lines[:max_context_lines]):
                escaped = line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                self._append_output(f"    - {escaped}\n", Theme.ERROR)
            if len(old_lines) > max_context_lines:
                self._append_output(f"    ... ({len(old_lines) - max_context_lines} more lines)\n", Theme.TEXT_MUTED)

            # Show added lines (green with +)
            for i, line in enumerate(new_lines[:max_context_lines]):
                escaped = line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                self._append_output(f"    + {escaped}\n", Theme.SUCCESS)
            if len(new_lines) > max_context_lines:
                self._append_output(f"    ... ({len(new_lines) - max_context_lines} more lines)\n", Theme.TEXT_MUTED)

    def _display_write_info(self, input_data: dict):
        """Display info for Write tool."""
        file_path = input_data.get("file_path", "unknown")
        content = input_data.get("content", "")
        lines = len(content.split('\n')) if content else 0

        self._append_output(f"\n> Write({file_path})\n", Theme.CLAUDE_COLOR)
        self._append_output(f"  | Writing {lines} lines\n", Theme.TEXT_MUTED)

        # Show preview of content
        if content:
            preview_lines = content.split('\n')[:5]
            for line in preview_lines:
                escaped = line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                self._append_output(f"    {escaped}\n", Theme.TEXT_MUTED)
            if len(content.split('\n')) > 5:
                self._append_output(f"    ... ({len(content.split(chr(10))) - 5} more lines)\n", Theme.TEXT_MUTED)

    def _display_read_info(self, input_data: dict):
        """Display info for Read tool."""
        file_path = input_data.get("file_path", "unknown")
        self._append_output(f"\n> Read({file_path})\n", Theme.CLAUDE_COLOR)

    def _display_bash_info(self, input_data: dict):
        """Display info for Bash tool."""
        command = input_data.get("command", "")
        description = input_data.get("description", "")

        if description:
            self._append_output(f"\n> Bash({description})\n", Theme.CLAUDE_COLOR)
        else:
            # Show truncated command
            cmd_preview = command[:50] + "..." if len(command) > 50 else command
            self._append_output(f"\n> Bash({cmd_preview})\n", Theme.CLAUDE_COLOR)

        # Show full command
        if command:
            escaped = command.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            self._append_output(f"  | {escaped}\n", Theme.TEXT_MUTED)

    @Slot(dict)
    def _on_tool_complete(self, event: dict):
        """Show tool execution completed."""
        self.activity_label.setText("Thinking...")

    @Slot(int, int)
    def _on_usage_updated(self, input_tokens: int, output_tokens: int):
        """Handle token usage updates."""
        self._input_tokens = input_tokens
        self._output_tokens = output_tokens

    @Slot(int, str, float)
    def _on_finished(self, exit_code: int, session_id: str, cost_usd: float):
        """Handle completion."""
        # Flush any remaining text buffer
        if self._text_buffer:
            self._process_markdown_line(self._text_buffer)
            self._text_buffer = ""
        # Close any open code block
        if self._in_code_block:
            self._append_raw_html('</pre></div>')
            self._in_code_block = False

        # Stop status tracking
        self._status_timer.stop()
        self.status_bar.hide()

        self._session_id = session_id
        self._worker = None
        self.status_label.setText("Claude Code - Ready")
        self.status_dot.set_color(Theme.SUCCESS)
        self.send_btn.setEnabled(True)
        self.input_field.setEnabled(True)
        self.input_field.setFocus()

        # Show completion info
        elapsed = time.time() - self._start_time if self._start_time else 0
        total_tokens = self._input_tokens + self._output_tokens
        info_parts = []
        if elapsed > 0:
            info_parts.append(f"{elapsed:.1f}s")
        if total_tokens > 0:
            info_parts.append(f"{total_tokens:,} tokens")
        if cost_usd > 0:
            info_parts.append(f"${cost_usd:.4f}")

        if exit_code != 0:
            self._append_output(f"\n[Completed with code: {exit_code}]\n", Theme.TEXT_MUTED)
        elif info_parts:
            self._append_output(f"\n[{' | '.join(info_parts)}]\n", Theme.TEXT_MUTED)
        else:
            self._append_output("\n", None)

    @Slot(str)
    def _on_error(self, error: str):
        """Handle errors."""
        # Stop status tracking
        self._status_timer.stop()
        self.status_bar.hide()

        self._worker = None
        self.status_label.setText("Claude Code - Error")
        self.status_dot.set_color(Theme.ERROR)
        self.send_btn.setEnabled(True)
        self.input_field.setEnabled(True)
        self._append_output(f"\nError: {error}\n", Theme.ERROR)

    def _update_status_display(self):
        """Update the status bar display (called by timer)."""
        # Update spinner
        self._spinner_index = (self._spinner_index + 1) % len(self._spinner_frames)
        self.spinner_label.setText(self._spinner_frames[self._spinner_index])

        # Update elapsed time
        if self._start_time:
            elapsed = time.time() - self._start_time
            if elapsed >= 60:
                mins = int(elapsed // 60)
                secs = int(elapsed % 60)
                self.time_label.setText(f"{mins}m {secs}s")
            else:
                self.time_label.setText(f"{elapsed:.0f}s")

        # Update tokens
        total_tokens = self._input_tokens + self._output_tokens
        if total_tokens > 0:
            self.tokens_label.setText(f"{total_tokens:,} tokens")

    def _append_output(self, text: str, color: str = None):
        """Append text to output with optional color."""
        cursor = self.output.textCursor()
        cursor.movePosition(cursor.MoveOperation.End)
        if color:
            # Escape HTML and handle newlines
            escaped = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            escaped = escaped.replace('\n', '<br>')
            cursor.insertHtml(f'<span style="color: {color};">{escaped}</span>')
        else:
            cursor.insertHtml(text.replace('\n', '<br>'))
        self.output.setTextCursor(cursor)
        self.output.ensureCursorVisible()

    def clear(self):
        """Clear the output."""
        self.output.clear()

    def cleanup(self):
        """Clean up resources before destruction."""
        self._status_timer.stop()
        if self._worker:
            self._worker.cancel()
            self._worker.wait(1000)
            self._worker = None

    def closeEvent(self, event):
        """Handle widget close - cleanup resources."""
        self.cleanup()
        super().closeEvent(event)


# ============================================================================
# Cisco Code Panel - Same UI as ClaudeCodePanel for Cisco AI
# ============================================================================

class CiscoCodePanel(QWidget):
    """Cisco AI panel with Claude Code-style terminal interface.

    Uses the exact same UI as ClaudeCodePanel but connects to AgentWorker
    instead of spawning a CLI subprocess.
    """

    message_sent = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._working_dir: Optional[str] = None
        self._is_processing = False

        # Status tracking
        self._start_time: Optional[float] = None
        self._total_tokens = 0
        self._total_cost = 0.0
        self._spinner_frames = ["|", "/", "-", "\\"]  # ASCII spinner
        self._spinner_index = 0

        # Text accumulation for markdown formatting
        self._text_buffer = ""
        self._in_code_block = False
        self._code_block_lang = ""
        self._code_blocks: List[str] = []  # Store code block contents for copy
        self._current_code_block = ""  # Current code block being accumulated

        # Timer for status updates
        self._status_timer = QTimer(self)
        self._status_timer.timeout.connect(self._update_status_display)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Header
        header = QWidget()
        header.setFixedHeight(32)
        header.setStyleSheet(f"background: {Theme.BG_SECONDARY}; border-bottom: 1px solid {Theme.BORDER};")
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(12, 0, 12, 0)

        self.status_dot = StatusDot(Theme.TEXT_MUTED, 6)
        header_layout.addWidget(self.status_dot)

        self.status_label = QLabel("Circuit AI")
        self.status_label.setStyleSheet(f"color: {Theme.ACCENT_BLUE}; font-size: 11px; font-weight: bold;")
        header_layout.addWidget(self.status_label)

        header_layout.addStretch()

        # Restart button
        self.restart_btn = IconButton(Icons.refresh, "New Conversation")
        self.restart_btn.clicked.connect(self._restart)
        header_layout.addWidget(self.restart_btn)

        layout.addWidget(header)

        # Terminal output area
        self.output = QTextEdit()
        self.output.setReadOnly(True)
        self.output.setFont(QFont("Menlo", 12))
        self.output.setStyleSheet(f"""
            QTextEdit {{
                background: {Theme.BG_DARK};
                color: {Theme.TEXT_PRIMARY};
                border: none;
                padding: 12px;
                selection-background-color: {Theme.BG_SELECTED};
            }}
        """)
        self.output.setAcceptRichText(True)
        self.output.setMouseTracking(True)
        self.output.mousePressEvent = self._on_output_click
        layout.addWidget(self.output)

        # Status bar (visible while processing)
        self.status_bar = QWidget()
        self.status_bar.setFixedHeight(28)
        self.status_bar.setStyleSheet(f"background: {Theme.BG_TERTIARY}; border-top: 1px solid {Theme.BORDER};")
        status_layout = QHBoxLayout(self.status_bar)
        status_layout.setContentsMargins(12, 0, 12, 0)
        status_layout.setSpacing(8)

        self.spinner_label = QLabel("|")
        self.spinner_label.setStyleSheet(f"color: {Theme.ACCENT_BLUE}; font-size: 14px; font-family: monospace;")
        self.spinner_label.setFixedWidth(16)
        status_layout.addWidget(self.spinner_label)

        self.activity_label = QLabel("Thinking...")
        self.activity_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
        status_layout.addWidget(self.activity_label)

        status_layout.addStretch()

        self.tokens_label = QLabel("")
        self.tokens_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
        status_layout.addWidget(self.tokens_label)

        self.time_label = QLabel("")
        self.time_label.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-size: 11px;")
        status_layout.addWidget(self.time_label)

        self.status_bar.hide()  # Hidden by default
        layout.addWidget(self.status_bar)

        # Input area
        input_container = QWidget()
        input_container.setStyleSheet(f"background: {Theme.BG_SECONDARY}; border-top: 1px solid {Theme.BORDER};")
        input_layout = QHBoxLayout(input_container)
        input_layout.setContentsMargins(12, 10, 12, 10)
        input_layout.setSpacing(10)

        # Prompt marker
        prompt = QLabel(">")
        prompt.setStyleSheet(f"color: {Theme.ACCENT_BLUE}; font-size: 14px; font-weight: bold;")
        prompt.setFixedWidth(16)
        input_layout.addWidget(prompt)

        self.input_field = CompactLineEdit("")
        self.input_field.setPlaceholderText("Message Circuit AI...")
        self.input_field.returnPressed.connect(self._send_input)
        self.input_field.setStyleSheet(f"""
            QLineEdit {{
                background: {Theme.BG_INPUT};
                border: 1px solid {Theme.BORDER};
                border-radius: 4px;
                padding: 8px 12px;
                color: {Theme.TEXT_PRIMARY};
                font-size: 13px;
                font-family: Menlo, monospace;
            }}
            QLineEdit:focus {{ border-color: {Theme.ACCENT_BLUE}; }}
        """)
        input_layout.addWidget(self.input_field)

        self.send_btn = CompactButton("Send")
        self.send_btn.setFixedSize(60, 32)
        self.send_btn.clicked.connect(self._send_input)
        self.send_btn.setStyleSheet(f"""
            QPushButton {{
                background: {Theme.ACCENT_BLUE};
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 12px;
                font-weight: 500;
            }}
            QPushButton:hover {{ background: #3b82f6; }}
            QPushButton:pressed {{ background: #1d4ed8; }}
        """)
        input_layout.addWidget(self.send_btn)

        layout.addWidget(input_container)

    def set_working_dir(self, path: str):
        """Set the working directory."""
        self._working_dir = path

    def set_connected(self, connected: bool):
        """Update connection status display."""
        if connected:
            self.status_label.setText("Circuit AI - Ready")
            self.status_dot.set_color(Theme.SUCCESS)
            self._append_output("Circuit AI connected and ready. Type a message to begin.\n", Theme.TEXT_MUTED)
        else:
            self.status_label.setText("Circuit AI - Disconnected")
            self.status_dot.set_color(Theme.TEXT_MUTED)

    def _restart(self):
        """Start a new conversation."""
        self.output.clear()
        self._code_blocks = []  # Reset code blocks for copy
        self._text_buffer = ""
        self._in_code_block = False
        self._total_tokens = 0
        self._total_cost = 0.0
        self.status_label.setText("Circuit AI - Ready")
        self.status_dot.set_color(Theme.SUCCESS)
        self._append_output("New conversation started.\n", Theme.TEXT_MUTED)

    def _send_input(self):
        """Send message to Circuit AI."""
        text = self.input_field.text().strip()
        if not text or self._is_processing:
            return

        self.input_field.clear()

        # Show user message with gap before response
        self._append_output(f"\n> {text}\n\n", Theme.CIRCUIT_COLOR)

        # Update status
        self.status_label.setText("Circuit AI - Thinking...")
        self.status_dot.set_color(Theme.WARNING)
        self.send_btn.setEnabled(False)
        self.input_field.setEnabled(False)
        self._is_processing = True

        # Start status tracking
        self._start_time = time.time()
        self._spinner_index = 0
        self._text_buffer = ""
        self._in_code_block = False
        self.activity_label.setText("Thinking...")
        self.tokens_label.setText("")
        self.time_label.setText("0s")
        self.status_bar.show()
        self._status_timer.start(100)

        # Emit signal for the main window to send to AgentWorker
        self.message_sent.emit(text)

    # Public slots for AgentWorker signals
    @Slot(str)
    def on_chunk(self, chunk: str):
        """Handle streaming message chunk from AgentWorker."""
        self._text_buffer += chunk

        # Process complete lines from buffer
        while '\n' in self._text_buffer:
            line, self._text_buffer = self._text_buffer.split('\n', 1)
            self._process_markdown_line(line + '\n')

    @Slot(str, str, str)
    def on_tool_call(self, name: str, detail: str, status: str):
        """Handle tool call from AgentWorker."""
        if status == "running":
            self.activity_label.setText(f"Using {name}...")
            self._append_output(f"\n> {name}", Theme.ACCENT_BLUE)
            if detail:
                # Truncate long details
                detail_preview = detail[:50] + "..." if len(detail) > 50 else detail
                self._append_output(f"({detail_preview})", Theme.TEXT_MUTED)
            self._append_output("\n", None)
        elif status == "done":
            self.activity_label.setText("Thinking...")
        elif status == "error":
            self._append_output(f"  | Error: {detail}\n", Theme.ERROR)

    @Slot(str)
    def on_complete(self, response: str):
        """Handle message completion from AgentWorker."""
        # Flush any remaining text buffer
        if self._text_buffer:
            self._process_markdown_line(self._text_buffer)
            self._text_buffer = ""
        # Close any open code block
        if self._in_code_block:
            self._append_raw_html('</pre></div>')
            self._in_code_block = False

        # Stop status tracking
        self._status_timer.stop()
        self.status_bar.hide()
        self._is_processing = False

        self.status_label.setText("Circuit AI - Ready")
        self.status_dot.set_color(Theme.SUCCESS)
        self.send_btn.setEnabled(True)
        self.input_field.setEnabled(True)
        self.input_field.setFocus()

        # Show completion info
        elapsed = time.time() - self._start_time if self._start_time else 0
        info_parts = []
        if elapsed > 0:
            info_parts.append(f"{elapsed:.1f}s")
        if self._total_tokens > 0:
            info_parts.append(f"{self._total_tokens:,} tokens")
        if self._total_cost > 0:
            info_parts.append(f"${self._total_cost:.4f}")

        if info_parts:
            self._append_output(f"\n[{' | '.join(info_parts)}]\n", Theme.TEXT_MUTED)
        else:
            self._append_output("\n", None)

    @Slot(str)
    def on_error(self, error: str):
        """Handle error from AgentWorker."""
        # Stop status tracking
        self._status_timer.stop()
        self.status_bar.hide()
        self._is_processing = False

        self.status_label.setText("Circuit AI - Error")
        self.status_dot.set_color(Theme.ERROR)
        self.send_btn.setEnabled(True)
        self.input_field.setEnabled(True)
        self._append_output(f"\nError: {error}\n", Theme.ERROR)

    @Slot(int, float)
    def on_stats(self, tokens: int, cost: float):
        """Handle stats update from AgentWorker."""
        self._total_tokens = tokens
        self._total_cost = cost

    def _process_markdown_line(self, line: str):
        """Process a single line with markdown formatting."""
        stripped = line.rstrip('\n')

        # Handle code block start/end
        if stripped.startswith('```'):
            if self._in_code_block:
                # End code block - save content and add to list
                self._code_blocks.append(self._current_code_block)
                self._current_code_block = ""
                self._in_code_block = False
                self._code_block_lang = ""
                self._append_raw_html(f'</pre></div>')
            else:
                # Start code block
                self._in_code_block = True
                self._current_code_block = ""
                self._code_block_lang = stripped[3:].strip()
                lang_label = f'<span style="color: {Theme.TEXT_MUTED}; font-size: 10px;">{self._code_block_lang}</span>' if self._code_block_lang else ''
                code_block_id = len(self._code_blocks)
                copy_btn = f'<a href="copy://{code_block_id}" style="color: {Theme.ACCENT_BLUE}; text-decoration: none; font-size: 10px;">Copy</a>'
                self._append_raw_html(
                    f'<div style="background: {Theme.BG_DARK}; border: 1px solid {Theme.BORDER}; '
                    f'border-radius: 4px; margin: 8px 0; padding: 0;">'
                    f'<div style="background: {Theme.BG_TERTIARY}; padding: 4px 8px; '
                    f'border-bottom: 1px solid {Theme.BORDER}; font-size: 11px; display: flex; justify-content: space-between;">'
                    f'{lang_label}<span style="margin-left: auto;">{copy_btn}</span></div>'
                    f'<pre style="margin: 0; padding: 8px; overflow-x: auto; font-family: Menlo, monospace; font-size: 12px;">'
                )
            return

        # Inside code block - accumulate content and append with syntax coloring
        if self._in_code_block:
            self._current_code_block += line
            escaped = line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            self._append_raw_html(f'<span style="color: {Theme.TEXT_PRIMARY};">{escaped}</span>')
            return

        # Format markdown elements
        html = self._format_markdown(stripped)
        if html:
            self._append_raw_html(html + '<br>')

    def _format_markdown(self, text: str) -> str:
        """Convert markdown text to HTML."""
        if not text:
            return ''

        # Escape HTML first
        text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        # Headers
        if text.startswith('### '):
            return f'<span style="color: {Theme.TEXT_PRIMARY}; font-size: 14px; font-weight: bold;">{text[4:]}</span>'
        if text.startswith('## '):
            return f'<span style="color: {Theme.TEXT_PRIMARY}; font-size: 15px; font-weight: bold;">{text[3:]}</span>'
        if text.startswith('# '):
            return f'<span style="color: {Theme.TEXT_PRIMARY}; font-size: 16px; font-weight: bold;">{text[2:]}</span>'

        # Bullet points
        if text.startswith('- ') or text.startswith('* '):
            return f'<span style="color: {Theme.CIRCUIT_COLOR};">-</span> {self._format_inline(text[2:])}'
        if text.startswith('  - ') or text.startswith('  * '):
            return f'&nbsp;&nbsp;<span style="color: {Theme.TEXT_MUTED};">-</span> {self._format_inline(text[4:])}'

        # Numbered lists (supports multi-digit numbers like 10, 11, etc.)
        import re
        num_match = re.match(r'^(\d+)\.\s', text)
        if num_match:
            num = num_match.group(1)
            rest = text[num_match.end():]
            return f'<span style="color: {Theme.CIRCUIT_COLOR};">{num}.</span> {self._format_inline(rest)}'

        # Blockquotes
        if text.startswith('> '):
            return f'<span style="color: {Theme.TEXT_MUTED}; border-left: 2px solid {Theme.CIRCUIT_COLOR}; padding-left: 8px; display: inline-block;">{self._format_inline(text[2:])}</span>'

        # Regular text with inline formatting
        return self._format_inline(text)

    def _format_inline(self, text: str) -> str:
        """Format inline markdown elements (bold, italic, code)."""
        # Inline code `code`
        result = ""
        i = 0
        while i < len(text):
            if text[i] == '`' and i + 1 < len(text):
                # Find closing backtick
                end = text.find('`', i + 1)
                if end != -1:
                    code = text[i+1:end]
                    result += f'<span style="background: {Theme.BG_TERTIARY}; color: {Theme.ACCENT_ORANGE}; padding: 1px 4px; border-radius: 3px; font-family: Menlo, monospace; font-size: 11px;">{code}</span>'
                    i = end + 1
                    continue
            result += text[i]
            i += 1
        text = result

        # Bold **text**
        parts = text.split('**')
        if len(parts) > 1:
            result = parts[0]
            for i, part in enumerate(parts[1:], 1):
                if i % 2 == 1:
                    result += f'<b style="color: {Theme.TEXT_PRIMARY};">{part}</b>'
                else:
                    result += part
            text = result

        # Italic *text* (but not **)
        parts = re.split(r'(?<!\*)\*(?!\*)', text)
        if len(parts) > 1:
            result = parts[0]
            for i, part in enumerate(parts[1:], 1):
                if i % 2 == 1:
                    result += f'<i>{part}</i>'
                else:
                    result += part
            text = result

        return f'<span style="color: {Theme.TEXT_PRIMARY};">{text}</span>'

    def _append_raw_html(self, html: str):
        """Append raw HTML to output."""
        cursor = self.output.textCursor()
        cursor.movePosition(cursor.MoveOperation.End)
        cursor.insertHtml(html)
        self.output.setTextCursor(cursor)
        self.output.ensureCursorVisible()

    def _on_output_click(self, event):
        """Handle clicks on output area, including copy links."""
        cursor = self.output.cursorForPosition(event.pos())
        anchor = cursor.charFormat().anchorHref()
        if anchor and anchor.startswith("copy://"):
            try:
                block_id = int(anchor.replace("copy://", ""))
                if 0 <= block_id < len(self._code_blocks):
                    code = self._code_blocks[block_id]
                    QApplication.clipboard().setText(code)
                    # Show toast notification
                    ToastNotification.success("Code copied to clipboard", self.window())
            except (ValueError, IndexError):
                pass
        # Call original handler for selection etc.
        QTextEdit.mousePressEvent(self.output, event)

    def _update_status_display(self):
        """Update the status bar display (called by timer)."""
        # Update spinner
        self._spinner_index = (self._spinner_index + 1) % len(self._spinner_frames)
        self.spinner_label.setText(self._spinner_frames[self._spinner_index])

        # Update elapsed time
        if self._start_time:
            elapsed = time.time() - self._start_time
            if elapsed >= 60:
                mins = int(elapsed // 60)
                secs = int(elapsed % 60)
                self.time_label.setText(f"{mins}m {secs}s")
            else:
                self.time_label.setText(f"{elapsed:.0f}s")

        # Update tokens
        if self._total_tokens > 0:
            self.tokens_label.setText(f"{self._total_tokens:,} tokens")

    def _append_output(self, text: str, color: str = None):
        """Append text to output with optional color."""
        cursor = self.output.textCursor()
        cursor.movePosition(cursor.MoveOperation.End)
        if color:
            # Escape HTML and handle newlines
            escaped = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            escaped = escaped.replace('\n', '<br>')
            cursor.insertHtml(f'<span style="color: {color};">{escaped}</span>')
        else:
            cursor.insertHtml(text.replace('\n', '<br>'))
        self.output.setTextCursor(cursor)
        self.output.ensureCursorVisible()

    def clear(self):
        """Clear the output."""
        self.output.clear()
        self._text_buffer = ""
        self._in_code_block = False

    def cleanup(self):
        """Clean up resources before destruction."""
        self._status_timer.stop()

    def closeEvent(self, event):
        """Handle widget close - cleanup resources."""
        self.cleanup()
        super().closeEvent(event)


# ============================================================================
# Agent Worker
# ============================================================================

class AgentWorker(QObject):
    """Background worker for agent operations with persistent event loop.

    Supports both Cisco AI and Anthropic/Claude as providers.
    """

    # Signals
    connected = Signal(bool, str)  # success, error_message
    service_ready = Signal(bool, str)  # success, error_message
    message_chunk = Signal(str)
    message_complete = Signal(str)
    message_error = Signal(str)
    tool_call = Signal(str, str, str)
    stats_updated = Signal(int, float)
    confirmation_needed = Signal(object)  # ConfirmationRequest
    mcp_status = Signal(bool, str, int)  # connected, server_id, tool_count

    def __init__(self):
        super().__init__()
        self.service = None  # Cisco AgentService
        self.anthropic_client = None  # Anthropic client
        self.provider = "cisco"  # 'cisco' or 'anthropic'
        self.model = "gpt-4.1"
        self.working_dir = "."
        self._loop = None
        self._loop_thread = None
        self._running = False
        self._conversation_history = []  # For Anthropic
        self._total_tokens = 0
        self._total_cost = 0.0

    def start_event_loop(self):
        """Start the persistent event loop in a background thread."""
        if self._running:
            return

        self._loop = asyncio.new_event_loop()

        def run_loop():
            asyncio.set_event_loop(self._loop)
            self._running = True
            self._loop.run_forever()

        self._loop_thread = threading.Thread(target=run_loop, daemon=True)
        self._loop_thread.start()

    def stop_event_loop(self):
        """Stop the event loop and cleanup."""
        if self._loop and self._running:
            self._loop.call_soon_threadsafe(self._loop.stop)
            self._running = False
            if self._loop_thread:
                self._loop_thread.join(timeout=2.0)
            self._loop = None

    def set_provider(self, provider: str):
        """Set the AI provider (cisco or anthropic)."""
        self.provider = provider

    def init_service(self, working_dir: str, provider: str = "cisco"):
        """Initialize the agent service for the specified provider."""
        self.working_dir = working_dir
        self.provider = provider

        if provider == "anthropic":
            self._init_anthropic()
        else:
            self._init_cisco(working_dir)

    def _init_cisco(self, working_dir: str):
        """Initialize Cisco AgentService."""
        try:
            from circuit_agent.service import AgentService, EventType

            self.service = AgentService(
                working_dir=working_dir,
                model=self.model if self.model in ["gpt-4.1", "gpt-4o", "gpt-4o-mini", "o4-mini", "o1"] else "gpt-4.1",
                auto_approve=False,
            )

            # Register event handlers
            self.service.on(EventType.MESSAGE_CHUNK, self._on_chunk)
            self.service.on(EventType.TOOL_CALL_STARTED, self._on_tool_started)
            self.service.on(EventType.TOOL_CALL_COMPLETED, self._on_tool_done)
            self.service.on(EventType.TOOL_CALL_ERROR, self._on_tool_error)
            self.service.on(EventType.CONFIRMATION_NEEDED, self._on_confirmation_needed)
            self.service.on(EventType.TOKENS_UPDATED, self._on_tokens_updated)

            # Start the persistent event loop
            self.start_event_loop()

            self.service_ready.emit(True, "")

        except Exception as e:
            error_msg = str(e)
            self.service_ready.emit(False, error_msg)

    def _init_anthropic(self):
        """Initialize Anthropic client."""
        try:
            from circuit_agent.config import load_anthropic_key

            api_key = load_anthropic_key()
            if not api_key:
                self.service_ready.emit(False, "Anthropic API key not configured")
                return

            # Try to import anthropic
            try:
                import anthropic
                self.anthropic_client = anthropic.Anthropic(api_key=api_key)
            except ImportError:
                self.service_ready.emit(False, "anthropic package not installed. Run: pip install anthropic")
                return

            self._conversation_history = []
            self.service_ready.emit(True, "")

        except Exception as e:
            self.service_ready.emit(False, str(e))

    def _on_chunk(self, event):
        """Handle streaming message chunks."""
        chunk = event.data.get("chunk", "")
        if chunk:
            self.message_chunk.emit(chunk)

    def _on_tool_started(self, event):
        """Handle tool execution start."""
        tc = event.data.get("tool_call")
        if tc:
            self.tool_call.emit(tc.name, tc.detail, "running")

    def _on_tool_done(self, event):
        """Handle tool execution completion."""
        tc = event.data.get("tool_call")
        if tc:
            self.tool_call.emit(tc.name, tc.detail, "done")

    def _on_tool_error(self, event):
        """Handle tool execution error."""
        tc = event.data.get("tool_call")
        if tc:
            self.tool_call.emit(tc.name, event.data.get("error", "")[:50], "error")

    def _on_confirmation_needed(self, event):
        """Handle confirmation request from agent."""
        request = event.data.get("request")
        if request:
            self.confirmation_needed.emit(request)

    def _on_tokens_updated(self, event):
        """Handle token usage updates."""
        session_total = event.data.get("session_total", 0)
        # Get cost from service
        if self.service:
            costs = self.service.get_cost_stats()
            cost = costs.get("total_cost_usd", 0)
            self.stats_updated.emit(session_total, cost)

    @Slot()
    def connect_agent(self):
        """Connect to the agent using saved credentials."""
        if self.provider == "anthropic":
            self._connect_anthropic()
        else:
            self._connect_cisco()

    def _connect_cisco(self):
        """Connect using Cisco credentials."""
        if not self.service:
            self.connected.emit(False, "Service not initialized")
            return

        if not self._loop or not self._running:
            self.start_event_loop()

        def do_connect():
            try:
                future = asyncio.run_coroutine_threadsafe(
                    self.service.connect_with_saved_credentials(),
                    self._loop
                )
                result = future.result(timeout=30)

                if result:
                    self.connected.emit(True, "")
                else:
                    error = self.service.state.error or "Connection failed"
                    self.connected.emit(False, error)

            except Exception as e:
                self.connected.emit(False, str(e))

        threading.Thread(target=do_connect, daemon=True).start()

    def _connect_anthropic(self):
        """Connect/verify Anthropic API."""
        if not self.anthropic_client:
            # Try to initialize
            self._init_anthropic()
            if not self.anthropic_client:
                self.connected.emit(False, "Anthropic client not initialized")
                return

        # Test the connection
        def do_test():
            try:
                # Simple API call to verify key works
                import httpx
                from circuit_agent.config import load_anthropic_key

                api_key = load_anthropic_key()
                with httpx.Client(timeout=10.0) as client:
                    response = client.get(
                        "https://api.anthropic.com/v1/models",
                        headers={
                            "x-api-key": api_key,
                            "anthropic-version": "2023-06-01",
                        }
                    )
                    if response.status_code == 200:
                        self.connected.emit(True, "")
                    else:
                        self.connected.emit(False, f"API error: {response.status_code}")
            except Exception as e:
                self.connected.emit(False, str(e))

        threading.Thread(target=do_test, daemon=True).start()

    def reinit_mcp(self):
        """Reinitialize MCP servers after settings change."""
        if self.provider != "cisco" or not self.service:
            return

        def do_reinit():
            try:
                from circuit_agent.config import load_github_pat, load_github_mcp_config
                from circuit_agent.mcp.servers.github import GitHubMCPServer

                # Get current config
                github_config = load_github_mcp_config()
                github_pat = load_github_pat()

                if not github_config.get("enabled") or not github_pat:
                    # Disconnect if disabled
                    if self.service and hasattr(self.service, '_agent') and self.service._agent:
                        self.service._agent.disconnect_mcp("github")
                    self.mcp_status.emit(False, "github", 0)
                    return

                # Connect to GitHub MCP
                if self.service and hasattr(self.service, '_agent') and self.service._agent:
                    agent = self.service._agent
                    toolsets = github_config.get("toolsets", [])

                    # Disconnect existing connection first
                    agent.mcp_manager.disconnect("github")

                    # Create new config
                    config = GitHubMCPServer.get_remote_config(
                        pat=github_pat,
                        toolsets=toolsets,
                        enabled=True
                    )

                    # Connect
                    success = agent.mcp_manager.connect(config)
                    if success:
                        agent._mcp_tools_cache = agent.mcp_manager.list_tools()
                        tool_count = len(agent._mcp_tools_cache)
                        self.mcp_status.emit(True, "github", tool_count)
                    else:
                        self.mcp_status.emit(False, "github", 0)

            except Exception as e:
                self.mcp_status.emit(False, "github", 0)

        threading.Thread(target=do_reinit, daemon=True).start()

    @Slot(str)
    def send_message(self, text: str):
        """Send a message to the agent."""
        if self.provider == "anthropic":
            self._send_message_anthropic(text)
        else:
            self._send_message_cisco(text)

    def _send_message_cisco(self, text: str):
        """Send message via Cisco service."""
        if not self.service:
            self.message_error.emit("Service not initialized")
            return

        if not self._loop or not self._running:
            self.message_error.emit("Event loop not running")
            return

        def do_send():
            try:
                future = asyncio.run_coroutine_threadsafe(
                    self.service.send_message(text),
                    self._loop
                )
                response = future.result(timeout=300)

                if response:
                    self.message_complete.emit(response)
                    stats = self.service.get_token_stats()
                    costs = self.service.get_cost_stats()
                    self.stats_updated.emit(
                        stats.get("session_total", 0),
                        costs.get("total_cost_usd", 0)
                    )
                else:
                    if self.service.state.error:
                        self.message_error.emit(self.service.state.error)
                    else:
                        self.message_error.emit("No response received")

            except asyncio.TimeoutError:
                self.message_error.emit("Request timed out")
            except Exception as e:
                self.message_error.emit(str(e))

        threading.Thread(target=do_send, daemon=True).start()

    def _send_message_anthropic(self, text: str):
        """Send message via Anthropic API with streaming."""
        if not self.anthropic_client:
            self.message_error.emit("Anthropic client not initialized")
            return

        def do_send():
            try:
                # Add user message to history
                self._conversation_history.append({
                    "role": "user",
                    "content": text
                })

                # Create system prompt
                system_prompt = f"""You are an AI coding assistant helping with a software project.
Working directory: {self.working_dir}

Be helpful, concise, and provide code examples when relevant."""

                # Stream the response
                full_response = ""

                with self.anthropic_client.messages.stream(
                    model=self.model,
                    max_tokens=4096,
                    system=system_prompt,
                    messages=self._conversation_history,
                ) as stream:
                    for text_chunk in stream.text_stream:
                        full_response += text_chunk
                        self.message_chunk.emit(text_chunk)

                # Get final message for token counts
                final_message = stream.get_final_message()

                # Add assistant response to history
                self._conversation_history.append({
                    "role": "assistant",
                    "content": full_response
                })

                # Calculate tokens and cost
                input_tokens = final_message.usage.input_tokens
                output_tokens = final_message.usage.output_tokens
                total_tokens = input_tokens + output_tokens
                self._total_tokens += total_tokens

                # Anthropic pricing (approximate)
                # Claude 3.5 Sonnet: $3/$15 per MTok
                # Claude 3 Opus: $15/$75 per MTok
                if "opus" in self.model.lower():
                    cost = (input_tokens * 15 + output_tokens * 75) / 1_000_000
                elif "haiku" in self.model.lower():
                    cost = (input_tokens * 0.25 + output_tokens * 1.25) / 1_000_000
                else:  # Sonnet
                    cost = (input_tokens * 3 + output_tokens * 15) / 1_000_000

                self._total_cost += cost

                self.message_complete.emit(full_response)
                self.stats_updated.emit(self._total_tokens, self._total_cost)

            except Exception as e:
                self.message_error.emit(str(e))

        threading.Thread(target=do_send, daemon=True).start()

    def set_model(self, model: str):
        """Set the AI model."""
        self.model = model
        if self.service and self.provider == "cisco":
            self.service.set_model(model)

    def set_auto_approve(self, enabled: bool):
        """Enable or disable auto-approve mode."""
        if self.service:
            self.service.set_auto_approve(enabled)

    def respond_to_confirmation(self, confirm_id: str, approved: bool):
        """Respond to a tool confirmation request."""
        if self.service and self._loop and self._running:
            asyncio.run_coroutine_threadsafe(
                self._do_respond_confirmation(confirm_id, approved),
                self._loop
            )

    async def _do_respond_confirmation(self, confirm_id: str, approved: bool):
        """Actually respond to confirmation (runs in event loop)."""
        if hasattr(self.service, 'respond_to_confirmation'):
            await self.service.respond_to_confirmation(confirm_id, approved)

    def clear_conversation(self):
        """Clear conversation history (for Anthropic)."""
        self._conversation_history = []
        self._total_tokens = 0
        self._total_cost = 0.0


# ============================================================================
# Enhanced Status Bar
# ============================================================================

class StatusBarItem(QPushButton):
    """Clickable status bar item with hover effects."""

    def __init__(self, text: str = "", tooltip: str = "", parent=None):
        super().__init__(text, parent)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFlat(True)
        self.setToolTip(tooltip)
        self._base_color = Theme.TEXT_MUTED
        self._update_style()

    def _update_style(self):
        self.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                border: none;
                color: {self._base_color};
                font-size: 10px;
                padding: 0 6px;
                min-height: 18px;
                max-height: 18px;
            }}
            QPushButton:hover {{
                background: {Theme.BG_HOVER};
            }}
            QPushButton:pressed {{
                background: {Theme.BG_TERTIARY};
            }}
        """)

    def set_color(self, color: str):
        """Set the text color."""
        self._base_color = color
        self._update_style()


class EnhancedStatusBar(QWidget):
    """VS Code-style status bar with clickable sections."""

    # Signals for clickable items
    branch_clicked = Signal()
    position_clicked = Signal()
    encoding_clicked = Signal()
    filetype_clicked = Signal()
    agent_clicked = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(22)
        self.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_DARK};
                border-top: 1px solid {Theme.BORDER};
            }}
        """)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Left side: Git branch (with special styling)
        self.branch_btn = StatusBarItem("", "Click to view branches")
        self.branch_btn.setStyleSheet(f"""
            QPushButton {{
                background: {Theme.ACCENT_BLUE};
                border: none;
                color: white;
                font-size: 10px;
                padding: 0 8px;
                min-height: 18px;
                max-height: 18px;
            }}
            QPushButton:hover {{
                background: {Theme.TEXT_LINK};
            }}
        """)
        self.branch_btn.clicked.connect(self.branch_clicked.emit)
        self.branch_btn.hide()
        layout.addWidget(self.branch_btn)

        # Errors
        self.errors_btn = StatusBarItem("0", "Errors")
        self.errors_btn.set_color(Theme.ERROR)
        layout.addWidget(self.errors_btn)

        # Warnings
        self.warnings_btn = StatusBarItem("0", "Warnings")
        self.warnings_btn.set_color(Theme.WARNING)
        layout.addWidget(self.warnings_btn)

        # Spacer
        layout.addStretch()

        # Center: status message (non-clickable)
        self.message_label = QLabel("Ready")
        self.message_label.setStyleSheet(f"""
            color: {Theme.TEXT_MUTED};
            font-size: 10px;
            padding: 0 8px;
        """)
        layout.addWidget(self.message_label)

        layout.addStretch()

        # Right side: clickable items
        self.position_btn = StatusBarItem("Ln 1, Col 1", "Go to line")
        self.position_btn.clicked.connect(self.position_clicked.emit)
        layout.addWidget(self.position_btn)

        self.encoding_btn = StatusBarItem("UTF-8", "Encoding")
        self.encoding_btn.clicked.connect(self.encoding_clicked.emit)
        layout.addWidget(self.encoding_btn)

        self.filetype_btn = StatusBarItem("Text", "Language")
        self.filetype_btn.clicked.connect(self.filetype_clicked.emit)
        layout.addWidget(self.filetype_btn)

        # Agent status
        self.agent_btn = StatusBarItem("AI: Off", "AI agent status")
        self.agent_btn.clicked.connect(self.agent_clicked.emit)
        layout.addWidget(self.agent_btn)

        # Token count (hidden - stats shown in AI Agent panel instead)
        self.token_btn = StatusBarItem("0 tokens", "Token usage")
        self.token_btn.set_color(Theme.SUCCESS)
        self.token_btn.hide()  # Removed per user request

    def set_message(self, msg: str):
        """Set the main status message."""
        self.message_label.setText(msg)

    def set_position(self, line: int, col: int):
        """Set cursor position."""
        self.position_btn.setText(f"Ln {line}, Col {col}")

    def set_file_type(self, file_type: str):
        """Set file type indicator."""
        self.filetype_btn.setText(file_type)

    def set_branch(self, branch: str):
        """Set git branch name."""
        if branch:
            self.branch_btn.setText(branch)
            self.branch_btn.setIcon(Icons.git_branch(12, Theme.TEXT_SECONDARY))
            self.branch_btn.show()
        else:
            self.branch_btn.hide()

    def set_agent_status(self, connected: bool, model: str = ""):
        """Set agent connection status."""
        if connected:
            text = f"AI: {model}" if model else "AI: Connected"
            self.agent_btn.setText(text)
            self.agent_btn.setIcon(Icons.check(12, Theme.SUCCESS))
            self.agent_btn.setStyleSheet(f"""
                QPushButton {{
                    background: rgba(35, 134, 54, 0.2);
                    border: none;
                    color: {Theme.SUCCESS};
                    font-size: 11px;
                    padding: 2px 10px;
                    border-radius: 2px;
                }}
                QPushButton:hover {{
                    background: rgba(35, 134, 54, 0.3);
                }}
            """)
        else:
            self.agent_btn.setText("AI: Offline")
            self.agent_btn.set_color(Theme.TEXT_MUTED)

    def set_tokens(self, tokens: int, cost: float = 0.0):
        """Set token usage."""
        if cost > 0:
            self.token_btn.setText(f"{tokens:,} | ${cost:.3f}")
        else:
            self.token_btn.setText(f"{tokens:,} tokens")

    def set_errors(self, errors: int, warnings: int = 0):
        """Set error and warning counts."""
        self.errors_btn.setText(str(errors))
        self.errors_btn.setIcon(Icons.error(12, Theme.ERROR))
        self.warnings_btn.setText(str(warnings))
        self.warnings_btn.setIcon(Icons.warning(12, Theme.WARNING))


# ============================================================================
# Main Window
# ============================================================================

class CircuitIDEWindow(QMainWindow):
    """Main application window."""

    send_message_signal = Signal(str)

    def __init__(self):
        super().__init__()

        self.project_dir: Optional[Path] = None
        self.worker = None
        self.worker_thread = None
        self._total_tokens = 0
        self._total_cost = 0.0

        self.setWindowTitle("Circuit IDE")
        self.setGeometry(100, 100, 1400, 900)
        self.setMinimumSize(1000, 600)

        self._apply_main_style()
        self._setup_ui()
        self._setup_menu()
        self._setup_shortcuts()

    def _apply_main_style(self):
        self.setStyleSheet(f"""
            QMainWindow {{ background: {Theme.BG_MAIN}; }}
            QMenuBar {{
                background: {Theme.BG_DARK};
                color: {Theme.TEXT_PRIMARY};
                padding: 2px 0;
                font-size: 12px;
                border-bottom: 1px solid {Theme.BORDER};
            }}
            QMenuBar::item {{ padding: 4px 10px; background: transparent; }}
            QMenuBar::item:selected {{ background: {Theme.BG_HOVER}; }}
            QMenu {{
                background: {Theme.BG_SECONDARY};
                color: {Theme.TEXT_PRIMARY};
                border: 1px solid {Theme.BORDER};
                padding: 4px;
                font-size: 12px;
            }}
            QMenu::item {{ padding: 6px 20px; }}
            QMenu::item:selected {{ background: {Theme.BG_SELECTED}; }}
            QStatusBar {{
                background: {Theme.BG_DARK};
                color: {Theme.TEXT_MUTED};
                font-size: 11px;
                border-top: 1px solid {Theme.BORDER};
            }}
        """)

    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)

        main = QHBoxLayout(central)
        main.setContentsMargins(0, 0, 0, 0)
        main.setSpacing(0)

        # Activity bar
        self.activity_bar = ActivityBar()
        self.activity_bar.view_changed.connect(self._on_view_changed)
        main.addWidget(self.activity_bar)

        # Side panel stack - only file explorer and search
        self.side_stack = QStackedWidget()
        self.side_stack.setFixedWidth(260)
        self.side_stack.setStyleSheet(f"background: {Theme.BG_SECONDARY};")

        self.file_explorer = FileExplorer()
        self.file_explorer.file_selected.connect(self._open_file_from_explorer)
        self.side_stack.addWidget(self.file_explorer)

        self.search_panel = SearchPanel()
        self.search_panel.file_selected.connect(self._open_file_from_explorer)
        self.side_stack.addWidget(self.search_panel)

        main.addWidget(self.side_stack)

        # Create panels that will open as tabs (but don't add to side stack)
        self.agent_panel = AgentControlPanel()
        self.agent_panel.reconnect_requested.connect(self._reconnect)
        self.agent_panel.clear_chat_requested.connect(self._new_chat)

        self.git_panel = GitPanel()
        self.git_panel.git_operation_completed.connect(self._update_git_branch)

        self.settings_panel = SettingsPanel()
        self.settings_panel.settings_changed.connect(self._on_settings_changed)
        self.settings_panel.provider_changed.connect(self._on_provider_changed)

        # Track which panels are open as tabs
        self._panel_tabs = {}  # "settings" -> tab_index, etc.

        # Main content splitter (horizontal: editor area | chat)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(1)
        splitter.setStyleSheet(f"QSplitter::handle {{ background: {Theme.BORDER}; }}")

        # Editor area with terminal (vertical splitter)
        editor_splitter = QSplitter(Qt.Orientation.Vertical)
        editor_splitter.setHandleWidth(1)
        editor_splitter.setStyleSheet(f"QSplitter::handle {{ background: {Theme.BORDER}; }}")

        # Editor area with split support
        self.editor_tabs = EditorArea()
        self.editor_tabs.cursor_position_changed.connect(self._on_cursor_position)
        self.welcome_screen = WelcomeScreen()
        self.welcome_screen.open_folder_requested.connect(self._open_folder)
        self.welcome_screen.open_file_requested.connect(self._open_file)
        self.welcome_screen.new_file_requested.connect(self._new_file)
        self.welcome_screen.project_selected.connect(self._open_project)
        self.editor_tabs.show_welcome(self.welcome_screen)
        editor_splitter.addWidget(self.editor_tabs)

        # Terminal panel
        self.terminal_panel = TerminalPanel()
        self.terminal_panel.setMinimumHeight(100)
        editor_splitter.addWidget(self.terminal_panel)

        # Set initial sizes (editor takes most space, terminal at bottom)
        editor_splitter.setSizes([500, 200])

        splitter.addWidget(editor_splitter)

        # Chat panel stack (for swapping between Cisco AI and Claude Code)
        self.chat_stack = QStackedWidget()
        self.chat_stack.setMinimumWidth(350)

        # Cisco AI panel (index 0) - Uses same terminal-style UI as Claude Code
        self.cisco_code_panel = CiscoCodePanel()
        self.cisco_code_panel.message_sent.connect(self._on_cisco_message)
        self.chat_stack.addWidget(self.cisco_code_panel)

        # Claude Code panel (index 1)
        self.claude_code_panel = ClaudeCodePanel()
        self.chat_stack.addWidget(self.claude_code_panel)

        splitter.addWidget(self.chat_stack)

        splitter.setSizes([750, 400])
        main.addWidget(splitter)

        # Enhanced status bar
        self.enhanced_status = EnhancedStatusBar()
        self.setStatusBar(None)  # Remove default status bar
        # Add as a widget at the bottom
        central_layout = central.layout()
        if isinstance(central_layout, QHBoxLayout):
            # Wrap in vertical layout to add status bar at bottom
            wrapper = QWidget()
            wrapper_layout = QVBoxLayout(wrapper)
            wrapper_layout.setContentsMargins(0, 0, 0, 0)
            wrapper_layout.setSpacing(0)

            # Move all widgets from main to a horizontal container
            content = QWidget()
            content_layout = QHBoxLayout(content)
            content_layout.setContentsMargins(0, 0, 0, 0)
            content_layout.setSpacing(0)
            content_layout.addWidget(self.activity_bar)
            content_layout.addWidget(self.side_stack)
            content_layout.addWidget(splitter)

            wrapper_layout.addWidget(content)
            wrapper_layout.addWidget(self.enhanced_status)

            self.setCentralWidget(wrapper)

        # Timer to periodically update git branch (every 30 seconds)
        self._git_branch_timer = QTimer(self)
        self._git_branch_timer.timeout.connect(self._update_git_branch)
        self._git_branch_timer.start(30000)  # 30 seconds

    def _setup_menu(self):
        menu = self.menuBar()

        file_menu = menu.addMenu("File")
        file_menu.addAction("New File", self._new_file, QKeySequence.StandardKey.New)
        file_menu.addAction("Open File...", self._open_file, QKeySequence.StandardKey.Open)
        file_menu.addAction("Open Folder...", self._open_folder, QKeySequence("Ctrl+K"))
        file_menu.addSeparator()
        file_menu.addAction("Save", self._save_file, QKeySequence.StandardKey.Save)
        file_menu.addSeparator()
        file_menu.addAction("Quit", self.close, QKeySequence.StandardKey.Quit)

        view_menu = menu.addMenu("View")
        view_menu.addAction("Explorer", lambda: self._on_view_changed("files"))
        view_menu.addAction("Search", lambda: self._on_view_changed("search"))
        view_menu.addAction("AI Agent", lambda: self._on_view_changed("agent"))
        view_menu.addAction("Source Control", lambda: self._on_view_changed("git"))
        view_menu.addAction("Settings", lambda: self._on_view_changed("settings"))
        view_menu.addSeparator()
        view_menu.addAction("Toggle Terminal", self._toggle_terminal, QKeySequence("Ctrl+`"))

        agent_menu = menu.addMenu("Agent")
        agent_menu.addAction("New Chat", self._new_chat)
        agent_menu.addAction("Reconnect", self._reconnect)

    def _setup_shortcuts(self):
        """Setup keyboard shortcuts and command palette."""
        # Command Palette (Ctrl+Shift+P)
        self.command_palette = CommandPalette(self)
        self.command_palette.command_selected.connect(self._execute_command)

        palette_shortcut = QShortcut(QKeySequence("Ctrl+Shift+P"), self)
        palette_shortcut.activated.connect(self._show_command_palette)

        # Quick Open (Ctrl+P)
        self.quick_open = QuickOpenDialog(parent=self)
        self.quick_open.file_selected.connect(self._open_file_from_quick_open)

        quick_open_shortcut = QShortcut(QKeySequence("Ctrl+P"), self)
        quick_open_shortcut.activated.connect(self._show_quick_open)

        # Additional shortcuts
        QShortcut(QKeySequence("Ctrl+Shift+E"), self).activated.connect(
            lambda: self._on_view_changed("files"))
        QShortcut(QKeySequence("Ctrl+Shift+F"), self).activated.connect(
            lambda: self._on_view_changed("search"))
        QShortcut(QKeySequence("Ctrl+Shift+G"), self).activated.connect(
            lambda: self._on_view_changed("git"))
        QShortcut(QKeySequence("Ctrl+,"), self).activated.connect(
            lambda: self._on_view_changed("settings"))

    def _show_command_palette(self):
        """Show the command palette."""
        self.command_palette.show()

    def _show_quick_open(self):
        """Show quick file open dialog."""
        if self.project_dir:
            self.quick_open.set_root(str(self.project_dir))
        self.quick_open.show()

    def _open_file_from_quick_open(self, path: Path):
        """Open a file selected from quick open."""
        self.editor_tabs.open_file(path)

    def _execute_command(self, command: str):
        """Execute a command from the command palette."""
        command_map = {
            "new_file": self._new_file,
            "open_file": self._open_file,
            "open_folder": self._open_folder,
            "save_file": self._save_file,
            "save_all": self._save_all,
            "close_tab": self._close_current_tab,
            "undo": self._undo,
            "redo": self._redo,
            "find": self._show_find,
            "replace": self._show_replace,
            "show_explorer": lambda: self._on_view_changed("files"),
            "show_search": lambda: self._on_view_changed("search"),
            "toggle_terminal": self._toggle_terminal,
            "show_settings": lambda: self._on_view_changed("settings"),
            "show_git": lambda: self._on_view_changed("git"),
            "new_chat": self._new_chat,
            "reconnect_ai": self._reconnect,
            "clear_terminal": self._clear_terminal,
            "restart_terminal": self._restart_terminal,
            "git_commit": self._git_commit,
            "git_push": self._git_push,
            "git_pull": self._git_pull,
            "split_right": self._split_editor_right,
            "split_down": self._split_editor_down,
            "close_split": self._close_split,
            "toggle_blame": self._toggle_blame,
            "fold_all": self._fold_all,
            "unfold_all": self._unfold_all,
        }
        if command in command_map:
            command_map[command]()

    def _save_all(self):
        """Save all open files."""
        for i in range(self.editor_tabs.count()):
            widget = self.editor_tabs.widget(i)
            if hasattr(widget, 'save'):
                widget.save()

    def _close_current_tab(self):
        """Close the current editor tab."""
        current = self.editor_tabs.currentIndex()
        if current >= 0:
            self.editor_tabs.close_tab(current)

    def _undo(self):
        """Undo in current editor."""
        current = self.editor_tabs.currentWidget()
        if hasattr(current, 'undo'):
            current.undo()

    def _redo(self):
        """Redo in current editor."""
        current = self.editor_tabs.currentWidget()
        if hasattr(current, 'redo'):
            current.redo()

    def _show_find(self):
        """Show find dialog in current editor."""
        current = self.editor_tabs.currentWidget()
        if hasattr(current, 'show_find'):
            current.show_find()

    def _show_replace(self):
        """Show find and replace dialog."""
        current = self.editor_tabs.currentWidget()
        if hasattr(current, 'show_replace'):
            current.show_replace()

    def _clear_terminal(self):
        """Clear terminal output."""
        if hasattr(self, 'terminal_panel'):
            self.terminal_panel._clear_terminal()

    def _restart_terminal(self):
        """Restart the terminal shell."""
        if hasattr(self, 'terminal_panel'):
            self.terminal_panel._restart_shell()

    def _git_commit(self):
        """Open git commit dialog."""
        self._on_view_changed("git")

    def _git_push(self):
        """Git push."""
        if hasattr(self, 'terminal_panel'):
            self.terminal_panel.shell.write(b'git push\n')

    def _git_pull(self):
        """Git pull."""
        if hasattr(self, 'terminal_panel'):
            self.terminal_panel.shell.write(b'git pull\n')

    def _split_editor_right(self):
        """Split editor horizontally (side by side)."""
        self.editor_tabs.split_horizontal()

    def _split_editor_down(self):
        """Split editor vertically (top/bottom)."""
        self.editor_tabs.split_vertical()

    def _close_split(self):
        """Close the current split pane."""
        self.editor_tabs.close_split()

    def _on_cursor_position(self, line: int, col: int):
        """Handle cursor position changes from editor."""
        self.enhanced_status.set_cursor_position(line, col)

    def _toggle_blame(self):
        """Toggle git blame view on the current editor."""
        current = self.editor_tabs.currentWidget()
        if hasattr(current, 'toggle_blame'):
            current.toggle_blame()

    def _fold_all(self):
        """Fold all collapsible regions in the current editor."""
        current = self.editor_tabs.currentWidget()
        if hasattr(current, 'fold_all'):
            current.fold_all()

    def _unfold_all(self):
        """Unfold all collapsed regions in the current editor."""
        current = self.editor_tabs.currentWidget()
        if hasattr(current, 'unfold_all'):
            current.unfold_all()

    def _open_folder(self):
        path = QFileDialog.getExistingDirectory(self, "Open Folder", str(Path.home()))
        if path:
            self._open_project(path)

    def _open_project(self, path: str):
        self.project_dir = Path(path).resolve()
        RecentProjects.add(str(self.project_dir))

        self.setWindowTitle(f"Circuit IDE - {self.project_dir.name}")

        # Update panels
        self.file_explorer.set_root(str(self.project_dir))
        self.search_panel.set_root(str(self.project_dir))
        # Ensure git panel is valid before using it
        self._ensure_panel_valid("git")
        try:
            self.git_panel.set_root(str(self.project_dir))
        except RuntimeError:
            pass  # Panel was deleted, will be recreated when opened
        self.terminal_panel.set_working_dir(str(self.project_dir))
        self.claude_code_panel.set_working_dir(str(self.project_dir))

        # Clear welcome, show empty state
        self.editor_tabs.clear()
        self.editor_tabs._open_files.clear()

        # Update git branch in status bar
        self._update_git_branch()

        # Load provider preference and initialize appropriate agent
        try:
            from circuit_agent.config import load_provider_preference
            provider = load_provider_preference()
        except ImportError:
            provider = "cisco"

        if provider == "anthropic":
            # Use Claude Code panel
            self.chat_stack.setCurrentIndex(1)
            self.claude_code_panel.start_claude()
            self.enhanced_status.set_agent_status(True, "Claude Code")
        else:
            # Use Cisco AI agent
            self.chat_stack.setCurrentIndex(0)
            self._init_agent(provider)

        self.enhanced_status.set_message(f"Opened {self.project_dir}")

    def _update_git_branch(self):
        """Update git branch display in status bar."""
        if not self.project_dir:
            self.enhanced_status.set_branch("")
            return

        try:
            result = subprocess.run(
                ["git", "rev-parse", "--abbrev-ref", "HEAD"],
                cwd=str(self.project_dir),
                capture_output=True,
                text=True,
                timeout=2
            )
            if result.returncode == 0:
                branch = result.stdout.strip()
                self.enhanced_status.set_branch(branch)
            else:
                self.enhanced_status.set_branch("")
        except Exception:
            self.enhanced_status.set_branch("")

    def _open_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Open File", str(Path.home()))
        if path:
            self.editor_tabs.open_file(Path(path))

    def _open_file_from_explorer(self, path: Path):
        self.editor_tabs.open_file(path)

    def _new_file(self):
        # Create untitled editor with find/replace support
        editor = EditorWithFindReplace()
        index = self.editor_tabs.addTab(editor, "Untitled")
        self.editor_tabs.setCurrentIndex(index)

    def _save_file(self):
        if widget := self.editor_tabs.currentWidget():
            # Handle both EditorWithFindReplace and CodeEditor
            if isinstance(widget, EditorWithFindReplace):
                if widget.current_file:
                    widget.save_file()
                    self.enhanced_status.set_message(f"Saved {widget.current_file.name}")
            elif isinstance(widget, CodeEditor) and widget.current_file:
                widget.save_file()
                self.enhanced_status.set_message(f"Saved {widget.current_file.name}")

    def _init_agent(self, provider: str = None):
        if not self.project_dir:
            return

        # Load provider preference if not specified
        if provider is None:
            try:
                from circuit_agent.config import load_provider_preference
                provider = load_provider_preference()
            except ImportError:
                provider = "cisco"

        # Clean up any existing worker and thread
        if self.worker:
            try:
                # Disconnect old signals first
                self.send_message_signal.disconnect(self.worker.send_message)
            except (RuntimeError, TypeError):
                pass
            if hasattr(self.worker, 'stop_event_loop'):
                self.worker.stop_event_loop()
            self.worker = None

        if hasattr(self, 'worker_thread') and self.worker_thread:
            try:
                self.worker_thread.quit()
                self.worker_thread.wait(1000)  # Wait up to 1 second
            except (RuntimeError, Exception):
                pass
            self.worker_thread = None

        self.worker_thread = QThread()
        self.worker = AgentWorker()
        self.worker.moveToThread(self.worker_thread)

        # Connect all signals
        self.worker.service_ready.connect(self._on_service_ready)
        self.worker.connected.connect(self._on_connected)
        self.worker.message_chunk.connect(self._on_chunk)
        self.worker.message_complete.connect(self._on_complete)
        self.worker.message_error.connect(self._on_error)
        self.worker.tool_call.connect(self._on_tool_call)
        self.worker.stats_updated.connect(self._on_stats)
        self.worker.confirmation_needed.connect(self._on_confirmation_needed)
        self.send_message_signal.connect(self.worker.send_message)

        self.worker_thread.start()
        self.worker.init_service(str(self.project_dir), provider)

        provider_name = "Claude" if provider == "anthropic" else "Cisco AI"
        self.enhanced_status.set_message(f"Initializing {provider_name} agent...")

    @Slot(bool, str)
    def _on_service_ready(self, success: bool, error: str):
        """Handle service initialization result."""
        if success:
            self.enhanced_status.set_message("Connecting to agent...")
            QTimer.singleShot(100, self.worker.connect_agent)
        else:
            self.cisco_code_panel._append_output(f"Service initialization failed: {error}\n", Theme.ERROR)
            self.enhanced_status.set_message(f"Error: {error[:50]}")
            self.cisco_code_panel.set_connected(False)
            self.agent_panel.set_connected(False, None)

    @Slot(bool, str)
    def _on_connected(self, connected: bool, error: str):
        """Handle agent connection result."""
        self.cisco_code_panel.set_connected(connected)
        project_name = str(self.project_dir) if self.project_dir else None
        self.agent_panel.set_connected(connected, project_name)

        # Get provider info
        provider = self.worker.provider if self.worker else "cisco"
        provider_name = "Claude" if provider == "anthropic" else "Circuit AI"
        model = self.worker.model if self.worker else ""

        # Update enhanced status bar
        self.enhanced_status.set_agent_status(connected, model)

        if connected:
            self.enhanced_status.set_message(f"{provider_name} connected")
            # CiscoCodePanel shows connected message in set_connected()
        else:
            error_msg = error if error else "Connection failed - check credentials in Settings"
            self.enhanced_status.set_message(f"Connection failed: {error_msg[:40]}")
            self.cisco_code_panel._append_output(f"Connection failed: {error_msg}\n", Theme.ERROR)

    @Slot(object)
    def _on_confirmation_needed(self, request):
        """Handle tool confirmation request - show dialog."""
        QTimer.singleShot(0, lambda: self._show_confirmation_dialog(request))

    def _show_confirmation_dialog(self, request):
        """Display confirmation dialog for dangerous operations."""
        from PySide6.QtWidgets import QMessageBox

        tool_name = request.tool_call.name if hasattr(request, 'tool_call') else "Unknown"
        details = request.details if hasattr(request, 'details') else ""
        is_dangerous = request.is_dangerous if hasattr(request, 'is_dangerous') else False

        msg = QMessageBox(self)
        msg.setWindowTitle("Confirm Action")
        msg.setText(f"Allow {tool_name}?")
        msg.setInformativeText(details[:200] if details else "The AI wants to perform this action.")

        if is_dangerous:
            msg.setIcon(QMessageBox.Icon.Warning)
        else:
            msg.setIcon(QMessageBox.Icon.Question)

        msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        msg.setDefaultButton(QMessageBox.StandardButton.No if is_dangerous else QMessageBox.StandardButton.Yes)

        result = msg.exec()
        approved = result == QMessageBox.StandardButton.Yes

        # Send response back to worker
        if self.worker and hasattr(request, 'id'):
            self.worker.respond_to_confirmation(request.id, approved)

    @Slot(str)
    def _on_chunk(self, chunk: str):
        """Handle streaming message chunk from AgentWorker."""
        self.cisco_code_panel.on_chunk(chunk)

    @Slot(str)
    def _on_complete(self, response: str):
        """Handle message completion from AgentWorker."""
        self.cisco_code_panel.on_complete(response)
        self.agent_panel.increment_messages()  # Assistant response
        self.enhanced_status.set_message("Ready")

    @Slot(str)
    def _on_error(self, error: str):
        """Handle error from AgentWorker."""
        self.cisco_code_panel.on_error(error)

    @Slot(str, str, str)
    def _on_tool_call(self, name: str, detail: str, status: str):
        """Handle tool call from AgentWorker."""
        self.cisco_code_panel.on_tool_call(name, detail, status)

    @Slot(int, float)
    def _on_stats(self, tokens: int, cost: float):
        """Handle stats update from AgentWorker."""
        self._total_tokens = tokens
        self._total_cost = cost
        self.cisco_code_panel.on_stats(tokens, cost)
        self.agent_panel.update_stats(tokens, cost)
        self.enhanced_status.set_tokens(tokens, cost)

    def _on_view_changed(self, view: str):
        """Handle activity bar view changes - opens panels as tabs or switches side panel."""
        # Side panel views (file explorer, search)
        if view == "files":
            self.side_stack.setCurrentIndex(0)
            self.side_stack.show()
        elif view == "search":
            self.side_stack.setCurrentIndex(1)
            self.side_stack.show()
        # Tab-based views (agent, git, settings)
        elif view in ("agent", "git", "settings"):
            self._open_panel_as_tab(view)

    def _ensure_panel_valid(self, panel_name: str):
        """Ensure panel is valid, recreating if the C++ object was deleted."""
        try:
            # Try to access a property - will raise RuntimeError if deleted
            if panel_name == "settings":
                _ = self.settings_panel.objectName()
            elif panel_name == "git":
                _ = self.git_panel.objectName()
            elif panel_name == "agent":
                _ = self.agent_panel.objectName()
            return True
        except RuntimeError:
            # C++ object was deleted, recreate the panel
            if panel_name == "settings":
                self.settings_panel = SettingsPanel()
                self.settings_panel.settings_changed.connect(self._on_settings_changed)
                self.settings_panel.provider_changed.connect(self._on_provider_changed)
            elif panel_name == "git":
                self.git_panel = GitPanel()
                self.git_panel.git_operation_completed.connect(self._update_git_branch)
            elif panel_name == "agent":
                self.agent_panel = AgentControlPanel()
                self.agent_panel.reconnect_requested.connect(self._reconnect)
                self.agent_panel.clear_chat_requested.connect(self._new_chat)
            return False

    def _open_panel_as_tab(self, panel_name: str):
        """Open a panel (settings, git, agent) as a tab in the editor area."""
        if panel_name not in ("settings", "git", "agent"):
            return

        # Ensure panel is valid (recreate if deleted)
        self._ensure_panel_valid(panel_name)

        # Tab names and icons
        tab_info = {
            "settings": ("Settings", self.settings_panel, Icons.settings),
            "git": ("Source Control", self.git_panel, Icons.git_branch),
            "agent": ("AI Agent", self.agent_panel, Icons.robot),
        }

        tab_title, panel, icon_method = tab_info[panel_name]

        # Check if tab already open - if so, switch to it
        for i in range(self.editor_tabs.count()):
            tab_widget = self.editor_tabs.widget(i)
            # Check if this is a container with our panel
            if hasattr(tab_widget, 'panel_name') and tab_widget.panel_name == panel_name:
                self.editor_tabs.setCurrentIndex(i)
                if panel_name == "git" and self.project_dir:
                    try:
                        self.git_panel.refresh()
                    except RuntimeError:
                        pass  # Panel was deleted, will be recreated on next open
                return

        # Create a container widget for the panel
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(0)

        # Add header for the panel tab
        header = QWidget()
        header.setFixedHeight(40)
        header.setStyleSheet(f"""
            QWidget {{
                background: {Theme.BG_SECONDARY};
                border-bottom: 1px solid {Theme.BORDER};
            }}
        """)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(16, 0, 16, 0)

        header_title = QLabel(tab_title)
        header_title.setStyleSheet(f"""
            color: {Theme.TEXT_PRIMARY};
            font-size: 14px;
            font-weight: 500;
        """)
        header_layout.addWidget(header_title)
        header_layout.addStretch()

        container_layout.addWidget(header)

        # Add the panel in a scroll area
        scroll = QScrollArea()
        scroll.setWidget(panel)
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(f"""
            QScrollArea {{
                border: none;
                background: {Theme.BG_MAIN};
            }}
            QScrollBar:vertical {{
                background: {Theme.BG_SECONDARY};
                width: 10px;
            }}
            QScrollBar::handle:vertical {{
                background: {Theme.BG_TERTIARY};
                border-radius: 5px;
                min-height: 20px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {Theme.BG_HOVER};
            }}
        """)
        container_layout.addWidget(scroll)

        # Store reference to panel in container for later lookup
        container.panel = panel
        container.panel_name = panel_name

        index = self.editor_tabs.addTab(container, tab_title)
        self.editor_tabs.setTabIcon(index, icon_method(14, Theme.TEXT_SECONDARY))
        self.editor_tabs.setCurrentIndex(index)
        self._panel_tabs[panel_name] = index

        # Refresh git if needed
        if panel_name == "git" and self.project_dir:
            self.git_panel.refresh()

    def _on_tab_changed(self, index: int):
        """Handle tab change - update status bar with file info."""
        # Guard against being called before UI is fully set up
        if not hasattr(self, 'enhanced_status'):
            return

        widget = self.editor_tabs.widget(index)

        # Check if this is a panel tab (settings, git, agent)
        if hasattr(widget, 'panel_name'):
            panel_types = {
                "settings": "Settings",
                "git": "Source Control",
                "agent": "AI Agent",
            }
            self.enhanced_status.set_file_type(panel_types.get(widget.panel_name, "Panel"))
            self.enhanced_status.set_position(0, 0)
            return

        # Determine file type based on extension
        file_type = "Plain Text"
        if isinstance(widget, EditorWithFindReplace) and widget.current_file:
            ext = widget.current_file.suffix.lower()
            file_types = {
                ".py": "Python",
                ".js": "JavaScript",
                ".ts": "TypeScript",
                ".tsx": "TypeScript React",
                ".jsx": "JavaScript React",
                ".html": "HTML",
                ".css": "CSS",
                ".json": "JSON",
                ".md": "Markdown",
                ".yaml": "YAML",
                ".yml": "YAML",
                ".xml": "XML",
                ".sql": "SQL",
                ".sh": "Shell",
                ".bash": "Bash",
                ".rs": "Rust",
                ".go": "Go",
                ".java": "Java",
                ".cpp": "C++",
                ".c": "C",
                ".h": "C Header",
                ".hpp": "C++ Header",
            }
            file_type = file_types.get(ext, "Plain Text")

            # Connect cursor position signal for this editor
            try:
                widget.cursor_position_changed.disconnect(self._on_cursor_position)
            except:
                pass
            widget.cursor_position_changed.connect(self._on_cursor_position)

            # Update position immediately
            cursor = widget.editor.textCursor()
            self._on_cursor_position(cursor.blockNumber() + 1, cursor.columnNumber() + 1)

        self.enhanced_status.set_file_type(file_type)

    def _on_cursor_position(self, line: int, col: int):
        """Update cursor position in status bar."""
        self.enhanced_status.set_position(line, col)

    def _toggle_terminal(self):
        """Toggle terminal panel visibility."""
        if self.terminal_panel.isVisible():
            self.terminal_panel.hide()
        else:
            self.terminal_panel.show()
            self.terminal_panel.input.setFocus()

    def _on_settings_changed(self, settings: dict):
        if self.worker and "model" in settings:
            self.worker.set_model(settings["model"])

        # Handle GitHub MCP settings update
        if self.worker and settings.get("github_mcp_updated"):
            self.worker.reinit_mcp()

    @Slot(str)
    def _on_provider_changed(self, provider: str):
        """Handle AI provider change - swap panels and reinitialize."""
        try:
            provider_name = "Claude Code" if provider == "anthropic" else "Cisco AI"
            self.enhanced_status.set_message(f"Switching to {provider_name}...")

            # Update chat message colors based on provider
            if provider == "anthropic":
                OutputEntry.set_provider("claude")
            else:
                OutputEntry.set_provider("circuit")

            if provider == "anthropic":
                # Switch to Claude Code panel
                self.chat_stack.setCurrentIndex(1)

                # Stop the Cisco agent worker if running
                if self.worker:
                    try:
                        if hasattr(self.worker, 'stop_event_loop'):
                            self.worker.stop_event_loop()
                    except Exception:
                        pass

                # Set working directory and start Claude Code
                if self.project_dir:
                    try:
                        self.claude_code_panel.set_working_dir(str(self.project_dir))
                        self.claude_code_panel.start_claude()
                    except Exception:
                        pass

                # Update status
                try:
                    self.enhanced_status.set_agent_status(True, "Claude Code")
                    self.agent_panel.set_connected(True, str(self.project_dir) if self.project_dir else None)
                except Exception:
                    pass
            else:
                # Switch to Cisco AI chat panel
                self.chat_stack.setCurrentIndex(0)

                # Stop Claude Code if running
                try:
                    self.claude_code_panel.stop_claude()
                except Exception:
                    pass

                # Clear and reinitialize Cisco agent
                try:
                    self.cisco_code_panel.set_connected(False)
                    self.agent_panel.set_connected(False, None)
                    self.agent_panel.reset_session()  # Reset stats for new session
                    self.cisco_code_panel.clear()
                    self.cisco_code_panel._append_output(f"Switching to {provider_name}...\n", Theme.TEXT_MUTED)
                except Exception:
                    pass

                # Reinitialize the Cisco agent
                if self.project_dir:
                    self._init_agent(provider)
        except Exception as e:
            print(f"Error switching provider: {e}")

    def _on_model_changed(self, model: str):
        if self.worker:
            self.worker.set_model(model)
        self.enhanced_status.set_message(f"Model changed to {model}")

    def _on_auto_approve_changed(self, enabled: bool):
        if self.worker:
            self.worker.set_auto_approve(enabled)
        state = "enabled" if enabled else "disabled"
        self.enhanced_status.set_message(f"Auto-approve {state}")

    def _on_cisco_message(self, text: str):
        """Handle message sent from CiscoCodePanel."""
        if not self.project_dir:
            self.cisco_code_panel.on_error("Please open a project folder first to use the AI assistant.\n\nUse File > Open Folder or click 'Open Folder' on the welcome screen.")
            return

        # Check if worker is initialized (either Cisco service or Anthropic client)
        if not self.worker:
            self.cisco_code_panel.on_error("AI agent not initialized. Please open a project folder first.")
            return

        if not self.worker.service and not self.worker.anthropic_client:
            self.cisco_code_panel.on_error("AI agent not connected. Check your credentials in Settings.")
            return

        self.enhanced_status.set_message("Processing...")
        self.agent_panel.increment_messages()  # User message
        self.send_message_signal.emit(text)

    def _new_chat(self):
        """Start a new chat conversation."""
        self.cisco_code_panel.clear()
        self.cisco_code_panel._restart()
        if self.worker:
            # Clear history for Cisco service
            if self.worker.service:
                self.worker.service.clear_history()
            # Clear conversation for Anthropic
            if self.worker.anthropic_client:
                self.worker.clear_conversation()

    def _reconnect(self):
        if not self.project_dir:
            self.enhanced_status.set_message("Open a project folder first")
            return

        if self.worker:
            self.enhanced_status.set_message("Reconnecting...")
            QTimer.singleShot(100, self.worker.connect_agent)
        else:
            # Initialize agent if not done yet
            self._init_agent()

    def closeEvent(self, event):
        # Cleanup all panels with resources
        if hasattr(self, 'claude_code_panel'):
            self.claude_code_panel.stop_claude()
            self.claude_code_panel.cleanup()

        if hasattr(self, 'cisco_code_panel'):
            self.cisco_code_panel.cleanup()

        if hasattr(self, 'terminal_panel'):
            self.terminal_panel.cleanup()

        # Stop the agent worker's event loop
        if self.worker:
            self.worker.stop_event_loop()

        # Stop the Qt thread
        if self.worker_thread:
            self.worker_thread.quit()
            self.worker_thread.wait(2000)  # 2 second timeout

        super().closeEvent(event)


# ============================================================================
# Main
# ============================================================================

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Circuit IDE")
    app.setStyle("Fusion")

    # Set application icon
    icon_path = Path(__file__).parent.parent / "assets" / "icon.png"
    if icon_path.exists():
        app.setWindowIcon(QIcon(str(icon_path)))

    # Dark palette
    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window, QColor(Theme.BG_MAIN))
    palette.setColor(QPalette.ColorRole.WindowText, QColor(Theme.TEXT_PRIMARY))
    palette.setColor(QPalette.ColorRole.Base, QColor(Theme.BG_SECONDARY))
    palette.setColor(QPalette.ColorRole.AlternateBase, QColor(Theme.BG_TERTIARY))
    palette.setColor(QPalette.ColorRole.Text, QColor(Theme.TEXT_PRIMARY))
    palette.setColor(QPalette.ColorRole.Button, QColor(Theme.BG_TERTIARY))
    palette.setColor(QPalette.ColorRole.ButtonText, QColor(Theme.TEXT_PRIMARY))
    palette.setColor(QPalette.ColorRole.Highlight, QColor(Theme.ACCENT_BLUE))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor(Theme.TEXT_PRIMARY))
    app.setPalette(palette)

    window = CircuitIDEWindow()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
